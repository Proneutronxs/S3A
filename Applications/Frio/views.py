from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from S3A.funcionesGenerales import *
from django.db import connections
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import datetime
from Applications.RRHH.views import buscaDatosParaInsertarHE, obtener_fecha_hora_actual_con_milisegundos

# Create your views here.


@login_required
@permission_required('Frio.puede_ingresar', raise_exception=True)
def Frio(request):
    return render (request, 'Frio/frio.html')

@login_required
def HorasExtrasFrio(request):
    return render (request, 'Frio/HorasExtras/horasExtrasFrio.html')

@login_required
def AutorizaHorasExtrasFrio(request):
    return render (request, 'Frio/HorasExtras/autorizaHorasExtras.html')

@login_required
def verHorasExtras(request):
    return render (request, 'Frio/HorasExtras/verHorasExtrasFrio.html')



@csrf_exempt
def cargaCentrosCostosFrio(request):### CAMBIO A CENTRO DE COSTOS MUESTRA LOS CENTROS DE COSTOS CARGADOS
    if request.method == 'GET':
        usuario = str(request.user)
        try:
            data = buscaCentroCostos("CENTROS-FRIO",usuario.upper())
            if data:
                return JsonResponse({'Message': 'Success', 'Datos': data})
            else:
                data = "No se encontraron Centros de Costos Asignados."
                return JsonResponse({'Message': 'Error', 'Nota': data})                
        except Exception as e:
            data = str(e)
            insertar_registro_error_sql("Frio","cargaCCFrio",request.user,data)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
def buscaCentroCostos(codigo,usuario):
    listado =  buscaCentroCostosPorUsuario(codigo,usuario)
    lista_json = []
    for item in listado:
        try:
            with connections['ISISPayroll'].cursor() as cursor:
                sql = "SELECT Regis_CCo, DescrCtroCosto " \
                    "FROM CentrosCostos " \
                    "WHERE Regis_CCo = %s"
                cursor.execute(sql,[item])
                consulta = cursor.fetchone()
                if consulta:
                    ids = str(consulta[0])
                    cc = str(consulta[1])
                    data = {'cc':cc, 'id':ids}
                    lista_json.append(data)
        except Exception as e:
            error = str(e)
            insertar_registro_error_sql("LOGISTICA","BUSCA CENTROS DE COSTO",usuario,error)
            return lista_json
        finally:
            cursor.close()
            connections['ISISPayroll'].close()
    return lista_json

# def buscaCentroCostosFrioHE():
#     data = []
#     try:
#         with connections['TRESASES_APLICATIVO'].cursor() as cursor:
#             sql = "SELECT Texto " \
#                 "FROM Parametros_Aplicativo " \
#                 "WHERE Codigo = 'APP-F-IDCC'"
#             cursor.execute(sql)
#             consulta = cursor.fetchone()
#             if consulta:
#                 data = str(consulta[0]).split('-')
#                 return data
#             else:
#                 return data
#     except Exception as e:
#         error = str(e)
#         insertar_registro_error_sql("Frio","buscaCentroCostosFrio","usuario",error)
#         return data
#     finally:
#         cursor.close()
#         connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def mostrarHorasCargadasPorCCFrio(request): ### MUESTRA LA TABLA DE HORAS
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Frio.puede_ver')
        if user_has_permission:
            cc = request.POST.get('ComboxTipoCCAutorizaFrio')
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT     RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, CONVERT(VARCHAR(10), " \
                                        "HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5), " \
                                        "HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5), HorasExtras_Procesadas.CantidadHoras) AS HORAS, " \
                                        "RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, HorasExtras_Procesadas.ID_HEP AS idHoras, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTROCOSTOS, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) " \
                                        "FROM TresAses_ISISPayroll.dbo.Empleados " \
                                        "WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, HorasExtras_Procesadas.EstadoEnvia " \
                        "FROM        TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                        "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                        "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                        "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                        "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON " \
                                        "TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                        "WHERE     (TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = %s) AND (HorasExtras_Procesadas.EstadoEnvia = '3') " \
                        "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                    cursor.execute(sql, [cc])
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            tipo = str(i[0])
                            legajo = str(i[1])
                            nombres = str(i[2])
                            desde = str(i[3]) + " - " + str(i[4])
                            hasta = str(i[5]) + " - " + str(i[6])
                            motivo = str(i[7])
                            descripcion = str(i[8])
                            horas = str(i[9])
                            idHoras = str(i[11])
                            centro = str(i[12])
                            solicita = str(i[13])
                            datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas , 'centro': centro, 'solicita': solicita}
                            #print(datos)
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("Frio","mostrarHorasCargadasPorCCFrio",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    

def eliminaHorasEstadoHEP(ID_HEP):
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '8' WHERE ID_HEP = %s"
            cursor.execute(sql, [ID_HEP])
            
            slq2 = "UPDATE HorasExtras_Sin_Procesar SET Estado='8' WHERE ID_HESP = (SELECT ID_HESP FROM HorasExtras_Procesadas WHERE ID_HEP = %s)"
            cursor.execute(slq2, [ID_HEP])
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("Frio","eliminaHorasEstadoHEP","usuario",error)
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def eliminaHorasSeleccionadasFrio(request): ### ELIMINA LAS HORAS SELECCIONADAS
    if request.method == 'POST':   ### METODO POST PUT DELETE GET
        user_has_permission = request.user.has_perm('Frio.puede_denegar')
        if user_has_permission:
            checkboxes_tildados = request.POST.getlist('idCheck')
            resultados = []
            for i in checkboxes_tildados:
                ID_HEP = str(i) 
                resultado = eliminaHorasEstadoHEP(ID_HEP)
                resultados.append(resultado)
            if 0 in resultados:
                data = "Se produjo un Error en alguna de las eliminaciones."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se eliminaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'}) 
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def autorizaHorasCargadas(request): ### INSERTA LAS HORAS SELECCIONADAS
    if request.method == 'POST': 
        user_has_permission = request.user.has_perm('Frio.puede_autorizar') 
        if user_has_permission: 
            usuario = str(request.user)
            checkboxes_tildados = request.POST.getlist('idCheck')
            horas = request.POST.getlist('cantHoras')
            tipo = request.POST.getlist('tipoHoraExtra')

           
            resultados = []
            importe = "0"
            pagada = "N"
            index = 0
            for i in checkboxes_tildados:
                ID_HEP = str(i) 
                tipo_hora = str(tipo[index])
                cant_hora = str(horas[index])
                fecha_y_hora = str(obtener_fecha_hora_actual_con_milisegundos())
                Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora = buscaDatosParaInsertarHE(ID_HEP) 
                resultado = insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, cant_hora, IdMotivo, IdAutoriza, Descripcion, tipo_hora, importe, pagada, fecha_y_hora,usuario.upper())
                resultados.append(resultado)
                index = index + 1

            if 0 in resultados:
                data = "Se produjo un Error en alguna de las inserciones."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se autorizaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'}) 
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
def insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user): ### FUNCION QUE INSERTA LAS HORAS EN EL S3A SOLO FUNCION (NO REQUIERE PERMISOS)
    try:
        with connections['S3A'].cursor() as cursor:
            sql = "INSERT INTO RH_HE_Horas_Extras (IdRepl,IdHoraExtra,IdLegajo,FechaDesde,HoraDesde,FechaHasta,HoraHasta,CantHoras,IdMotivo,IdAutoriza,Descripcion,TipoHoraExtra,Valor,Pagada,FechaAlta,UserID) " \
                    "VALUES ('100',(SELECT MAX (IdHoraExtra) + 1 FROM RH_HE_Horas_Extras),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            values = (Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user)
            cursor.execute(sql, values)
            actualizarEstadoHEP(ID_HEP,Thora,Choras)
            return 1
    except Exception as e:
        insertar_registro_error_sql("FRIGORIFICO","insertaHorasExtras","request.user",str(e))
        return 0
    finally:
        connections['S3A'].close()

def actualizarEstadoHEP(ID_HEP, tipo, horas): ### FUNCIÓN QUE ACTUALIZA EL ESTADO SI SE ENVÍO LA HS EXTRA SÓLO FUNCIÓN (NO REQUIERE PERMISOS)
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '0', CantidadHoras = %s, TipoHoraExtra = %s WHERE ID_HEP = %s"
            cursor.execute(sql, [horas,tipo,ID_HEP])
    except Exception as e:
        insertar_registro_error_sql("FRIGORIFICO","actualizarEstadoHEP","request.user",str(e))

# def autorizaHorasEstadoHEP(ID_HEP):
#     try:
#         with connections['TRESASES_APLICATIVO'].cursor() as cursor:
#             sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '1' WHERE ID_HEP = %s"
#             cursor.execute(sql, [ID_HEP])
#             #return True
#     except Exception as e:
#         error = str(e)
#         insertar_registro_error_sql("FRIO","autorizaHorasEstadoHEP","usuario",error)
#         #return False
#     finally:
#         cursor.close()
#         connections['TRESASES_APLICATIVO'].close()


####VER HORAS EXTRAS ESTADO DE LAS HORAS RECHAZADAS Y/O AUTORIZADAS
        
@login_required
@csrf_exempt
def listadoHorasExtrasEstado(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Frio.puede_ver')
        if user_has_permission:
            usuario = str(request.user)
            ###NUEVA
            listado =  buscaCentroCostosPorUsuario("CENTROS-FRIO",usuario)
            numeros_str = [str(numero) for numero in listado]
            numeros_concatenados = ', '.join(numeros_str)
            ###NUEVA
            cc = str(request.POST.get('ComboxCentrosCostos'))
            desde = request.POST.get('fechaBusquedaDesde')
            hasta = request.POST.get('fechaBusquedaHasta')
            estado = str(request.POST.get('ComboxEstadoHora'))
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = f""" 
                            DECLARE @@Desde DATE;
                            DECLARE @@Hasta DATE;
                            DECLARE @@Centro INT;
                            DECLARE @@Estado INT;
                            SET @@Desde = %s;
                            SET @@Hasta = %s;
                            SET @@Centro = %s;
                            SET @@Estado = %s;
                            SELECT        HorasExtras_Procesadas.ID_HEP AS ID, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25),TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRE, 
                                            (SELECT TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto FROM TresAses_ISISPayroll.dbo.CentrosCostos WHERE Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo) AS CC,
                                            (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) + ' Hs.') AS DESDE, (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) + ' Hs.') AS HASTA, 
                                            (SELECT RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) FROM S3A.dbo.RH_HE_Motivo WHERE S3A.dbo.RH_HE_Motivo.IdMotivo= HorasExtras_Procesadas.IdMotivo) AS MOTIVO, 
                                            RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, FORMAT(HorasExtras_Procesadas.CantidadHoras, '0.0') AS CANTIDAD, 
                                            CASE WHEN HorasExtras_Procesadas.ID_HESP IS NULL THEN 'WEB' ELSE 'APP' END AS SECTOR,
                                            CASE WHEN HorasExtras_Procesadas.EstadoEnvia = '8' THEN 'RECHAZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '0' THEN 'AUTORIZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '4' THEN 'PENDIENTE'
                                            WHEN HorasExtras_Procesadas.EstadoEnvia = '3' THEN 'PENDIENTE' ELSE 'DESCONOCIDO' END AS ESTADO, HorasExtras_Procesadas.EstadoEnvia AS ID_ESTADO,
                                            HorasExtras_Procesadas.DescripcionMotivo AS MT
                            FROM            HorasExtras_Procesadas INNER JOIN
                                                    TresAses_ISISPayroll.dbo.Empleados ON HorasExtras_Procesadas.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado
                            WHERE (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraDesde) >= @@Desde OR @@Desde IS NULL OR @@Desde = '')
                                        AND (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraHasta) <= @@Hasta OR @@Hasta IS NULL OR @@Hasta = '')
                                        AND ((@@Estado = '0' AND EstadoEnvia = '0') OR
                                            (@@Estado = '8' AND EstadoEnvia = '8') OR
                                            (@@Estado = '10' AND EstadoEnvia IN (0,8)))
                                        AND (HorasExtras_Procesadas.EstadoEnvia NOT IN (1,4,3))
                                        AND ((TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN ({numeros_concatenados}) AND @@Centro = '0')
                                        OR (TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN (@@Centro) AND @@Centro <> '0'))
                            ORDER BY TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple, HorasExtras_Procesadas.FechaHoraDesde, HorasExtras_Procesadas.FechaHoraHasta

                        """
                    cursor.execute(sql,[desde,hasta,cc,estado])
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            idHora = str(row[0])
                            legajo = str(row[1])
                            nombre = str(row[2])
                            cc = str(row[3])
                            desde = str(row[4])
                            hasta = str(row[5])
                            motivo = str(row[6]) + ' - ' + str(row[12])
                            tipo = str(row[7])
                            horas = str(row[8])
                            dispositivo = str(row[9])
                            estado = str(row[10])
                            id_estado = str(row[11])
                            datos = {'ID':idHora, 'Tipo':tipo, 'Legajo':legajo, 'Nombre':nombre, 'CC':cc, 'Desde':desde,
                                    'Hasta':hasta, 'Motivo': motivo, 'Horas':horas, 'Dispositivo':dispositivo, 'Estado':estado, 'ID_Estado': id_estado}
                            data.append(datos)

                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron horas.'})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("FRIO","LISTA HORAS PROCESADAS",str(request.user),error)
                return JsonResponse({'Message': 'Error', 'Nota': error})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})   
    
@login_required
@csrf_exempt
def creaExcelHorasMostradas(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Frio.puede_ver')
        if user_has_permission:
            usuario = str(request.user)
            ###NUEVA
            listado =  buscaCentroCostosPorUsuario("CENTROS-FRIO",usuario)
            numeros_str = [str(numero) for numero in listado]
            numeros_concatenados = ', '.join(numeros_str)
            ###NUEVA
            cc = str(request.POST.get('ComboxCentrosCostos'))
            desde = request.POST.get('fechaBusquedaDesde')
            hasta = request.POST.get('fechaBusquedaHasta')
            estado = str(request.POST.get('ComboxEstadoHora'))
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = f""" 
                            DECLARE @@Desde DATE;
                            DECLARE @@Hasta DATE;
                            DECLARE @@Centro INT;
                            DECLARE @@Estado INT;
                            SET @@Desde = %s;
                            SET @@Hasta = %s;
                            SET @@Centro = %s;
                            SET @@Estado = %s;
                            SELECT        HorasExtras_Procesadas.ID_HEP AS ID, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25),TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRE, 
                                            (SELECT TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto FROM TresAses_ISISPayroll.dbo.CentrosCostos WHERE Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo) AS CC,
                                            (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) + ' Hs.') AS DESDE, (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) + ' Hs.') AS HASTA, 
                                            HorasExtras_Procesadas.DescripcionMotivo AS MOTIVO, 
                                            RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, FORMAT(HorasExtras_Procesadas.CantidadHoras, '0.0') AS CANTIDAD, 
                                            CASE WHEN HorasExtras_Procesadas.ID_HESP IS NULL THEN 'WEB' ELSE 'APP' END AS SECTOR,
                                            CASE WHEN HorasExtras_Procesadas.EstadoEnvia = '8' THEN 'RECHAZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '0' THEN 'AUTORIZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '4' THEN 'PENDIENTE'
                                            WHEN HorasExtras_Procesadas.EstadoEnvia = '3' THEN 'PENDIENTE' ELSE 'DESCONOCIDO' END AS ESTADO
                            FROM            HorasExtras_Procesadas INNER JOIN
                                                    TresAses_ISISPayroll.dbo.Empleados ON HorasExtras_Procesadas.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado
                            WHERE (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraDesde) >= @@Desde OR @@Desde IS NULL OR @@Desde = '')
                                        AND (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraHasta) <= @@Hasta OR @@Hasta IS NULL OR @@Hasta = '')
                                        AND ((@@Estado = '0' AND EstadoEnvia = '0') OR
                                            (@@Estado = '8' AND EstadoEnvia = '8') OR
                                            (@@Estado = '10' AND EstadoEnvia IN (0,8)))
                                        AND (HorasExtras_Procesadas.EstadoEnvia NOT IN (1,4,3))
                                        AND ((TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN ({numeros_concatenados}) AND @@Centro = '0')
                                        OR (TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN (@@Centro) AND @@Centro <> '0'))
                            ORDER BY TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple, HorasExtras_Procesadas.FechaHoraDesde, HorasExtras_Procesadas.FechaHoraHasta

                        """
                    cursor.execute(sql,[desde,hasta,cc,estado])
                    results = cursor.fetchall()
                    if results:
                        
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = "Horas_Extras"
                        
                        headers = ['ID_HORA', 'LEGAJO', 'NOMBRE', 'CC', 'FECHA HORA INICIO', 'FECHA HORA FIN', 'DESCRIPCIÓN - MOTIVO', 'TIPO', 'CANTIDAD', 'SISTEMA', 'ESTADO']
                        
                        sheet.append(headers)
                        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        for cell in sheet[1]:
                            cell.fill = header_fill

                        border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
                        
                        for row in sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                cell.border = border
                        
                        for row in sheet.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=len(headers)):
                            for cell in row:
                                cell.border = border

                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)

                        sheet.insert_cols(1)

                        for row_number, row_data in enumerate(results, start=8):
                            for col_number, value in enumerate(row_data, start=2):
                                sheet.cell(row=row_number, column=col_number, value=value).alignment = Alignment(horizontal='left', vertical='center')

                        
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 3)
                            sheet.column_dimensions[column].width = adjusted_width

                        for row in sheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        img = Image('static/imagenes/TA.png')
                        img.width = 160 
                        img.height = 80
                        sheet.add_image(img, 'B2')

                        centro = str(NombreCentro(cc))
                        if cc == '0':
                            centro = "TODOS LOS CENTROS"
                        
                        sheet['E2'] = "HORAS EXTRAS - " + centro
                        sheet['E2'].font = Font(bold=True)
                        sheet['H2'] = fecha_hora_actual_texto()
                        sheet['H4'].font = Font(bold=True)
                        sheet['H4'] = "USUARIO: " + usuario.upper()

                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=datos.xlsx'

                        workbook.save(response)

                        return response
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron horas.'})
            except Exception as e:
                error = str(e)
                print(error)
                insertar_registro_error_sql("EMPAQUE","LISTA HORAS PROCESADAS",str(request.user),error)
                return JsonResponse({'Message': 'Error', 'Nota': error})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})   

@login_required
@csrf_exempt
def restauraHora(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Frio.puede_restaurar')
        if user_has_permission:
            usuario = str(request.user)
            Listado_ID_HEP = request.POST.getlist('idCheck')
            index = 0
            for ID_HEP in Listado_ID_HEP:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """ 
                                UPDATE HorasExtras_Procesadas
                                SET EstadoEnvia = CASE
                                                WHEN ID_LTFP IS NULL THEN '3'
                                                ELSE '4'
                                            END
                                WHERE ID_HEP = %s

                            """
                        cursor.execute(sql,[ID_HEP])
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("FRIO","ACTUALIZA ESTADO HORAS",usuario.upper(),error)
                    return JsonResponse ({'Message': 'Not Found', 'Nota': 'Se produjo un error al intentar resolver la petición.'}) 
                index = index + 1

            if len(Listado_ID_HEP) == index:
                return JsonResponse ({'Message': 'Success', 'Nota': 'Los items se restauraron correctamente.'}) 
            else:
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Se produjo un error al intentar restaurar algunos items.'}) 
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})   





























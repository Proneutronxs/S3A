from django.shortcuts import render
from S3A.funcionesGenerales import *
from django.contrib.auth.decorators import login_required # type: ignore
from django.views.decorators.csrf import csrf_exempt # type: ignore
from django.views.static import serve
import os
import datetime 
import openpyxl
from openpyxl.styles import Protection
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from django.db import connections # type: ignore
from django.http import JsonResponse, HttpResponse # type: ignore

def funcionGeneralPermisos(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Empaque.puede_ver')
        if user_has_permission:
            pass
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

# Create your views here.
### RENDERIZADO DE RRHH 
@login_required
def RRHH(request):
    return render (request, 'RRHH/rrhh.html')

@login_required
def Horarios(request):
    return render (request, 'RRHH/Horarios/horarios.html')

@login_required
def agregarHorarios(request):
    return render (request, 'RRHH/Horarios/agregarHorario.html')

@login_required
def Archivos(request):
    return render (request, 'RRHH/Archivos/archivos.html')

### RENDERIZADO DE HORAS EXTRAS
@login_required
def horasExtras(request):
    return render (request, 'RRHH/HorasExtras/horasExtras.html')

@login_required
def horasExtrasSabados(request):
    return render (request, 'RRHH/HorasExtras/heSabados.html')

### RENDERIZADO DE HORAS EXTRAS TRANSFERENCIA
@login_required
def transferenciaHorasExtras(request):
    return render (request, 'RRHH/HorasExtras/transferirhe.html')


@login_required
def Ranking(request):
    return render (request, 'RRHH/RankingFaltas/ranking.html')

@login_required
def RankingFaltas(request):
    return render (request, 'RRHH/RankingFaltas/rankingFaltas.html')

def obtener_fecha_hora_actual_con_milisegundos():
    now = datetime.datetime.now()
    fecha_hora_actual = now.strftime("%Y-%m-%dT%H:%M:%S")
    hora = str(fecha_hora_actual) + ".000"
    return hora

@login_required
@csrf_exempt
def mostrarHorasCargadas(request): ### MUESTRA HORAS PROCESADAS POR SECTOR Y TIPO DE HORA
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            tipo = request.POST.get('ComboxTipoHoraTransf')
            sector =request.POST.get('ComboxSectorHEHoras')
            if tipo == 'N50':
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = "SELECT        RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), " \
                                                    "TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTRO_COSTOS, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE,  " \
                                                    "CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5),  " \
                                                    "HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5),  " \
                                                    "HorasExtras_Procesadas.CantidadHoras) AS HORAS, RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, HorasExtras_Procesadas.ID_HEP AS idHoras, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, ISNULL(ImpArreglo, '0') AS ImpArreglo " \
                            "FROM            TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                                    "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON  " \
                                                    "S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                            "WHERE        (HorasExtras_Procesadas.Sector = 'C') AND (HorasExtras_Procesadas.TipoHoraExtra = 'N') AND (DATEPART(WEEKDAY, HorasExtras_Procesadas.FechaHoraDesde) >= 2 AND DATEPART(WEEKDAY, HorasExtras_Procesadas.FechaHoraDesde) <= 6) " \
                            "AND (HorasExtras_Procesadas.EstadoEnvia = '1') " \
                            "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                        cursor.execute(sql)
                        consulta = cursor.fetchall()
                        if consulta:
                            data = []
                            for i in consulta:
                                tipo = str(i[0])
                                legajo = str(i[1])
                                nombres = str(i[2])
                                centro = str(i[3])
                                desde = str(i[4]) + " - " + str(i[5])
                                hasta = str(i[6]) + " - " + str(i[7])
                                motivo = str(i[8])
                                descripcion = str(i[9])
                                horas = str(i[10])
                                idHoras = str(i[12])
                                solicita = str(i[13])
                                importe = str(i[14])
                                datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'centro': centro, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas, 'importe': importe, 'solicita': solicita}
                                data.append(datos)
                            return JsonResponse({'Message': 'Success', 'Datos': data})
                        else:
                            data = "No se encontrarón horas extras procesadas."
                            return JsonResponse({'Message': 'Error', 'Nota': data})
                except Exception as e:
                    insertar_registro_error_sql("RRHH","mostrarHorasCargadas",request.user,str(e))
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    connections['TRESASES_APLICATIVO'].close()
            if tipo == 'N100':
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = "SELECT        RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), " \
                                                    "TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTRO_COSTOS, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE,  " \
                                                    "CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5),  " \
                                                    "HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5),  " \
                                                    "HorasExtras_Procesadas.CantidadHoras) AS HORAS, RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, HorasExtras_Procesadas.ID_HEP AS idHoras, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, ISNULL(ImpArreglo, '0') AS ImpArreglo  " \
                            "FROM            TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                                    "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON  " \
                                                    "S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                            "WHERE        (HorasExtras_Procesadas.Sector = 'C') AND (HorasExtras_Procesadas.TipoHoraExtra = 'N') AND (DATEPART(WEEKDAY, HorasExtras_Procesadas.FechaHoraDesde) IN (1, 7)) " \
                            "AND (HorasExtras_Procesadas.EstadoEnvia = '1') " \
                            "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                        cursor.execute(sql)
                        consulta = cursor.fetchall()
                        if consulta:
                            data = []
                            for i in consulta:
                                tipo = str(i[0])
                                legajo = str(i[1])
                                nombres = str(i[2])
                                centro = str(i[3])
                                desde = str(i[4]) + " - " + str(i[5])
                                hasta = str(i[6]) + " - " + str(i[7])
                                motivo = str(i[8])
                                descripcion = str(i[9])
                                horas = str(i[10])
                                idHoras = str(i[12])
                                solicita = str(i[13])
                                importe = str(i[14])
                                datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'centro': centro, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas, 'importe': importe, 'solicita': solicita}
                                data.append(datos)
                            return JsonResponse({'Message': 'Success', 'Datos': data})
                        else:
                            data = "No se encontrarón horas extras procesadas."
                            return JsonResponse({'Message': 'Error', 'Nota': data})
                except Exception as e:
                    insertar_registro_error_sql("RRHH","mostrarHorasCargadas",request.user,str(e))
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    connections['TRESASES_APLICATIVO'].close()

            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT        RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), " \
                                                    "TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTRO_COSTOS, " \
                                                    "CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), " \
                                                    "HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, " \
                                                    "RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5), HorasExtras_Procesadas.CantidadHoras) AS HORAS, RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, " \
                                                    "HorasExtras_Procesadas.ID_HEP AS idHoras, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, ISNULL(ImpArreglo, '0') AS ImpArreglo  " \
                            "FROM            TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                                    "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON " \
                                                    "S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                            "WHERE        (HorasExtras_Procesadas.Sector = %s ) AND (HorasExtras_Procesadas.TipoHoraExtra = %s) AND (HorasExtras_Procesadas.EstadoEnvia = '1') " \
                            "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                    cursor.execute(sql, [sector,tipo])
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            tipo = str(i[0])
                            legajo = str(i[1])
                            nombres = str(i[2])
                            centro = str(i[3])
                            desde = str(i[4]) + " - " + str(i[5])
                            hasta = str(i[6]) + " - " + str(i[7])
                            motivo = str(i[8])
                            descripcion = str(i[9])
                            horas = str(i[10])
                            idHoras = str(i[12])
                            solicita = str(i[13])
                            importe = str(i[14])
                            datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'centro': centro, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas, 'importe': importe, 'solicita': solicita}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                insertar_registro_error_sql("RRHH","mostrarHorasCargadas",request.user,str(e))
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def traeTodo(request): ### MUESTRA HORAS PROCESADAS POR SECTOR Y TIPO DE HORA
    if request.method == 'GET':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT        RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), " \
                                                    "TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTRO_COSTOS, " \
                                                    "CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), " \
                                                    "HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, " \
                                                    "RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5), HorasExtras_Procesadas.CantidadHoras) AS HORAS, RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, " \
                                                    "HorasExtras_Procesadas.ID_HEP AS idHoras, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, ISNULL(ImpArreglo, '0') AS ImpArreglo " \
                            "FROM            TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                                    "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON " \
                                                    "S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                            "WHERE        (HorasExtras_Procesadas.Sector = 'C' ) AND (HorasExtras_Procesadas.EstadoEnvia = '1') " \
                            "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                    cursor.execute(sql)
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            tipo = str(i[0])
                            legajo = str(i[1])
                            nombres = str(i[2])
                            centro = str(i[3])
                            desde = str(i[4]) + " - " + str(i[5])
                            hasta = str(i[6]) + " - " + str(i[7])
                            motivo = str(i[8])
                            descripcion = str(i[9])
                            horas = str(i[10])
                            idHoras = str(i[12])
                            solicita = str(i[13])
                            importe = str(i[14])
                            datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'centro': centro, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas, 'importe': importe, 'solicita': solicita}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                insertar_registro_error_sql("RRHH","mostrarHorasCargadas",request.user,str(e))
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required  
@csrf_exempt
def mostrarHorasCargadasPorLegajo(request): ### MUESTRA HORAS CARGADAS Y PROCESADAS POR LEGAJO 
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            legajo = request.POST.get('ComboxTipoLegajoTransf')
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT        RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), " \
                                                    "TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTRO_COSTOS, " \
                                                    "CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), " \
                                                    "HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, " \
                                                    "RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5), HorasExtras_Procesadas.CantidadHoras) AS HORAS, RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, " \
                                                    "HorasExtras_Procesadas.ID_HEP AS idHoras, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, ISNULL(ImpArreglo, '0') AS ImpArreglo " \
                            "FROM            TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                                    "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                                    "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON " \
                                                    "S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                            "WHERE        (HorasExtras_Procesadas.Legajo = %s ) AND (HorasExtras_Procesadas.EstadoEnvia = '1') " \
                            "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                    cursor.execute(sql, [legajo])
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            tipo = str(i[0])
                            legajo = str(i[1])
                            nombres = str(i[2])
                            centro = str(i[3])
                            desde = str(i[4]) + " - " + str(i[5])
                            hasta = str(i[6]) + " - " + str(i[7])
                            motivo = str(i[8])
                            descripcion = str(i[9])
                            horas = str(i[10])
                            idHoras = str(i[12])
                            solicita = str(i[13])
                            importe = str(i[14])
                            datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'centro': centro, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas, 'importe': importe, 'solicita': solicita}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                insertar_registro_error_sql("RRHH","mostrarHorasCargadasPorLegajo",request.user,str(e))
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def enviarHorasCargadas(request): ### INSERTA LAS HORAS SELECCIONADAS EN S3A 
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_insertar')
        if user_has_permission:
            usuario = str(request.user)
            importe = request.POST.get('valor', None)
            pagada = request.POST.get('HePagada', 'N')
            checkboxes_tildados = request.POST.getlist('idCheck')
            resultados = []
            for i in checkboxes_tildados:
                ID_HEP = str(i) 
                fecha_y_hora = str(obtener_fecha_hora_actual_con_milisegundos())
                Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora = buscaDatosParaInsertarHE(ID_HEP) 
                resultado = insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,usuario.upper())
                resultados.append(resultado)
            if 0 in resultados:
                data = "Se produjo un Error en alguna de las inserciones."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se guardaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def cargaLegajos(request): ### MUESTRA LOS LEGAJOS DE LAS HORAS EXTRAS EN LA TABLA POR SECTOR 
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            sector = request.POST.get('ComboxSectorHELegajos')  
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT DISTINCT " \
                                                    "HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), CONVERT(VARCHAR(5), HorasExtras_Procesadas.Legajo) + ' - ' + TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple) AS DATOS " \
                            "FROM            TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                                    "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo " \
                            "WHERE (HorasExtras_Procesadas.EstadoEnvia = '1') AND (HorasExtras_Procesadas.Sector = %s)"
                    cursor.execute(sql, [sector])
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            legajo = str(i[0])
                            nombres = str(i[1])
                            datos = {'legajo':legajo, 'nombres':nombres}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                insertar_registro_error_sql("RRHH","cargaLegajos",request.user,str(e))
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def buscaDatosParaInsertarHE(idHEP): ### SOLO FUNCIÓN, BUSCA DATOS PARA INSERTAR (NO REQUIERE PERMISOS)
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "SELECT Legajo, CONVERT(VARCHAR(10), FechaHoraDesde, 126) AS FECHA_DESDE,CONVERT(VARCHAR(5), FechaHoraDesde, 108) AS HORA_DESDE, " \
                            "CONVERT(VARCHAR(10), FechaHoraHasta, 126) AS FECHA_HASTA, CONVERT(VARCHAR(5), FechaHoraHasta, 108) AS HORA_HASTA, CantidadHoras AS HORAS, " \
                            "IdMotivo AS MOTIVO, Autorizado AS AUTORIZADO, CONVERT(VARCHAR(99), RTRIM(DescripcionMotivo)) AS DESCRIPCION, TipoHoraExtra AS TIPO " \
                    "FROM HorasExtras_Procesadas " \
                    "WHERE ID_HEP = %s" 
            cursor.execute(sql, [idHEP])
            consulta = cursor.fetchone()
            if consulta:
                Legajo = str(consulta[0])
                Fdesde = str(consulta[1])
                Hdesde = str(consulta[2])
                Fhasta = str(consulta[3])
                Hhasta = str(consulta[4])
                Choras = str(consulta[5])
                IdMotivo = str(consulta[6])
                IdAutoriza = str(consulta[7])
                Descripcion = str(consulta[8])
                Thora = str(consulta[9])
                return Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora  
    except Exception as e:
        insertar_registro_error_sql("RRHH","buscaDatosParaInsertarHE","request.user",str(e))
        data = str(e)
        return data
    finally:
        connections['TRESASES_APLICATIVO'].close()

def insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user): ### FUNCION QUE INSERTA LAS HORAS EN EL S3A SOLO FUNCION (NO REQUIERE PERMISOS)
    try:
        with connections['S3A'].cursor() as cursor:
            sql = "INSERT INTO RH_HE_Horas_Extras (IdRepl,IdHoraExtra,IdLegajo,FechaDesde,HoraDesde,FechaHasta,HoraHasta,CantHoras,IdMotivo,IdAutoriza,Descripcion,TipoHoraExtra,Valor,Pagada,FechaAlta,UserID) " \
                    "VALUES ('100',(SELECT MAX (IdHoraExtra) + 1 FROM RH_HE_Horas_Extras),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            values = (Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user)
            cursor.execute(sql, values)
            cursor.close()
            actualizarEstadoHEP(ID_HEP)
            return 1
    except Exception as e:
        insertar_registro_error_sql("RRHH","insertaHorasExtras","request.user",str(e))
        return 0
    finally:
        connections['S3A'].close()

def actualizarEstadoHEP(ID_HEP): ### FUNCIÓN QUE ACTUALIZA EL ESTADO SI SE ENVÍO LA HS EXTRA SÓLO FUNCIÓN (NO REQUIERE PERMISOS)
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '0' WHERE ID_HEP = %s"
            cursor.execute(sql, [ID_HEP])
    except Exception as e:
        insertar_registro_error_sql("RRHH","actualizarEstadoHEP","request.user",str(e))

def eliminaHEP(ID_HEP): ### FUNCIÓN QUE ELIMNA LAS HORAS SELECCIONADS SEGUN EL ID SÓLO FUNCIÓN(NO REQUIERE PERMISOS)
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '8' WHERE ID_HEP = %s"
            cursor.execute(sql, [ID_HEP])

            slq2 = "UPDATE HorasExtras_Sin_Procesar SET Estado='8' WHERE ID_HESP = (SELECT ID_HESP FROM HorasExtras_Procesadas WHERE ID_HEP = %s)"
            cursor.execute(slq2, [ID_HEP])

    except Exception as e:
        insertar_registro_error_sql("RRHH","eliminaHEP","request.user",str(e))
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def eliminaHorasCargadas(request): ### PETICIÓN QUE ELIMINA LAS HORAS SELECCIONADAS (REQUIERE PERMISOS delete)
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_borrar')
        if user_has_permission:
            importe = request.POST.get('valor', None)
            pagada = request.POST.get('HePagada', 'N')
            checkboxes_tildados = request.POST.getlist('idCheck')
            resultados = []
            for i in checkboxes_tildados:
                ID_HEP = str(i)  
                resultado = eliminaHEP(ID_HEP)
                resultados.append(resultado)

            if 0 in resultados:
                data = "Se produjo un Error en alguna de las horas a eliminar."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se eliminaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def mostrarHorasArchivo(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            desde = request.POST.get('Inicio')
            hasta = request.POST.get('Final')
            legajos = request.POST.get('Legajo') or '0'
            values = [desde,hasta,legajos]
            Horas = []
            Nombres = [{'Legajo': '0', 'Nombre': 'TODOS'}]
            try:
                with connections['S3A'].cursor() as cursor:
                    sql = """  
                        DECLARE @@Inicio DATE;
                        DECLARE @@Final DATE;
                        DECLARE @@Legajo INT;
                        SET @@Inicio = %s;
                        SET @@Final = %s;
                        SET @@Legajo = %s;

                        SELECT 
                            RH_HE_Horas_Extras.IdLegajo AS LEGAJO, 
                            CASE 
                                WHEN ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) = 0 
                                THEN '-' 
                                ELSE CONVERT(VARCHAR, ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2)) 
                            END AS HORAS_50,
                            CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = RH_HE_Horas_Extras.IdLegajo) WHEN '10' THEN '534' WHEN '9' THEN '630' WHEN '2' THEN '760'
                            WHEN '4' THEN '825' WHEN '11' THEN '882' ELSE '-' END AS COD_50,
                            CASE 
                                WHEN ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) = 0 
                                THEN '-' 
                                ELSE CONVERT(VARCHAR, ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2)) 
                            END AS HORAS_100,
                            CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = RH_HE_Horas_Extras.IdLegajo) WHEN '10' THEN '536' WHEN '9' THEN '635' WHEN '2' THEN '765'
                            WHEN '4' THEN '820' WHEN '11' THEN '884' ELSE '-' END AS COD_100,
                            CASE 
                                WHEN ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'S' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) = 0 
                                THEN '-' 
                                ELSE CONVERT(VARCHAR, ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'S' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2)) 
                            END AS HORAS_S,
                            CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = RH_HE_Horas_Extras.IdLegajo) WHEN '10' THEN '532' ELSE '-' END AS COD_S,
                            CASE 
                                WHEN ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) = 0 
                                THEN '-' 
                                ELSE CONVERT(VARCHAR, ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2)) 
                            END AS HORAS_AC_50,
                            CASE 
                                WHEN ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) = 0 
                                THEN '-' 
                                ELSE CONVERT(VARCHAR, ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2)) 
                            END AS HORAS_AC_100,
                            (SELECT CASE Sindicatos.DescrSindicato WHEN NULL THEN 'SIN SINDICATO **' ELSE Sindicatos.DescrSindicato END
                                FROM   TresAses_ISISPayroll.dbo.Empleados INNER JOIN
                                            TresAses_ISISPayroll.dbo.Sindicatos ON Empleados.Regis_Sin = Sindicatos.Regis_Sin
                                WHERE (Empleados.CodEmpleado = RH_HE_Horas_Extras.IdLegajo)) AS SINDICATO,
                            (SELECT (RTRIM(RH_Legajo.Apellidos) + ' ' + RTRIM(RH_Legajo.Nombres)) FROM RH_Legajo WHERE RH_Legajo.IdLegajo = RH_HE_Horas_Extras.IdLegajo) AS NOMBRE,
                            (SELECT        CentrosCostos.AbrevCtroCosto
                                    FROM            TresAses_ISISPayroll.dbo.Empleados INNER JOIN
                                                            TresAses_ISISPayroll.dbo.CentrosCostos ON Empleados.Regis_CCo = CentrosCostos.Regis_CCo
                                    WHERE        (Empleados.CodEmpleado = RH_HE_Horas_Extras.IdLegajo)) AS CENTRO,RH_Legajo.Apellidos
                        FROM 
                            RH_HE_Horas_Extras INNER JOIN
                            RH_Legajo ON RH_HE_Horas_Extras.IdLegajo = RH_Legajo.IdLegajo
                        WHERE 
                            CONVERT(DATE, FechaDesde) >= @@Inicio
                            AND CONVERT(DATE, FechaHasta) <= @@Final
                            AND RTRIM(TipoHoraExtra) IN ('50', '100', 'S', 'AC-50', 'AC-100')
                            AND ((SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = RH_HE_Horas_Extras.IdLegajo) NOT IN ('8'))
                            AND (@@Legajo = '0' OR @@Legajo = RH_HE_Horas_Extras.IdLegajo)
                        GROUP BY 
                            RH_HE_Horas_Extras.IdLegajo,RH_Legajo.Apellidos
                        ORDER BY 
                            RH_Legajo.Apellidos;
                    """
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        for row in results:
                            legajo = str(row[0])
                            h_50 = str(row[1])
                            h_100 = str(row[3])
                            h_s = str(row[5])
                            ac_50 = str(row[7])
                            ac_100 = str(row[8])
                            sindicato = str(row[9])
                            nombre = str(row[10])
                            abrev = str(row[11])
                            datos = {'Legajo': legajo, 'Nombre': nombre, 'Abrev': abrev, 
                                     'Horas50': h_50, 'Horas100': h_100, 'HorasS': h_s, 
                                     'AC50': ac_50, 'AC100': ac_100, 'Sindicato':sindicato}
                            dato = {'Legajo': legajo, 'Nombre': legajo + ' - ' + nombre}
                            Horas.append(datos)
                            if dato not in Nombres:
                                Nombres.append(dato)
                        texto = ''
                        if retornaInicioFinalExcel() != '0':
                            inicio, final = retornaInicioFinalExcel()
                            fechaUno = formatear_fecha(inicio)
                            fechaDos = formatear_fecha(final)
                            texto = 'Fecha de Inicio: ' + '01/03/2025' + ', Fecha de Cierre: ' + '28/03/2025' + '.'
                        return JsonResponse({'Message': 'Success', 'Horas': Horas, 'Legajos': Nombres, 'Text':texto})
                    else:
                        texto = ''
                        if retornaInicioFinalExcel() != '0':
                            inicio, final = retornaInicioFinalExcel()
                            fechaUno = formatear_fecha(inicio)
                            fechaDos = formatear_fecha(final)
                            texto = 'Fecha de Inicio: ' + '01/03/2025' + ', Fecha de Cierre: ' + '28/03/2025' + '.'
                        data = "No se encontraron horas extras."
                        return JsonResponse({'Message': 'Error', 'Nota': data, 'Text':texto})

            except Exception as e:
                insertar_registro_error_sql("RRHH","mostrarHorasArchivo",str(request.user).upper(),str(e))
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Hubo un error al intentar resolver la petición. ' + str(e) })
            finally:
                cursor.close()
                connections['S3A'].close()
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def retornaInicioFinalExcel():
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """
                SELECT 
                    CONVERT(VARCHAR, YEAR(GETDATE())) + '-' + RIGHT('0' + CONVERT(VARCHAR, MONTH(GETDATE())), 2) + '-' + FORMAT(InicioINT, '00') AS INICIO,
                    CONVERT(VARCHAR, YEAR(GETDATE())) + '-' + RIGHT('0' + CONVERT(VARCHAR, MONTH(GETDATE())), 2) + '-' +FORMAT(FinalINT, '00') AS FINAL
                FROM 
                    Parametros_Aplicativo
                WHERE 
                    Codigo = 'EXCEL-ISIS-HE';
            """
            cursor.execute(sql)
            results = cursor.fetchone()
            if results:
                inicio = str(results[0])
                final = str(results[1])
                return inicio,final
            else:
                return '0'
    except Exception as e:
        insertar_registro_error_sql("RRHH","RETORNA INICIO FINAL","request.user",str(e))
        return '0'
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

def calcular_porcentajes(numero): 
    diez_por_ciento = round(numero * 0.10, 2)
    noventa_por_ciento = round(numero * 0.90, 2)
    return diez_por_ciento, noventa_por_ciento

def traeHorasExtras(): ### COLUMNA 0=LEGAJO
    try:
        with connections['S3A'].cursor() as cursor:
            sql = """
                DECLARE @@Inicio DATE;
                DECLARE @@Final DATE;
                SET @@Inicio = '2025-03-01';
                SET @@Final = '2025-03-28';
                SELECT 
                    IdLegajo AS LEGAJO, 
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_50,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '534' WHEN '9' THEN '630' WHEN '2' THEN '760'
                    WHEN '4' THEN '825' WHEN '11' THEN '882' ELSE '-' END AS COD_50,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_100,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '536' WHEN '9' THEN '635' WHEN '2' THEN '765'
                    WHEN '4' THEN '820' WHEN '11' THEN '884' ELSE '-' END AS COD_100,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'S' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_S,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '532' ELSE '-' END AS COD_S,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_AC_50,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_AC_100,
                    (SELECT CASE Sindicatos.DescrSindicato WHEN NULL THEN 'SIN SINDICATO **' ELSE Sindicatos.DescrSindicato END
                        FROM   TresAses_ISISPayroll.dbo.Empleados INNER JOIN
                                    TresAses_ISISPayroll.dbo.Sindicatos ON Empleados.Regis_Sin = Sindicatos.Regis_Sin
                        WHERE (Empleados.CodEmpleado = IdLegajo)) AS SINDICATO,
                    (SELECT (RTRIM(RH_Legajo.Apellidos) + ' ' + RTRIM(RH_Legajo.Nombres)) FROM RH_Legajo WHERE RH_Legajo.IdLegajo = RH_HE_Horas_Extras.IdLegajo) AS NOMBRE
                FROM 
                    RH_HE_Horas_Extras
                WHERE 
                    CONVERT(DATE, FechaDesde) >= @@Inicio
                    AND CONVERT(DATE, FechaHasta) <= @@Final
                    AND RTRIM(TipoHoraExtra) IN ('50', '100', 'S', 'AC-50', 'AC-100')
                    AND ((SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) NOT IN ('8'))
                    --AND IdLegajo IN ('59510','2278', '47087','52026','54965')
                GROUP BY 
                    IdLegajo
                ORDER BY 
                    IdLegajo;
            """
            inicio,final = retornaInicioFinalExcel()
            cursor.execute(sql)#,[inicio,final])
            results = cursor.fetchall()
            data = []
            if results:
                for row in results:
                    Legajo = str(row[0])
                    Sindicato = str(row[9])
                    Hora_50 = ''
                    Concepto_50 = ''
                    Hora_100 = ''
                    Concepto_100 = ''
                    Acuerdo_50 = str(row[7])
                    Acuerdo_100 = str(row[8])
                    Hora_50 = row[1]
                    Concepto_50 = str(row[2])
                    Hora_100 = row[3]
                    Concepto_100 = str(row[4])
                    Float_50 = float(row[7])
                    Float_100 = float(row[8])

                    if Legajo == '9':
                        Concepto_50 = '825'
                        Concepto_100 = '820'
                    if Legajo == '234':
                        Concepto_50 = '825'
                        Concepto_100 = '862'
                    if Legajo == '53979':
                        Concepto_50 = '825'
                        Concepto_100 = '862'
                    if Legajo == '56095':
                        Concepto_50 = '760'
                        Concepto_100 = '765'
                    if Legajo == '238':
                        Concepto_50 = '760'
                        Concepto_100 = '765'

                    total_horas_50 = 0.0
                    total_horas_100 = 0.0

                    if Sindicato == 'FRIGORIFICO':
                        if Acuerdo_100 != '0.0' or Acuerdo_50 != '0.0':
                            if Float_50 >= 6:
                                hora_10_50, hora_90_50 = calcular_porcentajes(Float_50)
                                total_horas_50 = round(hora_10_50 + float(Hora_50), 2)
                                datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_50), "CONCEPTO": Concepto_50}
                                data.append(datos)
                            else:
                                if Acuerdo_50 != '0.0':
                                    total_horas_50 = round(float(Acuerdo_50) + float(Hora_50), 2)
                                    datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_50).replace(',','.'), "CONCEPTO": Concepto_50}
                                    data.append(datos)
                            if Float_100 >= 6:
                                hora_10_100, hora_90_100 = calcular_porcentajes(Float_100)
                                total_horas_100 = round(hora_10_100 + float(Hora_100), 2)
                                datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_100), "CONCEPTO": Concepto_100}
                                data.append(datos)
                            else:
                                if Acuerdo_100 != '0.0':
                                    total_horas_100 = round(float(Acuerdo_100) + float(Hora_100), 2)
                                    datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_100).replace(',','.'), "CONCEPTO": Concepto_100}
                                    data.append(datos)
                            
                        else:
                            if str(Hora_50) != '0.0':
                                datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_50, 2)), "CONCEPTO": Concepto_50}
                                data.append(datos)
                            if str(Hora_100) != '0.0':
                                datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_100, 2)), "CONCEPTO": Concepto_100}
                                data.append(datos)
                    else:
                        if str(Hora_50) != '0.0':
                            datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_50, 2)), "CONCEPTO": Concepto_50}
                            data.append(datos)
                        if str(Hora_100) != '0.0':
                            datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_100, 2)), "CONCEPTO": Concepto_100}
                            data.append(datos)   
                return {"datos": data}
            else:
                return {"datos": []}
    except Exception as e:
        insertar_registro_error_sql("RRHH","TRAE HORAS EXTRAS","request.user",str(e))
        return '0'
    finally:
        cursor.close()
        connections['S3A'].close()
   
@login_required  
@csrf_exempt
def CreaExcelISIS(request): 
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('RRHH.puede_ver')
        if user_has_permission:
            data_dict = traeHorasExtras() 
            if data_dict != '0':
                file_name='horas_extras_isis.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active

                for idx, entry in enumerate(data_dict['datos'], start=1):
                    ws[f'A{idx}'] = entry['LEGAJO']
                    ws[f'B{idx}'] = entry['CONCEPTO']
                    ws[f'C{idx}'] = entry['HORAS']
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename=output.xlsx'
                return response
            else:
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'No se pudo completar la petición.'})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

# def generate_excel_from_dict():
#     data_dict = traeHorasExtras() 
#     file_name='output.xlsx'
#     wb = openpyxl.Workbook()
#     ws = wb.active

#     for idx, entry in enumerate(data_dict['datos'], start=1):
#         ws[f'A{idx}'] = entry['LEGAJO']
#         ws[f'B{idx}'] = entry['CONCEPTO']
#         ws[f'C{idx}'] = entry['HORAS']

#     # Save the workbook to a file
#     wb.save(file_name)
#     print(f"Excel file '{file_name}' generated successfully.")

# # Example usage


# generate_excel_from_dict(traeHorasExtras(), 'output.xlsx').

def carga_combox_sabados(request):
    if request.method == 'GET':
        try:
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ 
                        SELECT DISTINCT CONVERT(DATE,PCHE.Fecha) AS FECHA, 'SÁBADO: ' + CONVERT(VARCHAR(10), PCHE.Fecha, 103) AS DIA
                        FROM Pre_Carga_Horas_Extras AS PCHE
                        WHERE PCHE. Estado = 'P'
                            AND (DATENAME(WEEKDAY, CONVERT(DATE, Fecha)) = 'Sábado' OR DATENAME(WEEKDAY, CONVERT(DATE, Fecha)) = 'Saturday')
                            AND (DATEPART(YEAR, Fecha) = DATEPART(YEAR, GETDATE()))
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = []
                    for row in consulta:
                        idValue = str(row[0])
                        value = str(row[1])
                        lista_data.append({'IdValue': idValue, 'Value':value})
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
        return JsonResponse({'Message': 'Success', 'Reportes': reportes, "Empaques":empaques})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    

@csrf_exempt    
def data_listado_sabados(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idFecha = str(request.POST.get('IdFecha'))
            values = [idFecha,idFecha]

            with connections['principal'].cursor() as cursor:
                sql = """ 
                        SET DATEFORMAT ymd;
                        DECLARE @@FechaDesde DATETIME;
                        DECLARE @@FechaHasta DATETIME;
                        SET @@FechaDesde = %s;
                        SET @@FechaHasta = CONVERT(VARCHAR(10), DATEADD(DAY, +1, %s), 120);

                        SELECT TL.legLegajo AS LEGAJO, TL.legNombre AS NOMBRE,
                            CONVERT(VARCHAR, (FORMAT(TF.ficFecha, 'yyyy-MM-dd'))) + ' ' + CONVERT(VARCHAR, (FORMAT(TF.ficHora, 'HH:mm:ss'))) AS DATE_TIME,
                            CONVERT(VARCHAR, (FORMAT(TF.ficHora, 'HH:mm'))) AS TIME_DATE, CONVERT(VARCHAR(10), TF.ficFecha,103) AS FECHA
                        FROM T_Fichadas AS TF LEFT JOIN 
                                T_Legajos AS TL ON TL.legCodigo = TF.legCodigo
                        WHERE TL.legLegajo IN (SELECT Legajo
                                                FROM TRESASES_APLICATIVO.dbo.Pre_Carga_Horas_Extras
                                                WHERE TRY_CONVERT(DATE, Fecha) = @@FechaDesde)
                            AND ISNUMERIC(TL.legLegajo) = 1
                            AND TF.ficFecha + TF.ficHora >= @@FechaDesde + ' 05:00:00'
                            AND TF.ficFecha + TF.ficHora <= @@FechaHasta + ' 05:00:00'
                        ORDER BY TL.legLegajo, TF.ficFecha + TF.ficHora 
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = []
                    for row in consulta:
                        legajo = str(row[0])
                        nombre = str(row[1])
                        time = str(row[3])
                        fecha = 'SÁBADO - ' + str(row[4])
                        
                        encontrado = False
                        for item in lista_data:
                            if item["Legajo"] == legajo:
                                item["Fichadas"].append(time)  
                                encontrado = True
                                break
                            
                        if not encontrado:
                            nuevo_registro = {
                                "Legajo": legajo,
                                "Nombre": nombre,
                                "Dia": fecha,
                                "Fichadas": [time]  
                            }

                        
                            lista_data.append(nuevo_registro)

                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = 'Legajo'
                    ws['B1'] = 'Nombre'
                    ws['C1'] = ''
                    for i in range(6):
                        ws[f'D{i+1}'] = f' {i+1}'
                    ws['J1'] = '50'
                    ws['K1'] = '100'

                    for i, item in enumerate(lista_data, start=2):
                        ws[f'A{i}'] = item['Legajo']
                        ws[f'B{i}'] = item['Nombre']
                        ws[f'C{i}'] = item['Dia']
                        for j, fichada in enumerate(item['Fichadas']):
                            if j < 6:
                                ws.cell(row=i, column=3+j).value = fichada
                        ws[f'J{i}'] = ''
                        ws[f'K{i}'] = ''

                    wb.save('Applications/RRHH/Archivos/Excel/Archivo_Horas_Sabados.xlsx')
                    #print(lista_data)
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def descarga_excel_creados(request, filename):
    nombre = filename
    filename = 'Applications/RRHH/Archivos/Excel/' + filename
    if os.path.exists(filename):
        response = serve(request, os.path.basename(filename), os.path.dirname(filename))
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        return response
    else:
        raise Http404
    
@csrf_exempt 
def recibir_archivo_excel(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ 
                    DECLARE @Legajo VARCHAR(10);
                    DECLARE @Desde DATETIME;
                    DECLARE @Hasta DATETIME;
                    DECLARE @Tipo VARCHAR(10);
                    DECLARE @Cantidad VARCHAR(10);

                    SET @Legajo = %s;
                    SET @Desde = %s;
                    SET @Hasta = %s;
                    SET @Tipo = %s;
                    SET @Cantidad = %s;

                    IF EXISTS (
                        SELECT 1 
                        FROM HorasExtras_Procesadas 
                        WHERE Legajo = @Legajo AND FechaHoraDesde = @Desde AND TipoHoraExtra = @Tipo
                    )
                    BEGIN
                        UPDATE HorasExtras_Procesadas 
                        SET 
                            FechaHoraHasta = @Hasta, 
                            IdMotivo = '17', 
                            DescripcionMotivo = 'CALC-RRHH', 
                            Autorizado = '100', 
                            UsuarioEncargado = '0', 
                            CantidadHoras = @Cantidad, 
                            ImpArreglo = '0', 
                            Sector = '', 
                            ID_LTFP = '1001', 
                            EstadoEnvia = '4'
                        WHERE Legajo = @Legajo AND FechaHoraDesde = @Desde AND TipoHoraExtra = @Tipo
                    END
                    ELSE
                    BEGIN
                        INSERT INTO HorasExtras_Procesadas 
                        (
                            Legajo, 
                            FechaHoraDesde, 
                            FechaHoraHasta, 
                            IdMotivo, 
                            DescripcionMotivo, 
                            Autorizado, 
                            UsuarioEncargado, 
                            TipoHoraExtra, 
                            CantidadHoras, 
                            ImpArreglo, 
                            Sector, 
                            ID_LTFP, 
                            EstadoEnvia
                        ) 
                        VALUES 
                        (
                            @Legajo, 
                            @Desde, 
                            @Hasta, 
                            '17', 
                            'CALC-RRHH', 
                            '100', 
                            '0', 
                            @Tipo, 
                            @Cantidad, 
                            '0', 
                            '', 
                            '1001', 
                            '4'
                        )
                    END
                """
                # cursor.execute(sql)

                idFecha = str(request.POST.get('IdFecha'))
                archivo_excel = request.FILES['archivoExcel']
                wb = load_workbook(archivo_excel)
                ws = wb.active

                valido = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        int(row[0])
                        valido = True
                    except ValueError:
                        raise ValueError(f"La columna A contiene un valor no entero: {row[0]} ({type(row[0]).__name__})")
                    break
                if not valido:
                    raise ValueError("La columna A no contiene valores enteros")
                if not valido:
                    raise ValueError("La columna A no contiene valores enteros")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[9] is not None and not isinstance(row[9], datetime.time):
                        raise ValueError("La columna 50 no tiene datos de tipo tiempo HH:mm")
                    if row[10] is not None and not isinstance(row[10], datetime.time):
                        raise ValueError("La columna 100 no tiene datos de tipo tiempo HH:mm")
                

                for row in ws.iter_rows(min_row=2, values_only=True):
                    legajo = str(row[0])
                    nombre = str(row[1])
                    f1 = row[2]
                    f2 = row[3]
                    f3 = row[4]
                    f4 = row[5]
                    f5 = row[6]
                    f6 = row[7]
                    _50 = '0' if row[9] is None or row[9] in ['0:00', '00:00', '00:00:00'] else str(row[9])
                    _100 = '0' if row[10] is None or row[10] in ['0:00', '00:00', '00:00:00'] else str(row[10])
                    values_50 = [legajo,idFecha,idFecha,'50',str(tiempo_a_decimal(_50))]
                    values_100 = [legajo,idFecha,idFecha,'100',str(tiempo_a_decimal(_100))]
                    if _50 == '0' and _100 == '0':
                        cursor.execute(sql,values_50)
                    if _50 != '0':
                        cursor.execute(sql,values_50)
                    if _100 != '0':
                        cursor.execute(sql,values_100)

                sql_update = """ UPDATE Pre_Carga_Horas_Extras SET Estado = 'C' WHERE CONVERT(DATE,Fecha) = %s """
                cursor.execute(sql_update,[idFecha])
            return JsonResponse({'Message': 'Success', 'Nota': 'La horas se cargaron correctamente.'}) 
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
def tiempo_a_decimal(tiempo):
    if tiempo == '0' or tiempo == '':
        return '0'
    partes = tiempo.split(':')
    if len(partes) == 2:  
        horas, minutos = partes
        segundos = 0
    elif len(partes) == 3:
        horas, minutos, segundos = partes
    else:
        raise ValueError("Formato de tiempo inválido")
    
    horas = int(horas)
    minutos = int(minutos)
    segundos = int(segundos)
    decimal_horas = horas
    decimal_minutos = 0
    if minutos >= 45:
        decimal_minutos = 1
    elif minutos >= 15:
        decimal_minutos = 0.5
    return str(decimal_horas + decimal_minutos)

@csrf_exempt    
def recibe_data_listado_sabados(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            fecha = str(request.POST.get('IdFecha'))
            legajos = request.POST.getlist('Legajo[]')
            time_input_50 = request.POST.getlist('timeInput50[]')
            time_input_100 = request.POST.getlist('timeInput100[]')
            
            for i in range(len(legajos)):
                legajo = legajos[i]
                _50 = '00:00' if time_input_50[i] == '' else time_input_50[i]
                _100 = '00:00' if time_input_100[i] == '' else time_input_100[i]
                decimal_50 = horas_a_decimales(_50)
                decimal_100 = horas_a_decimales(_100)
                if decimal_100 != '0.00':
                    inserta_horas_sabado(legajo,fecha,'100',decimal_100)
                    actualiza_pre_carga(fecha,legajo)
                if decimal_50 != '0.0':
                    inserta_horas_sabado(legajo,fecha,'50',decimal_50)
                    actualiza_pre_carga(fecha,legajo)
            return JsonResponse({'Message': 'Success', 'Nota': 'Las horas se guardaron correctamente.'})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def inserta_horas_sabado(legajo,fecha,tipo,cantidad):
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """   

                DECLARE @@Legajo INT;
                DECLARE @@Fecha DATE;
                DECLARE @@TipoHora VARCHAR(10);
                DECLARE @@CantidadHora VARCHAR(10);

                SET @@Legajo = %s;
                SET @@Fecha = %s;
                SET @@TipoHora = %s;
                SET @@CantidadHora = %s;

                INSERT INTO HorasExtras_Procesadas (Legajo,FechaHoraDesde,FechaHoraHasta, IdMotivo, DescripcionMotivo, Autorizado, UsuarioEncargado, TipoHoraExtra, CantidadHoras,EstadoEnvia)
			            VALUES(@@Legajo, @@Fecha, @@Fecha, '17', 'REVISADO POR RRHH', '100','100', @@TipoHora, @@CantidadHora,'4')

                """
            cursor.execute(sql, [legajo,fecha,tipo,cantidad])    
            return True
    except Exception as e:
        return False
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

def actualiza_pre_carga(fecha,legajo):
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """   

                UPDATE Pre_Carga_Horas_Extras SET Estado = 'C' WHERE Fecha = %s AND Legajo = %s

                """
            cursor.execute(sql, [fecha,legajo])    
            return True
    except Exception as e:
        return False
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

def horas_a_decimales(hora):
    horas, minutos = hora.split(":")
    horas = int(horas)
    minutos = int(minutos)
    if minutos > 45:
        horas += 1
        minutos = 0
    elif minutos > 15:
        minutos = 30
    else:
        minutos = 0
    decimales = horas + minutos / 60
    return str(round(decimales, 2))


"""


SET DATEFORMAT ymd;
DECLARE @@FechaDesde DATETIME;
DECLARE @@FechaHasta DATETIME;
SET @@FechaDesde = '2025-02-08';
SET @@FechaHasta = CONVERT(VARCHAR(10), DATEADD(DAY, +1, '2025-02-08'), 120);
SELECT TL.legLegajo AS LEGAJO, TL.legNombre AS NOMBRE,
    CONVERT(VARCHAR, (FORMAT(TF.ficFecha, 'yyyy-MM-dd'))) + ' ' + CONVERT(VARCHAR, (FORMAT(TF.ficHora, 'HH:mm:ss'))) AS DATE_TIME,
	CONVERT(VARCHAR, (FORMAT(TF.ficHora, 'HH:mm'))) AS TIME_DATE
FROM T_Fichadas AS TF LEFT JOIN 
		T_Legajos AS TL ON TL.legCodigo = TF.legCodigo
WHERE TL.legLegajo IN (SELECT Legajo
						FROM TRESASES_APLICATIVO.dbo.Pre_Carga_Horas_Extras
						WHERE TRY_CONVERT(DATE, Fecha) = @@FechaDesde)
	AND ISNUMERIC(TL.legLegajo) = 1
    AND TF.ficFecha + TF.ficHora >= @@FechaDesde + ' 05:00:00'
    AND TF.ficFecha + TF.ficHora <= @@FechaHasta + ' 05:00:00'
ORDER BY TL.legLegajo, TF.ficFecha + TF.ficHora 

"""


"""
SELECT DISTINCT CONVERT(DATE,PCHE.Fecha) AS FECHA, 'SÁBADO: ' + CONVERT(VARCHAR(10), PCHE.Fecha, 103) AS DIA
FROM Pre_Carga_Horas_Extras AS PCHE
WHERE PCHE. Estado = 'P'
	AND (DATENAME(WEEKDAY, CONVERT(DATE, Fecha)) = 'Sábado' OR DATENAME(WEEKDAY, CONVERT(DATE, Fecha)) = 'Saturday')
	AND (DATEPART(YEAR, Fecha) = DATEPART(YEAR, GETDATE()))

"""
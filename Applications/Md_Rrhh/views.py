from S3A.funcionesGenerales import obtenerHorasArchivo, registroRealizado, formatear_moneda
from openpyxl.styles import PatternFill, Font, Border, Side
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from S3A.funcionesGenerales import obtenerHorasArchivo, registroRealizado
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseForbidden
from openpyxl.drawing.image import Image
from django.views.static import serve
from openpyxl.styles import Alignment
from django.shortcuts import render
from openpyxl import load_workbook
from django.db import connections
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
import openpyxl
import json
import os


# Create your views here.

@login_required
def Md_Rrhh(request):
    user_has_permission = request.user.has_perm('Md_Rrhh.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Rrhh/index.html')
    return render (request, 'Md_Rrhh/404.html')

@login_required
def Listado_Labores(request):
    user_has_permission = request.user.has_perm('Md_Rrhh.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Rrhh/Chacras/listadoLabores.html')
    return render (request, 'Md_Rrhh/404.html')

@login_required
def Listado_Sipreta(request):
    user_has_permission = request.user.has_perm('Md_Rrhh.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Rrhh/Chacras/listadoSipreta.html')
    return render (request, 'Md_Rrhh/404.html')

@login_required
def Archivo_Isis(request):
    user_has_permission = request.user.has_perm('Md_Rrhh.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Rrhh/HorasExtras/archivo_isis.html')
    return render (request, 'Md_Rrhh/404.html')







def listaCentrosChacras(request):
    if request.method == 'GET':
        try:
            with connections['ISISPayroll'].cursor() as cursor:
                sql = """
                        SELECT Regis_Cco AS ID, UPPER(CONVERT(VARCHAR(45),(AbrevCtroCosto + ' - ' + DescrCtroCosto))) AS CENTRO
                        FROM CentrosCostos
                        WHERE AbrevCtroCosto LIKE ('C%')
                        ORDER BY CENTRO
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    lista_tarea = [{'Codigo': '', 'Descripcion': 'TODO'},{'Codigo': 'P', 'Descripcion': 'PODA'},{'Codigo': 'R', 'Descripcion': 'RALEO'}]
                    lista_data = [{'Codigo': '', 'Descripcion': 'TODOS - (NO ROMMIK)'},{'Codigo': '18', 'Descripcion': 'RMK - ROMMIK'}]
                    for row in consulta:
                        codigo = str(row[0])
                        descripcion = str(row[1])
                        datos = {'Codigo': codigo, 'Descripcion': descripcion}
                        lista_data.append(datos)
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data, 'Tareas':lista_tarea })
                else:
                    return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})
        except Exception as e:
            error = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': error})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def descarga_archivo_excel(request, filename):
    nombre = filename
    filename = 'Applications/Md_Rrhh/Archivos/Excel/' + filename
    if os.path.exists(filename):
        response = serve(request, os.path.basename(filename), os.path.dirname(filename))
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        return response
    else:
        raise 

def listaCentros(request):
    if request.method == 'GET':
        try:
            with connections['ISISPayroll'].cursor() as cursor:
                sql = """
                        SELECT Regis_Cco AS ID, UPPER(CONVERT(VARCHAR(45),(AbrevCtroCosto + ' - ' + DescrCtroCosto))) AS CENTRO
                        FROM CentrosCostos
                        WHERE AbrevCtroCosto NOT LIKE ('C%')
                        ORDER BY CENTRO
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [{'Codigo': '', 'Descripcion': 'TODOS'}]
                    for row in consulta:
                        codigo = str(row[0])
                        descripcion = str(row[1])
                        datos = {'Codigo': codigo, 'Descripcion': descripcion}
                        lista_data.append(datos)
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})
        except Exception as e:
            error = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': error})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt    
def data_legajos_x_centro(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Md_Rrhh.puede_ver')
        if user_has_permission:
            try:
                centro = str(request.POST.get('Centro'))
                values = [centro]
                lista_data = []
                with connections['ISISPayroll'].cursor() as cursor:
                    sql = """
                            SELECT CodEmpleado AS LEGAJO, CONVERT(VARCHAR,CodEmpleado) + ' - ' + ApellidoEmple + ' ' + NombresEmple AS NOMBRES
                            FROM Empleados
                            WHERE Regis_Cco = %s  AND BajaDefinitivaEmple = 2
                            ORDER BY ApellidoEmple + ' ' + NombresEmple
                        """
                    cursor.execute(sql,values)
                    consulta = cursor.fetchall()
                    if consulta:
                        lista_data = []
                        for row in consulta:
                            codigo = str(row[0])
                            descripcion = str(row[1])
                            datos = {'Codigo': codigo, 'Descripcion': descripcion}
                            lista_data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})
            except Exception as e:
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Error', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt    
def data_listado_horas_extras_isis(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Md_Rrhh.puede_ver')
        if user_has_permission:
            try:
                inicio = str(request.POST.get('Incio'))
                final = str(request.POST.get('Final'))
                centro = str(request.POST.get('Centro'))
                legajo = str(request.POST.get('Legajo'))
                archivo = str(request.POST.get('Archivo'))
                values = [inicio,final,centro,legajo]
                values_dos = [inicio,final]
                if archivo == "N":
                    lista_data = jsonArchivoIsis(values)
                    if lista_data:
                        return JsonResponse({'Message': 'Success', 'Datos': lista_data})  
                    else:
                        data = 'No se encontraron datos.'
                        return JsonResponse({'Message': 'Error', 'Nota': data})
                else:
                    excel_response = response_excel(values_dos)            
                    return excel_response 
            except Exception as e:
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Error', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def calcular_porcentajes(numero): 
    diez_por_ciento = round(numero * 0.10, 2)
    noventa_por_ciento = round(numero * 0.90, 2)
    return diez_por_ciento, noventa_por_ciento

def traeHorasExtras(values):
    try:
        with connections['S3A'].cursor() as cursor:
            sql = """
                DECLARE @@Inicio DATE;
                DECLARE @@Final DATE;
                SET @@Inicio = %s;
                SET @@Final = %s;
                SELECT 
                    IdLegajo AS LEGAJO, 
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_50,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '534' WHEN '9' THEN '630' WHEN '2' THEN '760'
                    WHEN '4' THEN '825' WHEN '11' THEN '882' ELSE '-' END AS COD_50,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = '100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_100,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '536' WHEN '9' THEN '635' WHEN '2' THEN '765'
                    WHEN '4' THEN '820' WHEN '11' THEN '884' ELSE '-' END AS COD_100,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'S' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_S,
                    CASE (SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) WHEN '10' THEN '532' ELSE '-' END AS COD_S,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-50' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_AC_50,
                    ROUND(SUM(CASE WHEN RTRIM(TipoHoraExtra) = 'AC-100' THEN CONVERT(FLOAT, CantHoras) ELSE 0 END), 2) AS HORAS_AC_100,
                    (SELECT CASE Sindicatos.DescrSindicato WHEN NULL THEN 'SIN SINDICATO **' ELSE Sindicatos.DescrSindicato END
                        FROM   TresAses_ISISPayroll.dbo.Empleados INNER JOIN
                                    TresAses_ISISPayroll.dbo.Sindicatos ON Empleados.Regis_Sin = Sindicatos.Regis_Sin
                        WHERE (Empleados.CodEmpleado = IdLegajo)) AS SINDICATO,
                    (SELECT (RTRIM(RH_Legajo.Apellidos) + ' ' + RTRIM(RH_Legajo.Nombres)) FROM RH_Legajo WHERE RH_Legajo.IdLegajo = RH_HE_Horas_Extras.IdLegajo) AS NOMBRE
                FROM 
                    RH_HE_Horas_Extras
                WHERE 
                    CONVERT(DATE, FechaDesde) >= @@Inicio
                    AND CONVERT(DATE, FechaHasta) <= @@Final
                    AND RTRIM(TipoHoraExtra) IN ('50', '100', 'S', 'AC-50', 'AC-100')
                    AND ((SELECT Regis_Sin FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = IdLegajo) NOT IN ('8'))
                    --AND IdLegajo IN ('59510','2278', '47087','52026','54965')
                GROUP BY 
                    IdLegajo
                ORDER BY 
                    IdLegajo;
            """
            cursor.execute(sql,values)
            results = cursor.fetchall()
            data = []
            if results:
                for row in results:
                    Legajo = str(row[0])
                    Sindicato = str(row[9])
                    Hora_50 = ''
                    Concepto_50 = ''
                    Hora_100 = ''
                    Concepto_100 = ''
                    Acuerdo_50 = str(row[7])
                    Acuerdo_100 = str(row[8])
                    Hora_50 = row[1]
                    Concepto_50 = str(row[2])
                    Hora_100 = row[3]
                    Concepto_100 = str(row[4])
                    Float_50 = float(row[7])
                    Float_100 = float(row[8])

                    if Legajo == '9':
                        Concepto_50 = '825'
                        Concepto_100 = '820'
                    if Legajo == '234':
                        Concepto_50 = '825'
                        Concepto_100 = '862'
                    if Legajo == '53979':
                        Concepto_50 = '825'
                        Concepto_100 = '862'
                    if Legajo == '56095':
                        Concepto_50 = '760'
                        Concepto_100 = '765'
                    if Legajo == '238':
                        Concepto_50 = '760'
                        Concepto_100 = '765'

                    total_horas_50 = 0.0
                    total_horas_100 = 0.0

                    if Sindicato == 'FRIGORIFICO':
                        if Acuerdo_100 != '0.0' or Acuerdo_50 != '0.0':
                            if Float_50 >= 6:
                                hora_10_50, hora_90_50 = calcular_porcentajes(Float_50)
                                total_horas_50 = round(hora_10_50 + float(Hora_50), 2)
                                datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_50), "CONCEPTO": Concepto_50}
                                data.append(datos)
                            else:
                                if Acuerdo_50 != '0.0':
                                    total_horas_50 = round(float(Acuerdo_50) + float(Hora_50), 2)
                                    datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_50).replace(',','.'), "CONCEPTO": Concepto_50}
                                    data.append(datos)
                            if Float_100 >= 6:
                                hora_10_100, hora_90_100 = calcular_porcentajes(Float_100)
                                total_horas_100 = round(hora_10_100 + float(Hora_100), 2)
                                datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_100), "CONCEPTO": Concepto_100}
                                data.append(datos)
                            else:
                                if Acuerdo_100 != '0.0':
                                    total_horas_100 = round(float(Acuerdo_100) + float(Hora_100), 2)
                                    datos = {"LEGAJO": Legajo, "HORAS": str(total_horas_100).replace(',','.'), "CONCEPTO": Concepto_100}
                                    data.append(datos)
                        else:
                            if str(Hora_50) != '0.0':
                                datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_50, 2)), "CONCEPTO": Concepto_50}
                                data.append(datos)
                            if str(Hora_100) != '0.0':
                                datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_100, 2)), "CONCEPTO": Concepto_100}
                                data.append(datos)
                    else:
                        if str(Hora_50) != '0.0':
                            datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_50, 2)), "CONCEPTO": Concepto_50}
                            data.append(datos)
                        if str(Hora_100) != '0.0':
                            datos = {"LEGAJO": Legajo, "HORAS": str(round(Hora_100, 2)), "CONCEPTO": Concepto_100}
                            data.append(datos)   
                return {"datos": data}
            else:
                return {"datos": []}
    except Exception as e:
        return  '0'
    
def response_excel(values):
    data_dict = traeHorasExtras(values)
    try:
        if data_dict != '0':
            wb = openpyxl.Workbook()
            ws = wb.active

            for idx, entry in enumerate(data_dict['datos'], start=1):
                ws[f'A{idx}'] = entry['LEGAJO']
                ws[f'B{idx}'] = entry['CONCEPTO']
                ws[f'C{idx}'] = entry['HORAS']
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename=output.xlsx'
            return response
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No se pudo completar la petición.'})
    except Exception as e:
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'Hubo un error al intentar resolver la petición. ' + str(e) })
    
def jsonArchivoIsis(values):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute("EXEC MDR_LISTADO_ARCHIVO_ISIS %s,%s,%s,%s", values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    lista_data.append({
                        'Legajo':str(row[0]), 
                        'Horas50':str(row[1]), 
                        'Cod50':str(row[2]), 
                        'Horas100':str(row[3]), 
                        'Cod100':str(row[4]), 
                        'HorasS':str(row[5]), 
                        'CodS':str(row[6]), 
                        'Ac50':str(row[7]), 
                        'Ac100':str(row[8]), 
                        'Sindicato':str(row[9]), 
                        'Nombre':str(row[10]), 
                        'Centro':str(row[11])
                    })
                return lista_data 
            else:
                return lista_data
    except Exception as e:
        return lista_data

@csrf_exempt    
def data_listado_planilla_labores(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Md_Rrhh.puede_ver')
        if user_has_permission:
            try:
                inicio = str(request.POST.get('Inicio'))
                final = str(request.POST.get('Final'))
                centro = str(request.POST.get('IdCentro'))
                labor = str(request.POST.get('IdLabor'))
                archivo = str(request.POST.get('Tipo'))
                filtros = request.POST.get('Filtros')
                filtros_dict = json.loads(filtros)
                values = [inicio,final,centro,labor]
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """ 
                        DECLARE @Inicio DATE;
                        DECLARE @Final DATE;
                        DECLARE @IdCentro VARCHAR(10);
                        DECLARE @Labor VARCHAR(10);

                        SET @Inicio = %s;
                        SET @Final = %s;
                        SET @IdCentro = %s;
                        SET @Labor = %s;


                        WITH Datos AS (
                        SELECT        ST.FECHA,
                        CASE WHEN ST.USUARIO = 'JCASTILLO' THEN
                                                    (SELECT        CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple))
                                                    FROM            Rommik_isispayroll.dbo.Empleados
                                                    WHERE        CodEmpleado = ST.ID_LEGAJO) ELSE
                                                    (SELECT        CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple))
                                                    FROM            TresAses_ISISPayroll.dbo.Empleados
                                                    WHERE        CodEmpleado = ST.ID_LEGAJO) END AS NOMBRES, CASE WHEN ST.USUARIO = 'JCASTILLO' THEN
                                                    (SELECT        CC.AbrevCtroCosto
                                                    FROM            Rommik_isispayroll.dbo.Empleados AS EMP INNER JOIN
                                                                                Rommik_isispayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EMP.Regis_CCo
                                                    WHERE        EMP.CodEmpleado = ST.ID_LEGAJO) ELSE
                                                    (SELECT        CC.AbrevCtroCosto
                                                    FROM            TresAses_ISISPayroll.dbo.Empleados AS EMP INNER JOIN
                                                                                TresAses_ISISPayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EMP.Regis_CCo
                                                    WHERE        EMP.CodEmpleado = ST.ID_LEGAJO) END AS ABREV_CENTRO, ST.ID_LEGAJO AS LEGAJO, CASE WHEN ST.USUARIO = 'JCASTILLO' THEN
                                                    (SELECT        CC.Regis_CCo
                                                    FROM            Rommik_isispayroll.dbo.Empleados AS EMP INNER JOIN
                                                                                Rommik_isispayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EMP.Regis_CCo
                                                    WHERE        EMP.CodEmpleado = ST.ID_LEGAJO) ELSE
                                                    (SELECT        CC.Regis_CCo
                                                    FROM            TresAses_ISISPayroll.dbo.Empleados AS EMP INNER JOIN
                                                                                TresAses_ISISPayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EMP.Regis_CCo
                                                    WHERE        EMP.CodEmpleado = ST.ID_LEGAJO) END AS REGIS_CCO, SL.NOMBRE_LABOR AS LABOR, ST.VALOR AS IMPORTE, 
                                                    CASE WHEN P_1.DIAS_TRABAJADOS IS NULL 
                                                THEN 0 ELSE P_1.DIAS_TRABAJADOS END AS DIAS_TRABAJADOS, SPA_QR.ID_FILA,
                                                CASE
                                                    WHEN SL.ALIAS_LABOR = 'P' THEN '51'
                                                    WHEN SL.ALIAS_LABOR = 'R' THEN '52'
                                                    WHEN SL.ALIAS_LABOR = 'C' THEN '53'
                                                END  AS CODIGO_ASISTENCIA,
                                                SL.ALIAS_LABOR AS ALIAS_LABOR
                        FROM            SPA_QR INNER JOIN
                                                SPA_CUADRO ON SPA_QR.ID_CUADRO = SPA_CUADRO.ID_CUADRO RIGHT OUTER JOIN
                                                SPA_TAREA AS ST INNER JOIN
                                                SPA_LABOR AS SL ON SL.ID_LABOR = ST.ID_LABOR ON SPA_QR.ID_QR = ST.ID_QR_FILA LEFT OUTER JOIN
                                                    (SELECT        IdLegajo, SUM(CASE WHEN P.M = 'P' THEN 0.5 ELSE 0 END + CASE WHEN P. T = 'P' THEN 0.5 ELSE 0 END) AS DIAS_TRABAJADOS
                                                    FROM            S3A.dbo.RH_Presentismo AS P
                                                    WHERE        (CONVERT(DATE, Fecha) >= @Inicio) AND (CONVERT(DATE, Fecha) <= @Final)
                                                    GROUP BY IdLegajo) AS P_1 ON P_1.IdLegajo = ST.ID_LEGAJO
                        )
                        SELECT NOMBRES, ABREV_CENTRO, LEGAJO, REGIS_CCO, LABOR, IMPORTE, DIAS_TRABAJADOS, ID_FILA, CODIGO_ASISTENCIA
                        FROM Datos 
                        WHERE 
                            CONVERT(DATE, FECHA) >= @Inicio 
                            AND CONVERT(DATE, FECHA) <= @Final 
                            AND (REGIS_CCO = @IdCentro OR (@IdCentro = '' AND REGIS_CCO NOT IN ('18','20')))
                            AND (ALIAS_LABOR = @Labor OR @Labor = '')
                        ORDER BY NOMBRES
                     """
                    cursor.execute(sql,values)  ##INSERTAR VALUES
                    results = cursor.fetchall()
                    listado_json = {}

                    if results:
                        for row in results:
                            legajo = str(row[2])
                            tarea = str(row[4])

                            if legajo not in listado_json:
                                listado_json[legajo] = {
                                    "LEGAJO": legajo,
                                    "NOMBRES": str(row[0]),
                                    "CENTRO": str(row[1]),
                                    "REGIS": str(row[3]),
                                    "TAREA": tarea,
                                    "DETALLES": [
                                        {
                                            "ITEM": str(row[8]), 
                                            "CANTIDAD": str(row[6]),
                                            "IMPORTE": "",
                                            "SUMA_IMPORTE": ""
                                        }
                                    ]
                                }

                            importe = float(row[5])

                            # Definir rangos de ítems según la tarea
                            start_item = 200 if tarea == "PODA" else 300
                            max_item = 215 if tarea == "PODA" else 315

                            detalles = listado_json[legajo]["DETALLES"]

                            # Buscar si ya existe un detalle con el mismo importe (a partir del segundo elemento)
                            found = False
                            for detalle in detalles[1:]:
                                if detalle["IMPORTE"] == str(importe):
                                    detalle["CANTIDAD"] = str(int(detalle["CANTIDAD"]) + 1)
                                    detalle["SUMA_IMPORTE"] = str(float(detalle["SUMA_IMPORTE"]) + importe)
                                    found = True
                                    break

                            # Si no existe, agregar nuevo ítem si no se supera el límite
                            if not found:
                                current_count = len(detalles) - 1  # excluye el ítem de asistencia
                                if start_item + current_count <= max_item:
                                    new_item = str(start_item + current_count)
                                    detalles.append({
                                        "ITEM": new_item,
                                        "CANTIDAD": "1",
                                        "IMPORTE": str(importe),
                                        "SUMA_IMPORTE": str(importe)
                                    })
                                else:
                                    #print(f"¡Límite de ítems alcanzado para LEGAJO {legajo} y tarea {tarea}!")
                                    pass


                        listado_json = {k: [v] for k, v in listado_json.items()}

                        if archivo == 'TT':
                            #listado_json = dict(sorted(listado_json.items(), key=lambda x: x[1]['NOMBRES']))
                            return JsonResponse({'Message': 'Success', 'Datos': listado_json})
                        else:
                            nombre_excel = crear_excel_rrhh(archivo,listado_json,filtros_dict)
                            return JsonResponse({'Message': 'Success', 'Datos': listado_json, 'Archivo':nombre_excel})
                    else:
                        data = 'No se encontraron datos.'
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Error', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def crear_excel_rrhh(tipo, json_data, filtros):
    if tipo == 'RP':
        try:
            lista_filas = []
            for legajo, registros in json_data.items():
                for reg in registros:
                    for detalle in reg['DETALLES']:
                        fila = {
                            'LEGAJO': str(reg['LEGAJO']),
                            'APELLIDO Y NOMBRE': str(reg['NOMBRES']),
                            'CENTRO': str(reg['CENTRO']),
                            'TAREA': str(reg['TAREA']),
                            'CODIGO': str(detalle['ITEM']),
                            'CANTIDAD': formatear_numero(detalle['CANTIDAD']),
                            'IMPORTE': formatear_importe(str(detalle['IMPORTE'])),
                            'SUMA TOTAL': formatear_importe(str(detalle['SUMA_IMPORTE']))
                        }
                        lista_filas.append(fila)
            lista_filas = sorted(lista_filas, key=lambda x: x['APELLIDO Y NOMBRE'])
            df = pd.DataFrame(lista_filas)
            output = BytesIO()
            columnas = ['LEGAJO', 'APELLIDO Y NOMBRE', 'CENTRO', 'TAREA', 'CODIGO', 'CANTIDAD', 'IMPORTE', 'SUMA TOTAL']
            df = df[columnas]
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='DETALLE IMPORTES')
                ws = writer.sheets['DETALLE IMPORTES']

                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                ws.add_image(logo, 'J2')
                ws['E5'] = 'DETALLE DE IMPORTES'
                ws['E5'].font = Font(size=14, bold=True)
                ws['E5'].alignment = Alignment(horizontal='center', vertical='center')

                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                ws['E2'] = fecha_actual
                ws['E2'].alignment = Alignment(horizontal='right', vertical='center')

                ws['A2'] = 'DESDE:'
                ws['B2'] = filtros['DESDE']
                ws['A2'].font = Font(size=12, bold=True)

                ws['A3'] = 'HASTA:'
                ws['B3'] = filtros['HASTA']
                ws['A3'].font = Font(size=12, bold=True)

                ws['A4'] = 'CENTRO:'
                ws['B4'] = filtros['CENTRO']
                ws['A4'].font = Font(size=12, bold=True)

                ws['A5'] = 'TAREA:'
                ws['B5'] = filtros['LABOR']
                ws['A5'].font = Font(size=12, bold=True)
                
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                for cell in ws[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=11, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 8
                    ws.column_dimensions[column].width = adjusted_width

                    columnas = ['E', 'F', 'G', 'H']
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        for cell in row:
                            if cell.column_letter in columnas:
                                cell.alignment = Alignment(horizontal='right')

                    columnas = ['A', 'C', 'D']
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        for cell in row:
                            if cell.column_letter in columnas:
                                cell.alignment = Alignment(horizontal='center')

            output.seek(0)
            nombre_excel = f'DETALLE_IMPORTES_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
            with open('Applications/Md_Rrhh/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())
            return nombre_excel
        except Exception as e:
            #print(f"Error: {e}")
            return 'e'
    
    else:
        try:
            lista_filas = []
            for legajo, registros in json_data.items():
                for reg in registros:
                    for detalle in reg['DETALLES']:
                        fila = {
                            'LEGAJO': str(reg['LEGAJO']),
                            'CODIGO': str(detalle['ITEM']),
                            'CC': "1",
                            'CD': "0",
                            'CE': "1",
                            'CF': "1",
                            'CG': "1",
                            'CANTIDAD': f"{float(detalle['CANTIDAD']):.2f}",
                            'CI': "1",
                            'CJ': "1",
                            'CK': "1",
                            'IMPORTE': f"{float(detalle['IMPORTE'] or 0):.2f}"
                        }
                        lista_filas.append(fila)

            df = pd.DataFrame(lista_filas)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, header=False, startrow=0, sheet_name='ARCHIVO ISIS')
                ws = writer.sheets['ARCHIVO ISIS']            
                columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'L', 'I', 'J', 'K']
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        if cell.column_letter in columnas:
                            cell.alignment = Alignment(horizontal='right')

            output.seek(0)
            nombre_excel = f'ARCHIVO_ISIS_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
            with open('Applications/Md_Rrhh/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())
            return nombre_excel
        except Exception as e:
            #print(f"Error: {e}")
            return 'e'

def formatear_numero(numero):
    if not numero:
        return ''
    return f"{float(numero):.2f}".replace(',', '.')

def formatear_importe(numero):
    if not numero:
        return ''
    return f"$ {float(numero):,.2f}".replace('.', '#').replace(',', '.').replace('#', ',')


############# LISTADO LABORES

def carga_inicial_listado_labores(request):
    if request.method == 'GET':
        try:
            listado_productores = []
            listado_personal = []
            listado_labores = [{'Codigo': 'P', 'Descripcion':'PODA'},{'Codigo': 'R', 'Descripcion':'RALEO'}]
            listado_encargados = []
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                #### listado productores
                sql = """ EXEC LISTADO_PRODUCTORES_HABILITADOS """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_productores.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })

                #### listado personal
                sql = """
                        SELECT EM.CodEmpleado AS LEGAJO, ((CONVERT(VARCHAR, EM.CodEmpleado)) + ' - ' + EM.ApellidoEmple + ' ' + EM.NombresEmple) AS NOMBRES
                        FROM TresAses_ISISPayroll.dbo.Empleados AS EM INNER JOIN
                            TresAses_ISISPayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EM.Regis_CCo
                        WHERE CC.AbrevCtroCosto LIKE ('C%%')
                                AND EM.BajaDefinitivaEmple = '2'
                        ORDER BY EM.ApellidoEmple + ' ' + EM.NombresEmple
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_personal.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })

                #### listado encargados
                sql = """
                        SELECT RTRIM(US.Usuario) AS ENCARGADO, CASE WHEN ( EM.ApellidoEmple + ' ' + EM.NombresEmple) IS NULL THEN US.Usuario
                            ELSE ( EM.ApellidoEmple + ' ' + EM.NombresEmple) END AS NOMBRES
                        FROM USUARIOS AS US LEFT JOIN
                            TresAses_ISISPayroll.dbo.Empleados AS EM ON EM.CodEmpleado = US.CodEmpleado
                        WHERE US.Estado = 'A'
                            AND US.CodEmpleado NOT IN('99999')
                            AND US.Tipo IN ('EC','G')
                        ORDER BY EM.ApellidoEmple + ' ' + EM.NombresEmple

                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_encargados.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })
            return JsonResponse({'Message': 'Success', 'Productores':listado_productores, 'Personal':listado_personal, 'Labores':listado_labores, 'Encargados':listado_encargados})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def cuadro_personal_x_chacra(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_cuadros = []
            listado_personal = []
            idChacra = request.POST.get('IdChacra')
            values = [idChacra]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql_cuadros = """
                    SELECT ID_CUADRO, NOMBRE_CUADRO
                    FROM SPA_CUADRO
                    WHERE ID_CHACRA = %s
                    ORDER BY NOMBRE_CUADRO
                    """
                cursor.execute(sql_cuadros, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_cuadros.append({
                            "IdCuadro": str(row[0]),
                            "Cuadro":str(row[1])
                        })

                sql_personal = """
                    SELECT DISTINCT(ST.ID_LEGAJO)  AS LEGAJO,
                        CASE
                            WHEN SC.ID_CHACRA = '1000001' THEN (SELECT CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple)) FROM Rommik_isispayroll.dbo.Empleados WHERE CodEmpleado = ST.ID_LEGAJO)
                            ELSE (SELECT CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple)) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = ST.ID_LEGAJO)
                        END AS NOMBRES
                    FROM SPA_TAREA AS ST INNER JOIN
                        SPA_QR AS QR ON QR.ID_QR = ST.ID_QR_FILA INNER JOIN
                        SPA_CUADRO AS SC ON SC.ID_CUADRO = QR.ID_CUADRO
                    WHERE (CONVERT(DATE,ST.FECHA) >= '2025-06-08')
                        AND	SC.ID_CHACRA = %s
                    ORDER BY NOMBRES

                    """
                cursor.execute(sql_personal,values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_personal.append({
                            "Legajo":str(row[0]),
                            "Nombre":str(row[1])
                        })
            return JsonResponse({'Message': 'Success', 'Cuadros': listado_cuadros, 'Personal': listado_personal})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def listado_detalle_labores(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idLegajo = str(request.POST.get('IdLegajo'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            idEncargado = str(request.POST.get('IdEncargado'))
            idLabor = str(request.POST.get('IdLabor'))
            values = [inicio,final,idLegajo,idChacra,idEncargado,idLabor,idCuadro]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ EXEC SP_SELECT_DETALLE_LABORES %s, %s, %s, %s, %s, %s, %s  """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "LEGAJO":row[0],
                            "NOMBRES":row[1],
                            "FECHA":row[2],
                            "QR":row[3],
                            "ID_CUADRO":row[4],
                            "ID_CHACRA":row[5],
                            "ID_PRODUCTOR":row[6],
                            "PRODUCTOR":row[7],
                            "CHACRA":row[8],
                            "CUADRO":row[9],
                            "FILA":row[10],
                            "VARIEDADES":row[11],
                            "CANT_PLANTAS":row[12],
                            "LABOR":row[13],
                            "IMPORTE":row[14],
                            "ID_QR_FILA":row[15],
                            "PRESUPUESTO":row[16],
                            "SUPERFICIE":row[17],
                            "TIPO_DIA":row[20]
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': listado_data})
                return JsonResponse({'Message': 'Error', 'Nota': 'No se encontraron datos.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})


def listado_combox_productor(request):
    if request.method == 'GET':
        try:
            lista_data = []
            lista_especies = []
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_PRODUCTORES_HABILITADOS
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        lista_data.append({
                            "IdProductor":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                sql_especies = """
                        EXEC LISTADO_ESPECIES_HABILITADAS
                """
                cursor.execute(sql_especies)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        lista_especies.append({
                            "IdEspecie":str(row[0]),
                            "Descripcion":str(row[1])
                        })


            if lista_data:
                return JsonResponse({'Message': 'Success', 'Datos': lista_data, 'Especies':lista_especies})
            else:
                data = "No se encontraron Datos."
                return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def listado_combox_chacras_x_productor(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idProductor = str(request.POST.get('IdProductor'))
            values = [idProductor]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_CHACRAS_X_PRODUCTOR %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] #{'IdChacra': '', 'Descripcion':'TODOS'}
                    for row in consulta:
                        lista_data.append({
                            "IdChacra":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data, 'Chacras':lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt
def listado_combox_variedades_x_especies(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idProductor = str(request.POST.get('IdEspecie'))
            values = [idProductor]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_VARIEDADES_X_ESPECIE  %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] #{'IdChacra': '', 'Descripcion':'TODOS'}
                    for row in consulta:
                        lista_data.append({
                            "IdVariedad":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt
def listado_combox_cuadros_x_chacra(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idChacra = str(request.POST.get('IdChacra'))
            values = [idChacra]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_CUADROS_X_CHACRA %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] 
                    for row in consulta:
                        lista_data.append({
                            "IdCuadro":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def convert_json_labores(json_data,tipo):
    try:
        if tipo == 'L':
            result = {}
            #print(json_data)
            for item in json_data:
                nombre = item['NOMBRES']
                legajo = item['LEGAJO']
                importe = float(item['IMPORTE'])
                detalle = {
                    'LEGAJO': str(legajo),  
                    'FECHA': item['FECHA'],
                    'TIPO_DIA': item['TIPO_DIA'],
                    'PRODUCTOR': item['PRODUCTOR'],
                    'CHACRA': item['CHACRA'],
                    'CUADRO': item['CUADRO'],
                    'FILA': item['FILA'],
                    'QR': item['QR'],
                    'LABOR': item['LABOR'],
                    'IMPORTE': formatear_moneda(importe),
                    'PRESUPUESTO': formatear_moneda(item['PRESUPUESTO']),
                    'VARIEDADES': item['VARIEDADES'],
                    'PLANTAS': item['CANT_PLANTAS'],
                    'SUPERFICIE': item['SUPERFICIE']
                }
                if nombre in result:
                    result[nombre]['IMPORTE_TOTAL'] += importe
                    result[nombre]['DETALLES'].append(detalle)
                else:
                    result[nombre] = {
                        'NOMBRES': nombre,
                        'IMPORTE_TOTAL': importe,
                        'DETALLES': [detalle]
                    }
            result_json = sorted(list(result.values()), key=lambda x: x['NOMBRES'])
            return result_json
        elif tipo == 'C':
            result = {}
            for item in json_data:
                chacra = item['CHACRA']
                importe = float(item['IMPORTE'])
                detalle = {
                    'LEGAJO': str(item['LEGAJO']),
                    'NOMBRES': item['NOMBRES'],
                    'FECHA': item['FECHA'],
                    'PRODUCTOR': item['PRODUCTOR'],
                    'CUADRO': item['CUADRO'],
                    'FILA': item['FILA'],
                    'QR': item['QR'],
                    'LABOR': item['LABOR'],
                    'IMPORTE': importe,
                    'PRESUPUESTO': item['PRESUPUESTO'],
                    'VARIEDADES': item['VARIEDADES'],
                    'PLANTAS': item['CANT_PLANTAS'],
                    'SUPERFICIE': item['SUPERFICIE']
                }
                if chacra in result:
                    result[chacra]['IMPORTE_TOTAL'] += importe
                    result[chacra]['DETALLES'].append(detalle)
                else:
                    result[chacra] = {
                        'CHACRA': chacra,
                        'IMPORTE_TOTAL': importe,
                        'DETALLES': [detalle]
                    }
            for chacra in result:
                result[chacra]['DETALLES'] = sorted(result[chacra]['DETALLES'], key=lambda x: x['NOMBRES'])

            result_json = list(result.values())
            return result_json
    except Exception as e:
        return 'e'

@csrf_exempt
def archivo_detalle_labores(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idLegajo = str(request.POST.get('IdLegajo'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            idEncargado = str(request.POST.get('IdEncargado'))
            idLabor = str(request.POST.get('IdLabor'))
            archivo = str(request.POST.get('Archivo'))
            tipo = str(request.POST.get('Tipo'))
            values = [inicio,final,idLegajo,idChacra,idEncargado,idLabor,idCuadro]
            filtros = request.POST.get('Filtros')
            filtros_dict = json.loads(filtros)
            if tipo == 'RP':
                values = [inicio,final,idLegajo,idChacra,idEncargado]
                nombre_excel = consulta_resumido_persona(values,filtros_dict)
                return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel})
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ EXEC SP_SELECT_DETALLE_LABORES %s, %s, %s, %s, %s, %s, %s  """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "LEGAJO":row[0],
                            "NOMBRES":row[1],
                            "FECHA":row[2],
                            "QR":row[3],
                            "ID_CUADRO":row[4],
                            "ID_CHACRA":row[5],
                            "ID_PRODUCTOR":row[6],
                            "PRODUCTOR":row[7],
                            "CHACRA":row[8],
                            "CUADRO":row[9],
                            "FILA":row[10],
                            "VARIEDADES":row[11],
                            "CANT_PLANTAS":row[12],
                            "LABOR":row[13],
                            "IMPORTE":row[14],
                            "ID_QR_FILA":row[15],
                            "PRESUPUESTO":row[16],
                            "SUPERFICIE":row[17],
                            "SUM_PLANTAS":row[18],
                            "SUM_SUPERFICIE":row[19],
                            "TIPO_DIA":row[20]
                        })
                        suma_plantas = row[18]
                        suma_superficie = row[19]
                    totales = {
                        "IMPORTE": sum(item["IMPORTE"] for item in listado_data),
                        "PRESUPUESTO": sum(item["PRESUPUESTO"] for item in listado_data),
                        "CANT_PLANTAS": sum(item["CANT_PLANTAS"] for item in listado_data),
                        "SUPERFICIE": sum(item["SUPERFICIE"] for item in listado_data)
                    }
                    totales_censo = {
                        "SUMA_PLANTAS":suma_plantas,
                        "SUMA_SUPERFICIE":suma_superficie,
                        "PORCEN_PLANTAS":round(float((totales["CANT_PLANTAS"]/suma_plantas)*100),2),
                        "PORCEN_SUPERFICIE":round(float((totales["SUPERFICIE"]/suma_superficie)*100),2) 
                    }
                    if archivo == 'excel':
                        nombre_excel = crear_excel_labores(listado_data,tipo,filtros_dict,totales,totales_censo)
                        return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel})
                    if archivo == 'pdf':
                        return JsonResponse({'Message': 'Error', 'Nota': 'Este tipo de Archivo aún no esta disponible.'})
                return JsonResponse({'Message': 'Error', 'Nota': 'No se encontraron datos para generar el archivo.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def crear_excel_labores(jsonData,tipo,filtros,totales,total_censo):
    if tipo == 'DC':
        lista_data = convert_json_labores(jsonData,"C")
        try:
            df = pd.json_normalize(lista_data, 'DETALLES', ['CHACRA'])
            output = BytesIO()
            columns1 = ['CHACRA', 'NOMBRES', 'LEGAJO', 'FECHA', 'PRODUCTOR', 'CUADRO', 'FILA', 'QR', 'LABOR', 'IMPORTE', 'VARIEDADES', 'PLANTAS']
            df = df[columns1]
            #df.fillna('', inplace=True)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=5, sheet_name='RESUMIDO POR LEGAJO', header=columns1)
                worksheet = writer.sheets['RESUMIDO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'G2')
                worksheet['C4'] = 'DETALLE LABORES RESUMIDO POR LEGAJO'
                worksheet['C4'].font = Font(size=14, bold=True)
                worksheet['C4'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                for cell in worksheet[6]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=7, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[6:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width

                # for cell in worksheet['J']:
                #    if cell.row > 6 and isinstance(cell.value, (int, float)):
                #       cell.number_format = '#.##0,0'
                    
            output.seek(0)
            nombre_excel = f'Listado_Labores_Resumido_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Rrhh/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'
    if tipo == 'DP':
        lista_data = convert_json_labores(jsonData,"L")
        try:
            df = pd.json_normalize(lista_data, 'DETALLES', ['NOMBRES'])
            output = BytesIO()
            columns1 = ['NOMBRES', 'LEGAJO', 'CHACRA', 'FECHA', 'TIPO_DIA', 'PRODUCTOR', 'CUADRO', 'FILA', 'QR', 'LABOR', 'IMPORTE', 'PRESUPUESTO', 'VARIEDADES', 'PLANTAS', 'SUPERFICIE']
            df = df[columns1]
            #df.fillna('', inplace=True)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='DETALLADO POR LEGAJO', header=columns1)
                worksheet = writer.sheets['DETALLADO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'J2')
                worksheet['E5'] = 'LABORES POR LEGAJO'
                worksheet['E5'].font = Font(size=14, bold=True)
                worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros['DESDE']
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros['HASTA']
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A4'] = 'Productor:'
                worksheet['B4'] = filtros['PRODUCTOR']
                worksheet['A4'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros['CHACRA']
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A6'] = 'Cuadro:'
                worksheet['B6'] = filtros['CUADRO']
                worksheet['A6'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Personal:'
                worksheet['B7'] = filtros['PERSONAL']
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Labor:'
                worksheet['B8'] = filtros['LABOR']
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width
                    
                worksheet.cell(row=worksheet.max_row + 1, column=9).value = "TOTALES:"
                worksheet.cell(row=worksheet.max_row, column=10).value = formatear_moneda(totales["IMPORTE"])
                worksheet.cell(row=worksheet.max_row, column=11).value = formatear_moneda(totales["PRESUPUESTO"])
                worksheet.cell(row=worksheet.max_row, column=13).value = totales["CANT_PLANTAS"]
                worksheet.cell(row=worksheet.max_row, column=14).value = totales["SUPERFICIE"]
                
                for i in range(9, worksheet.max_column + 1):
                    cell = worksheet.cell(row=worksheet.max_row, column=i)
                    cell.font = Font(bold=True)
                    cell.border = border_style
                
                # Obtén la fila actual
                fila_actual = worksheet.max_row + 2

                # Agrega la primera fila de información
                worksheet.cell(row=fila_actual, column=12).value = "TOTAL CENSO:"
                worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                worksheet.cell(row=fila_actual, column=13).value = total_censo["SUMA_PLANTAS"]
                worksheet.cell(row=fila_actual, column=14).value = total_censo["SUMA_SUPERFICIE"]

                # Agrega la segunda fila de información
                fila_actual += 1
                worksheet.cell(row=fila_actual, column=12).value = "% REALIZADO:"
                worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                worksheet.cell(row=fila_actual, column=13).value = str(total_censo["PORCEN_PLANTAS"]) + ' %'
                worksheet.cell(row=fila_actual, column=14).value = str(total_censo["PORCEN_SUPERFICIE"]) + ' %'

                # Aplica borde a las celdas
                for i in range(12, 15):
                    for j in range(fila_actual - 1, fila_actual + 1):
                        cell = worksheet.cell(row=j, column=i)
                        cell.border = border_style
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    for cell in row:
                        if cell.column_letter in ['J', 'K', 'M', 'N']:
                            cell.alignment = Alignment(horizontal='right')
                    
            output.seek(0)
            nombre_excel = f'Listado_Labores_Detallado_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Rrhh/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'
    
def crear_excel_resumido(tipo,lista_data,filtros):
    if tipo == 'RP':
        try:
            df = pd.DataFrame(lista_data)
            output = BytesIO()
            columns1 = ['NOMBRES', 'LEGAJO', 'CANT_PLANTAS', 'LABOR', 'IMPORTE', 'VALOR_REFERENCIA', 'DIAS_TRABAJADOS', 'IMPORTE_DIA']
            df = df[columns1]
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='RESUMIDO POR LEGAJO')
                worksheet = writer.sheets['RESUMIDO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'J2')
                worksheet['E5'] = 'RESUMIDO POR LEGAJO'
                worksheet['E5'].font = Font(size=14, bold=True)
                worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros['DESDE']
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros['HASTA']
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A4'] = 'Productor:'
                worksheet['B4'] = filtros['PRODUCTOR']
                worksheet['A4'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros['CHACRA']
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Personal:'
                worksheet['B7'] = filtros['PERSONAL']
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Labor:'
                worksheet['B8'] = filtros['LABOR']
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width
                    
                # worksheet.cell(row=worksheet.max_row + 1, column=9).value = "TOTALES:"
                # worksheet.cell(row=worksheet.max_row, column=10).value = formatear_moneda(totales["IMPORTE"])
                # worksheet.cell(row=worksheet.max_row, column=11).value = formatear_moneda(totales["PRESUPUESTO"])
                # worksheet.cell(row=worksheet.max_row, column=13).value = totales["CANT_PLANTAS"]
                # worksheet.cell(row=worksheet.max_row, column=14).value = totales["SUPERFICIE"]
                
                # for i in range(9, worksheet.max_column + 1):
                #     cell = worksheet.cell(row=worksheet.max_row, column=i)
                #     cell.font = Font(bold=True)
                #     cell.border = border_style
                
                # Obtén la fila actual
                # fila_actual = worksheet.max_row + 2

                # # Agrega la primera fila de información
                # worksheet.cell(row=fila_actual, column=12).value = "TOTAL CENSO:"
                # worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                # worksheet.cell(row=fila_actual, column=13).value = total_censo["SUMA_PLANTAS"]
                # worksheet.cell(row=fila_actual, column=14).value = total_censo["SUMA_SUPERFICIE"]

                # # Agrega la segunda fila de información
                # fila_actual += 1
                # worksheet.cell(row=fila_actual, column=12).value = "% REALIZADO:"
                # worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                # worksheet.cell(row=fila_actual, column=13).value = str(total_censo["PORCEN_PLANTAS"]) + ' %'
                # worksheet.cell(row=fila_actual, column=14).value = str(total_censo["PORCEN_SUPERFICIE"]) + ' %'

                # # Aplica borde a las celdas
                # for i in range(12, 15):
                #     for j in range(fila_actual - 1, fila_actual + 1):
                #         cell = worksheet.cell(row=j, column=i)
                #         cell.border = border_style
                # for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                #     for cell in row:
                #         if cell.column_letter in ['J', 'K', 'M', 'N']:
                #             cell.alignment = Alignment(horizontal='right')
                    
            output.seek(0)
            nombre_excel = f'Resumido_Persona_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Rrhh/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'

def consulta_resumido_persona(values,filtros):
    listado_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """ 
                EXEC SP_SELECT_RESUMEN_LABORES %s, %s, %s, %s, %s

              """
            cursor.execute(sql, values)
            results = cursor.fetchall()
            if results:
                for row in results:
                    listado_data.append({
                        "NOMBRES":row[0],
                        "LEGAJO":row[1],
                        "CANT_PLANTAS":row[2],
                        "LABOR":row[3],
                        "IMPORTE":row[4],
                        "VALOR_REFERENCIA":row[5],
                        "DIAS_TRABAJADOS":row[6],
                        "IMPORTE_DIA":row[7],
                    })
            nombre_excel = crear_excel_resumido('RP',listado_data,filtros)
        return nombre_excel
    except Exception as e:
        return 'e'



























































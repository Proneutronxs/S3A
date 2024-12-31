from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from S3A.funcionesGenerales import *
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connections, transaction
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from Applications.RRHH.views import buscaDatosParaInsertarHE, obtener_fecha_hora_actual_con_milisegundos
from Applications.NotificacionesPush.notificaciones_push import notificaciones_Fruit_Truck
import datetime 

# Create your views here.
@login_required
def Logistica(request):
    return render (request, 'Logistica/logistica.html')

@login_required
def seguimientoViajes(request):
    return render (request, 'Logistica/Seguimiento/seguimientoViajes.html')

@login_required
def horasExtras(request):
    return render (request, 'Logistica/HorasExtras/horasExtrasLogistica.html')

@login_required
def horasExtras_autoriza(request):
    return render (request, 'Logistica/HorasExtras/autorizaHorasExtras.html')

@login_required
def verHorasExtras(request):
    return render (request, 'Logistica/HorasExtras/verHorasExtrasLogistica.html')

@login_required
def pedidos_flete(request):
    return render (request, 'Logistica/PedidosFlete/pedidosFlete.html')

@login_required
def asignacion_pedidos_flete(request):
    return render (request, 'Logistica/PedidosFlete/verPedidos.html')

@login_required
def baja_pedidos_flete(request):
    return render (request, 'Logistica/PedidosFlete/bajaPedidos.html')

@login_required
def ultima_ubicacion_choferes(request):
    return render (request, 'Logistica/PedidosFlete/ubicacionChoferes.html')

@login_required
@csrf_exempt
def cargaCentrosEmpaque(request):### CAMBIO A CENTRO DE COSTOS MUESTRA LOS CENTROS DE COSTOS CARGADOS
    if request.method == 'GET':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        usuario = str(request.user)
        if user_has_permission:
            try:
                data = buscaCentroCostos("CENTROS-LOGISTICA",usuario.upper())
                if data:
                    return JsonResponse({'Message': 'Success', 'Datos': data})
                else:
                    data = "No se encontraron Centros de Costos Asignados."
                    return JsonResponse({'Message': 'Error', 'Nota': data})                
            except Exception as e:
                data = str(e)
                insertar_registro_error_sql("Empaque","cargaLegajosEmpaque",request.user,data)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})   
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def buscaCentroCostos(codigo,usuario):
    listado =  buscaCentroCostosPorUsuario(codigo,usuario)
    lista_json = []
    for item in listado:
        try:
            with connections['ISISPayroll'].cursor() as cursor:
                sql = "SELECT Regis_CCo, DescrCtroCosto " \
                    "FROM CentrosCostos " \
                    "WHERE Regis_CCo = %s"
                cursor.execute(sql,[item])
                consulta = cursor.fetchone()
                if consulta:
                    ids = str(consulta[0])
                    cc = str(consulta[1])
                    data = {'cc':cc, 'id':ids}
                    lista_json.append(data)
        except Exception as e:
            error = str(e)
            insertar_registro_error_sql("LOGISTICA","BUSCA CENTRSO DE COSTO",usuario,error)
            return lista_json
        finally:
            cursor.close()
            connections['ISISPayroll'].close()
    return lista_json

@login_required
@csrf_exempt
def mostrarHorasCargadasPorCC(request): ### MUESTRA LA TABLA DE HORAS
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            cc = request.POST.get('ComboxTipoCCAutorizaFrio')
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = "SELECT     RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25), TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRES, CONVERT(VARCHAR(10), " \
                                        "HorasExtras_Procesadas.FechaHoraDesde, 103) AS FECHA_DESDE, CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) AS HORA_DESDE, CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) AS FECHA_HASTA, CONVERT(VARCHAR(5), " \
                                        "HorasExtras_Procesadas.FechaHoraHasta, 108) AS HORA_HASTA, RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) AS MOTIVO, RTRIM(HorasExtras_Procesadas.DescripcionMotivo) AS DESCRIPCION, CONVERT(VARCHAR(5), HorasExtras_Procesadas.CantidadHoras) AS HORAS, " \
                                        "RTRIM(S3A.dbo.RH_HE_Autoriza.Apellidos) AS AUTORIZADO, HorasExtras_Procesadas.ID_HEP AS idHoras, TresAses_ISISPayroll.dbo.CentrosCostos.DescrCtroCosto AS CENTROCOSTOS, (SELECT CONVERT(VARCHAR(21),ApellidoEmple + ' ' + NombresEmple) " \
                                        "FROM TresAses_ISISPayroll.dbo.Empleados " \
                                        "WHERE CodEmpleado = HorasExtras_Procesadas.UsuarioEncargado) AS SOLICITA, HorasExtras_Procesadas.EstadoEnvia " \
                        "FROM        TresAses_ISISPayroll.dbo.CentrosCostos INNER JOIN " \
                                        "S3A.dbo.RH_HE_Autoriza INNER JOIN " \
                                        "S3A.dbo.RH_HE_Motivo INNER JOIN " \
                                        "TresAses_ISISPayroll.dbo.Empleados INNER JOIN " \
                                        "HorasExtras_Procesadas ON TresAses_ISISPayroll.dbo.Empleados.CodEmpleado = HorasExtras_Procesadas.Legajo ON S3A.dbo.RH_HE_Motivo.IdMotivo = HorasExtras_Procesadas.IdMotivo ON S3A.dbo.RH_HE_Autoriza.IdAutoriza = HorasExtras_Procesadas.Autorizado ON " \
                                        "TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo " \
                        "WHERE     (TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo = %s) AND (HorasExtras_Procesadas.EstadoEnvia IN (1,3)) " \
                        "ORDER BY LEGAJO, HorasExtras_Procesadas.FechaHoraDesde"
                    cursor.execute(sql, [cc])
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            tipo = str(i[0])
                            legajo = str(i[1])
                            nombres = str(i[2])
                            desde = str(i[3]) + " - " + str(i[4])
                            hasta = str(i[5]) + " - " + str(i[6])
                            motivo = str(i[7])
                            descripcion = str(i[8])
                            horas = str(i[9])
                            idHoras = str(i[11])
                            centro = str(i[12])
                            solicita = str(i[13])
                            datos = {'ID':idHoras,'tipo':tipo, 'legajo':legajo, 'nombres':nombres, 'desde':desde, 'hasta':hasta, 'motivo':motivo, 'descripcion':descripcion, 'horas':horas , 'centro': centro, 'solicita': solicita}
                            
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No se encontrarón horas extras procesadas."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","mostrarHorasCargadasPorCCFrio",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def autorizaHorasCargadas(request): ### INSERTA LAS HORAS SELECCIONADAS
    if request.method == 'POST': 
        user_has_permission = request.user.has_perm('Logistica.puede_autorizar') 
        if user_has_permission: 
            usuario = str(request.user)
            checkboxes_tildados = request.POST.getlist('idCheck')
            horas = request.POST.getlist('cantHoras')
            tipo = request.POST.getlist('tipoHoraExtra')
            resultados = []
            importe = "0"
            pagada = "N"
            index = 0
            for i in checkboxes_tildados:
                ID_HEP = str(i) 
                tipo_hora = str(tipo[index])
                cant_hora = str(horas[index])
                fecha_y_hora = str(obtener_fecha_hora_actual_con_milisegundos())
                Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora = buscaDatosParaInsertarHE(ID_HEP) 
                resultado = insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, cant_hora, IdMotivo, IdAutoriza, Descripcion, tipo_hora, importe, pagada, fecha_y_hora,usuario.upper())
                resultados.append(resultado)
                index = index + 1

            if 0 in resultados:
                data = "Se produjo un Error en alguna de las inserciones."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se autorizaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'}) 
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
def insertaHorasExtras(ID_HEP,Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user): ### FUNCION QUE INSERTA LAS HORAS EN EL S3A SOLO FUNCION (NO REQUIERE PERMISOS)
    try:
        with connections['S3A'].cursor() as cursor:
            sql = "INSERT INTO RH_HE_Horas_Extras (IdRepl,IdHoraExtra,IdLegajo,FechaDesde,HoraDesde,FechaHasta,HoraHasta,CantHoras,IdMotivo,IdAutoriza,Descripcion,TipoHoraExtra,Valor,Pagada,FechaAlta,UserID) " \
                    "VALUES ('100',(SELECT MAX (IdHoraExtra) + 1 FROM RH_HE_Horas_Extras),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            values = (Legajo, Fdesde, Hdesde, Fhasta, Hhasta, Choras, IdMotivo, IdAutoriza, Descripcion, Thora, importe, pagada, fecha_y_hora,user)
            cursor.execute(sql, values)
            actualizarEstadoHEP(ID_HEP,Thora,Choras)
            return 1
    except Exception as e:
        insertar_registro_error_sql("LOGISTICA","insertaHorasExtras","request.user",str(e))
        return 0
    finally:
        connections['S3A'].close()

def actualizarEstadoHEP(ID_HEP, tipo, horas): ### FUNCIÓN QUE ACTUALIZA EL ESTADO SI SE ENVÍO LA HS EXTRA SÓLO FUNCIÓN (NO REQUIERE PERMISOS)
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '0', CantidadHoras = %s, TipoHoraExtra = %s WHERE ID_HEP = %s"
            cursor.execute(sql, [horas,tipo,ID_HEP])
    except Exception as e:
        insertar_registro_error_sql("LOGISTICA","actualizarEstadoHEP","request.user",str(e))

@login_required
@csrf_exempt
def eliminaHorasSeleccionadas(request): ### ELIMINA LAS HORAS SELECCIONADAS
    if request.method == 'POST':   ### METODO POST PUT DELETE GET
        user_has_permission = request.user.has_perm('Logistica.puede_denegar')
        if user_has_permission:
            checkboxes_tildados = request.POST.getlist('idCheck')
            resultados = []
            for i in checkboxes_tildados:
                ID_HEP = str(i) 
                resultado = eliminaHorasEstadoHEP(ID_HEP)
                resultados.append(resultado)
            if 0 in resultados:
                data = "Se produjo un Error en alguna de las eliminaciones."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                data = "Las horas se eliminaron correctamente."
                return JsonResponse({'Message': 'Success', 'Nota': data})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'}) 
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
def eliminaHorasEstadoHEP(ID_HEP):
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = "UPDATE HorasExtras_Procesadas SET EstadoEnvia = '8' WHERE ID_HEP = %s"
            cursor.execute(sql, [ID_HEP])
            
            slq2 = "UPDATE HorasExtras_Sin_Procesar SET Estado='8' WHERE ID_HESP = (SELECT ID_HESP FROM HorasExtras_Procesadas WHERE ID_HEP = %s)"
            cursor.execute(slq2, [ID_HEP])
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","eliminaHorasEstadoHEP","usuario",error)
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def listadoHorasExtrasEstadoL(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            usuario = str(request.user)
            ###NUEVA
            listado =  buscaCentroCostosPorUsuario("CENTROS-LOGISTICA",usuario)
            numeros_str = [str(numero) for numero in listado]
            numeros_concatenados = ', '.join(numeros_str)
            ###NUEVA
            cc = str(request.POST.get('ComboxCentrosCostos'))
            desde = request.POST.get('fechaBusquedaDesde')
            hasta = request.POST.get('fechaBusquedaHasta')
            estado = str(request.POST.get('ComboxEstadoHora'))
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = f""" 
                            DECLARE @@Desde DATE;
                            DECLARE @@Hasta DATE;
                            DECLARE @@Centro INT;
                            DECLARE @@Estado INT;
                            SET @@Desde = %s;
                            SET @@Hasta = %s;
                            SET @@Centro = %s;
                            SET @@Estado = %s;
                            SELECT        HorasExtras_Procesadas.ID_HEP AS ID, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25),TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRE, 
                                            (SELECT TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto FROM TresAses_ISISPayroll.dbo.CentrosCostos WHERE Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo) AS CC,
                                            (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) + ' Hs.') AS DESDE, (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) + ' Hs.') AS HASTA, 
                                            (SELECT RTRIM(S3A.dbo.RH_HE_Motivo.Descripcion) FROM S3A.dbo.RH_HE_Motivo WHERE S3A.dbo.RH_HE_Motivo.IdMotivo= HorasExtras_Procesadas.IdMotivo) AS MOTIVO, 
                                            RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, FORMAT(HorasExtras_Procesadas.CantidadHoras, '0.0') AS CANTIDAD, 
                                            CASE WHEN HorasExtras_Procesadas.ID_HESP IS NULL THEN 'WEB' ELSE 'APP' END AS SECTOR,
                                            CASE WHEN HorasExtras_Procesadas.EstadoEnvia = '8' THEN 'RECHAZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '0' THEN 'AUTORIZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '4' THEN 'PENDIENTE'
                                            WHEN HorasExtras_Procesadas.EstadoEnvia = '3' THEN 'PENDIENTE' ELSE 'DESCONOCIDO' END AS ESTADO, HorasExtras_Procesadas.EstadoEnvia AS ID_ESTADO,
                                            HorasExtras_Procesadas.DescripcionMotivo AS MT
                            FROM            HorasExtras_Procesadas INNER JOIN
                                                    TresAses_ISISPayroll.dbo.Empleados ON HorasExtras_Procesadas.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado
                            WHERE (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraDesde) >= @@Desde OR @@Desde IS NULL OR @@Desde = '')
                                        AND (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraHasta) <= @@Hasta OR @@Hasta IS NULL OR @@Hasta = '')
                                        AND ((@@Estado = '0' AND EstadoEnvia = '0') OR
                                            (@@Estado = '8' AND EstadoEnvia = '8') OR
                                            (@@Estado = '10' AND EstadoEnvia IN (0,8)))
                                        AND (HorasExtras_Procesadas.EstadoEnvia NOT IN (1,4,3))
                                        AND ((TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN ({numeros_concatenados}) AND @@Centro = '0')
                                        OR (TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN (@@Centro) AND @@Centro <> '0'))
                            ORDER BY TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple, HorasExtras_Procesadas.FechaHoraDesde, HorasExtras_Procesadas.FechaHoraHasta

                        """
                    cursor.execute(sql,[desde,hasta,cc,estado])
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            idHora = str(row[0])
                            legajo = str(row[1])
                            nombre = str(row[2])
                            cc = str(row[3])
                            desde = str(row[4])
                            hasta = str(row[5])
                            motivo = str(row[6]) + ' - ' + str(row[12])
                            tipo = str(row[7])
                            horas = str(row[8])
                            dispositivo = str(row[9])
                            estado = str(row[10])
                            id_estado = str(row[11])
                            datos = {'ID':idHora, 'Tipo':tipo, 'Legajo':legajo, 'Nombre':nombre, 'CC':cc, 'Desde':desde,
                                    'Hasta':hasta, 'Motivo': motivo, 'Horas':horas, 'Dispositivo':dispositivo, 'Estado':estado, 'ID_Estado': id_estado}
                            data.append(datos)

                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron horas.'})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("FRIO","LISTA HORAS PROCESADAS",str(request.user),error)
                return JsonResponse({'Message': 'Error', 'Nota': error})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})  

@login_required
@csrf_exempt
def creaExcelHorasMostradas(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Empaque.puede_ver')
        if user_has_permission:
            usuario = str(request.user)
            ###NUEVA
            listado =  buscaCentroCostosPorUsuario("CENTROS-LOGISTICA",usuario)
            numeros_str = [str(numero) for numero in listado]
            numeros_concatenados = ', '.join(numeros_str)
            ###NUEVA
            cc = str(request.POST.get('ComboxCentrosCostos'))
            desde = request.POST.get('fechaBusquedaDesde')
            hasta = request.POST.get('fechaBusquedaHasta')
            estado = str(request.POST.get('ComboxEstadoHora'))
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = f""" 
                            DECLARE @@Desde DATE;
                            DECLARE @@Hasta DATE;
                            DECLARE @@Centro INT;
                            DECLARE @@Estado INT;
                            SET @@Desde = %s;
                            SET @@Hasta = %s;
                            SET @@Centro = %s;
                            SET @@Estado = %s;
                            SELECT        HorasExtras_Procesadas.ID_HEP AS ID, HorasExtras_Procesadas.Legajo AS LEGAJO, CONVERT(VARCHAR(25),TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple) AS NOMBRE, 
                                            (SELECT TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto FROM TresAses_ISISPayroll.dbo.CentrosCostos WHERE Regis_CCo = TresAses_ISISPayroll.dbo.Empleados.Regis_CCo) AS CC,
                                            (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraDesde, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraDesde, 108) + ' Hs.') AS DESDE, (CONVERT(VARCHAR(10), HorasExtras_Procesadas.FechaHoraHasta, 103) + ' - ' + CONVERT(VARCHAR(5), HorasExtras_Procesadas.FechaHoraHasta, 108) + ' Hs.') AS HASTA, 
                                            HorasExtras_Procesadas.DescripcionMotivo AS MOTIVO, 
                                            RTRIM(HorasExtras_Procesadas.TipoHoraExtra) AS TIPO, FORMAT(HorasExtras_Procesadas.CantidadHoras, '0.0') AS CANTIDAD, 
                                            CASE WHEN HorasExtras_Procesadas.ID_HESP IS NULL THEN 'WEB' ELSE 'APP' END AS SECTOR,
                                            CASE WHEN HorasExtras_Procesadas.EstadoEnvia = '8' THEN 'RECHAZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '0' THEN 'AUTORIZADO' WHEN HorasExtras_Procesadas.EstadoEnvia = '4' THEN 'PENDIENTE'
                                            WHEN HorasExtras_Procesadas.EstadoEnvia = '3' THEN 'PENDIENTE' ELSE 'DESCONOCIDO' END AS ESTADO
                            FROM            HorasExtras_Procesadas INNER JOIN
                                                    TresAses_ISISPayroll.dbo.Empleados ON HorasExtras_Procesadas.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado
                            WHERE (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraDesde) >= @@Desde OR @@Desde IS NULL OR @@Desde = '')
                                        AND (TRY_CONVERT(DATE, HorasExtras_Procesadas.FechaHoraHasta) <= @@Hasta OR @@Hasta IS NULL OR @@Hasta = '')
                                        AND ((@@Estado = '0' AND EstadoEnvia = '0') OR
                                            (@@Estado = '8' AND EstadoEnvia = '8') OR
                                            (@@Estado = '10' AND EstadoEnvia IN (0,8)))
                                        AND (HorasExtras_Procesadas.EstadoEnvia NOT IN (1,4,3))
                                        AND ((TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN ({numeros_concatenados}) AND @@Centro = '0')
                                        OR (TresAses_ISISPayroll.dbo.Empleados.Regis_CCo IN (@@Centro) AND @@Centro <> '0'))
                            ORDER BY TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple, HorasExtras_Procesadas.FechaHoraDesde, HorasExtras_Procesadas.FechaHoraHasta


                        """
                    cursor.execute(sql,[desde,hasta,cc,estado])
                    results = cursor.fetchall()
                    if results:
                        
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = "Horas_Extras"
                        
                        headers = ['ID_HORA', 'LEGAJO', 'NOMBRE', 'CC', 'FECHA HORA INICIO', 'FECHA HORA FIN', 'DESCRIPCIÓN - MOTIVO', 'TIPO', 'CANTIDAD', 'SISTEMA', 'ESTADO']
                        
                        sheet.append(headers)
                        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        for cell in sheet[1]:
                            cell.fill = header_fill

                        border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
                        
                        for row in sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                cell.border = border
                        
                        for row in sheet.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=len(headers)):
                            for cell in row:
                                cell.border = border

                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)
                        sheet.insert_rows(1)

                        sheet.insert_cols(1)

                        for row_number, row_data in enumerate(results, start=8):
                            for col_number, value in enumerate(row_data, start=2):
                                sheet.cell(row=row_number, column=col_number, value=value).alignment = Alignment(horizontal='left', vertical='center')

                        
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 3)
                            sheet.column_dimensions[column].width = adjusted_width

                        for row in sheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        img = Image('static/imagenes/TA.png')
                        img.width = 160 
                        img.height = 80
                        sheet.add_image(img, 'B2')

                        centro = str(NombreCentro(cc))
                        if cc == '0':
                            centro = "TODOS LOS CENTROS"
                        
                        sheet['E2'] = "HORAS EXTRAS - " + centro
                        sheet['E2'].font = Font(bold=True)
                        sheet['H2'] = fecha_hora_actual_texto()
                        sheet['H4'].font = Font(bold=True)
                        sheet['H4'] = "USUARIO: " + usuario.upper()

                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename=datos.xlsx'

                        workbook.save(response)

                        return response
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron horas.'})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("EMPAQUE","LISTA HORAS PROCESADAS",str(request.user),error)
                return JsonResponse({'Message': 'Error', 'Nota': error})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})   
    
@login_required
@csrf_exempt
def restauraHoraL(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_restaurar')
        if user_has_permission:
            usuario = str(request.user)
            Listado_ID_HEP = request.POST.getlist('idCheck')
            index = 0
            for ID_HEP in Listado_ID_HEP:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """ 
                                UPDATE HorasExtras_Procesadas
                                SET EstadoEnvia = CASE
                                                WHEN ID_LTFP IS NULL THEN '3'
                                                ELSE '4'
                                            END
                                WHERE ID_HEP = %s
                            """
                        cursor.execute(sql,[ID_HEP])
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("FRIO","ACTUALIZA ESTADO HORAS",usuario.upper(),error)
                    return JsonResponse ({'Message': 'Not Found', 'Nota': 'Se produjo un error al intentar resolver la petición.'}) 
                index = index + 1

            if len(Listado_ID_HEP) == index:
                return JsonResponse ({'Message': 'Success', 'Nota': 'Los items se restauraron correctamente.'}) 
            else:
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Se produjo un error al intentar restaurar algunos items.'}) 
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})   
    






@login_required
@csrf_exempt
def mostrar_pedidos_flete(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            Tipo = request.POST.get('Tipo')
            Estado = request.POST.get('Estado')
            values = [Tipo,Estado]
            try:
                with connections['S3A'].cursor() as cursor:
                    sql =   """ 
                            DECLARE @@Tipo VARCHAR(5);
                            DECLARE @@Estado VARCHAR(2);
                            SET @@Tipo = %s;
                            SET @@Estado = %s;

                            SELECT PF.IdPedidoFlete AS ID_PF, CASE WHEN PF.TipoDestino = 'P' THEN 'PLANTA' WHEN PF.TipoDestino = 'U' THEN 'CAMBIO DOM.' END AS TIPO, CONVERT(VARCHAR(20),RTRIM(PF.Solicitante)) AS SOLICITA,  
                                CASE WHEN RTRIM(PF.TipoCarga) = 'RAU' THEN 'COSECHA' WHEN RTRIM(PF.TipoCarga) = 'VAC' THEN 'VACÍOS' WHEN RTRIM(PF.TipoCarga) = 'FBI' THEN 'FRUTA EN BINS' WHEN RTRIM(PF.TipoCarga) = 'VAR' THEN 'VARIOS'
                                WHEN RTRIM(PF.TipoCarga) = 'MAT' THEN 'MATERIALES' WHEN RTRIM(PF.TipoCarga) = 'EMB' THEN 'EMBALADO' ELSE RTRIM(PF.TipoCarga) END AS TIPO, CONVERT(VARCHAR(10), PF.FechaPedido, 103) AS FECHA_PEDIDO,
                                CONVERT(VARCHAR(5), PF.HoraPedido, 108) AS HORA_PEDIDO, CONVERT(VARCHAR(10), PF.FechaRequerida, 103) AS FECHA_REQUERIDO, CASE WHEN CONVERT(VARCHAR(5), PF.HoraRequerida, 108) IS NULL THEN '--:--' ELSE CONVERT(VARCHAR(5), 
                                PF.HoraRequerida, 108) END AS HORA_REQUERIDO, CASE WHEN RTRIM(ZN.Nombre) IS NULL THEN '0' ELSE RTRIM(ZN.Nombre) END AS ZONA, CASE WHEN RTRIM(CH.Nombre) IS NULL THEN '0' ELSE RTRIM(CH.Nombre) END AS DESTINO,
                                CASE WHEN PF.Bins IS NULL THEN '-' ELSE PF.Bins END AS BINS, CASE WHEN RTRIM(ES.Nombre) IS NULL THEN '-' ELSE RTRIM(ES.Nombre) END AS ESPECIE, CASE WHEN RTRIM(VR.Nombre) IS NULL THEN '0' ELSE RTRIM(VR.Nombre) END AS VARIEDAD,
                                CASE WHEN RTRIM(PF.Vacios) = 'N' THEN 'NO' WHEN RTRIM(PF.Vacios) = 'S' THEN 'SI' ELSE '' END AS VACIOS, ISNULL(PF.CantVacios,0) AS CANT_VACIOS, CASE WHEN RTRIM(PF.Cuellos) = 'N' THEN 'NO' WHEN RTRIM(PF.Cuellos) = 'S' THEN 'SI' ELSE '' END AS CUELLOS,
                                CASE WHEN RTRIM(PF.Obs) IS NULL THEN '' ELSE RTRIM(PF.Obs) END AS OBSERVACIONES, ISNULL(PF.IdTransportista,0) AS ID_TRANSPORTISTA, COALESCE(CONVERT(VARCHAR(15), RTRIM(TR.RazonSocial)),'') AS NOMBRE_TRANSPORTISTA, ISNULL(PF.IdCamion,0) AS ID_CAMION,
                                COALESCE(RTRIM(CM.Nombre),'') AS NOMBRE_CAMION, ISNULL(PF.IdAcoplado,0) AS ID_ACOPLADO, COALESCE(RTRIM(AC.Nombre),'') AS NOMBRE_ACOPLADO, ISNULL(PF.IdChofer,0) AS ID_CHOFER, COALESCE(RTRIM(Chofer),'') AS NOMBRE_CHOFER,
                                COALESCE(RTRIM(UB1.Descripcion),'0') AS DESTINO, COALESCE(RTRIM(UB2.Descripcion),'0') AS ORIGEN, CASE WHEN (SELECT Estado FROM TRESASES_APLICATIVO.dbo.Chofer_Detalle_Chacras_Viajes WHERE IdPedidoFlete = PF.IdPedidoFlete AND Estado IN ('A','V')) = 'V' THEN '#008f39' WHEN
                                (SELECT Estado FROM TRESASES_APLICATIVO.dbo.Chofer_Detalle_Chacras_Viajes WHERE IdPedidoFlete = PF.IdPedidoFlete AND Estado IN ('A','V')) = 'A' THEN '#ff8000' ELSE '#000000' END AS COLOR, CASE WHEN (SELECT ID_CVN FROM TRESASES_APLICATIVO.dbo.Chofer_Detalle_Chacras_Viajes WHERE IdPedidoFlete = PF.IdPedidoFlete AND Estado IN ('A','V'))
                                IS NULL THEN '0' ELSE (SELECT ID_CVN FROM TRESASES_APLICATIVO.dbo.Chofer_Detalle_Chacras_Viajes WHERE IdPedidoFlete = PF.IdPedidoFlete AND Estado IN ('A','V')) END AS NUMERO_VIAJE, (SELECT Estado FROM TRESASES_APLICATIVO.dbo.Chofer_Detalle_Chacras_Viajes WHERE IdPedidoFlete = PF.IdPedidoFlete AND Estado IN ('A','V')) AS ESTADO
                            FROM PedidoFlete AS PF LEFT JOIN
                                    Zona AS ZN ON ZN.IdZona = PF.IdZona LEFT JOIN
                                    Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                    Especie AS ES ON ES.IdEspecie = PF.IdEspecie LEFT JOIN 
                                    Variedad AS VR ON VR.IdVariedad = PF.IdVariedad LEFT JOIN
                                    Transportista AS TR ON TR.IdTransportista = PF.IdTransportista LEFT JOIN
                                    Camion AS CM ON CM.IdCamion = PF.IdCamion LEFT JOIN
                                    Acoplado AS AC ON AC.IdAcoplado = PF.IdAcoplado LEFT JOIN
                                    Chofer AS CF ON CF.IdChofer = PF.IdChofer LEFT JOIN
                                    Ubicacion AS UB1 ON UB1.IdUbicacion = PF.IdPlantaDestino LEFT JOIN
                                    Ubicacion AS UB2 ON UB2.IdUbicacion = PF.IdPlanta
                            WHERE PF.Estado = @@Estado
                                    AND (@@Tipo = '0' OR @@Tipo = PF.TipoDestino)
                            ORDER BY NUMERO_VIAJE DESC
                            """
                    cursor.execute(sql, values)
                    consulta = cursor.fetchall()
                    if consulta:
                        data = []
                        for i in consulta:
                            ID_PF = str(i[0])
                            TIPO = str(i[1])
                            SOLICITA = str(i[2])
                            TIPO_DESTINO = str(i[3])
                            FECHA_PEDIDO = str(i[4])
                            HORA_PEDIDO = str(i[5])
                            FECHA_REQUERIDO = str(i[6])
                            HORA_REQUERIDO = str(i[7])
                            ZONA = str(i[8])
                            DESTINO = str(i[9])
                            BINS = str(i[10])
                            ESPECIE = str(i[11])
                            VARIEDAD = str(i[12])
                            VACIOS = str(i[13])
                            CUELLOS = str(i[15])
                            OBSERVACIONES = str(i[16])
                            ID_TRANSPORTISTA = str(i[17])
                            NOMBRE_TRANSPORTISTA = str(i[18])
                            ID_CAMION = str(i[19])
                            NOMBRE_CAMION = str(i[20])
                            ID_ACOPLADO = str(i[21])
                            NOMBRE_ACOPLADO = str(i[22])
                            ID_CHOFER = str(i[23])
                            NOMBRE_CHOFER = str(i[24])
                            DESTINO2 = str(i[25])
                            ORIGEN = str(i[26])
                            COLOR = str(i[27])
                            ID_VIAJE = str(i[28])
                            ESTADO = str(i[29])
                            datos = {'ID':ID_PF,'Tipo':TIPO, 'Solicita':SOLICITA, 'TipoDestino':TIPO_DESTINO, 'FechaPedido':FECHA_PEDIDO, 'HoraPedido':HORA_PEDIDO, 'FechaRequerido':FECHA_REQUERIDO, 'HoraRequerido':HORA_REQUERIDO, 
                                     'Zona':ZONA , 'Destino': DESTINO, 'Bins': BINS, 'Especie':ESPECIE, 'Variedad':VARIEDAD, 'Vacios':VACIOS, 'Cuellos':CUELLOS, 'Obs':OBSERVACIONES, 'IdTransportista':ID_TRANSPORTISTA,
                                     'Transportista':NOMBRE_TRANSPORTISTA, 'IdCamion':ID_CAMION, 'Camion':NOMBRE_CAMION, 'IdAcoplado':ID_ACOPLADO, 'Acoplado':NOMBRE_ACOPLADO, 'IdChofer':ID_CHOFER, 'Chofer':NOMBRE_CHOFER,
                                     'Origen':ORIGEN, 'Destino2':DESTINO2, 'Color':COLOR, 'IdViaje':ID_VIAJE, 'Estado':ESTADO}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No existen Pedidos de Flete Asignados y/o Pendientes."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS FLETES",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['S3A'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def mostrar_choferes(request):
    if request.method == 'GET':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        usuario = str(request.user)
        if user_has_permission:
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """ 
                            SELECT CA.ID_CA AS ID, CA.NombreChofer AS CHOFER, CA.NombreTransporte AS TRANSPORTE,
                                    CA.NombreCamion AS CAMION, CVN.ID_CVN AS ID_VIAJE
                            FROM Chofer_Alta AS CA LEFT JOIN
                                    Chofer_Viajes_Notificacion AS CVN ON CVN.ID_CA = CA.ID_CA
                            WHERE CA.Estado = 'A' AND CVN.Estado = 'V'
                        """
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            ID_CA = str(row[0])
                            CHOFER = str(row[1])
                            TRASNPORTE = str(row[2])
                            CAMION = str(row[3])
                            ID_VIAJE = str(row[4])
                            datos = {'IdCA':ID_CA, 'Chofer':CHOFER, 'Transporte':TRASNPORTE, 'Camion':CAMION, 'IdViaje':ID_VIAJE}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No existen Choferes en viaje."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR CHOFERES",usuario,data)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})   
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def mapeo_Ultima_Ubicacion(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            IdChofer = request.POST.get('Chofer')
            usuario = str(request.user)
            values = [IdChofer]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """ 
                            SELECT TOP 1 ISNULL(CDVC.ID_CDVC,0) AS ID_CDVC, CASE WHEN CDVC.Latitud IS NULL THEN '-38.942632' ELSE CDVC.Latitud END AS LATITUD, 
                                        CASE WHEN CDVC.Longitud IS NULL THEN '-68.003042' ELSE CDVC.Longitud END AS LONGITUD, CA.NombreChofer AS CHOFER, 
                                CASE WHEN 
                                    CASE 
                                        WHEN DATEDIFF(HOUR, CDVC.FechaAlta, GETDATE()) = 0 THEN 
                                            'Hace: ' + CONVERT(VARCHAR, (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) - (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) / 60) * 60)) + ' Minutos.'
                                        ELSE 
                                            'Hace: ' + CONVERT(VARCHAR, DATEDIFF(HOUR, CDVC.FechaAlta, GETDATE())) + ' Horas y ' + 
                                            CONVERT(VARCHAR, (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) - (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) / 60) * 60)) + ' Minutos.'
                                    END IS NULL
                                THEN 'Hace 0 Minutos'
                                ELSE
                                    CASE 
                                        WHEN DATEDIFF(HOUR, CDVC.FechaAlta, GETDATE()) = 0 THEN 
                                            'Hace: ' + CONVERT(VARCHAR, (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) - (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) / 60) * 60)) + ' Minutos.'
                                        ELSE 
                                            'Hace: ' + CONVERT(VARCHAR, DATEDIFF(HOUR, CDVC.FechaAlta, GETDATE())) + ' Horas y ' + 
                                            CONVERT(VARCHAR, (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) - (DATEDIFF(MINUTE, CDVC.FechaAlta, GETDATE()) / 60) * 60)) + ' Minutos.'
                                    END
                                END AS HACE 
                            FROM Chofer_Alta AS CA LEFT JOIN
                                Chofer_Viajes_Notificacion AS CVN ON CVN.ID_CA = CA.ID_CA LEFT JOIN 
                                Chofer_Detalle_Viajes_Coordenadas AS CDVC ON CDVC.ID_CVN = CVN.ID_CVN
                            WHERE CA.ID_CA = %s AND CVN.Estado = 'V'
                            ORDER BY CDVC.ID_CDVC DESC  
                        """
                    cursor.execute(sql, values)
                    results = cursor.fetchone()
                    if results:
                        ID_CDVC = str(results[0])
                        LATITUD = str(results[1])
                        LONGITUD = str(results[2])
                        CHOFER = str(results[3])
                        HACE = str(results[4])
                        datos = {'IdCdvc':ID_CDVC,'Latitud':LATITUD,'Longitud':LONGITUD,'Chofer':CHOFER,'Hace':HACE}
                        return JsonResponse({'Message': 'Success', 'Datos': datos})
                    else:
                        data = "No existen Coordenadas del viaje."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                insertar_registro_error_sql("LOGISTICA","CREAR UBICACION",usuario,data)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})   
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def mostrar_detalles_pedido_flete(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            IdPedidoFlete = request.POST.get('IdPedidoFlete')
            values = [IdPedidoFlete]
            try:
                with connections['S3A'].cursor() as cursor:
                    sql =   """ 
                            DECLARE @@IdPedidoFlete INT;
                            SET @@IdPedidoFlete = %s;

                            SELECT PF.IdPedidoFlete AS ID_PF, CASE WHEN PF.TipoDestino = 'P' THEN 'PLANTA' WHEN PF.TipoDestino = 'U' THEN 'CAMBIO DOM.' END AS TIPO, CONVERT(VARCHAR(20), RTRIM(PF.Solicitante)) AS SOLICITA,  
                                    CASE WHEN RTRIM(PF.TipoCarga) = 'RAU' THEN 'COSECHA' WHEN RTRIM(PF.TipoCarga) = 'VAC' THEN 'VACÍOS' WHEN RTRIM(PF.TipoCarga) = 'FBI' THEN 'FRUTA EN BINS' WHEN RTRIM(PF.TipoCarga) = 'VAR' THEN 'VARIOS'
                                    WHEN RTRIM(PF.TipoCarga) = 'MAT' THEN 'MATERIALES' WHEN RTRIM(PF.TipoCarga) = 'EMB' THEN 'EMBALADO' ELSE RTRIM(PF.TipoCarga) END AS TIPO, CONVERT(VARCHAR(10), PF.FechaPedido, 103) AS FECHA_PEDIDO,
                                    CONVERT(VARCHAR(5), PF.HoraPedido, 108) AS HORA_PEDIDO, CONVERT(VARCHAR(10), PF.FechaRequerida, 103) AS FECHA_REQUERIDO, CASE WHEN CONVERT(VARCHAR(5), PF.HoraRequerida, 108) IS NULL THEN '--:--' ELSE CONVERT(VARCHAR(5), 
                                    PF.HoraRequerida, 108) END AS HORA_REQUERIDO, CASE WHEN RTRIM(ZN.Nombre) IS NULL THEN '0' ELSE RTRIM(ZN.Nombre) END AS ZONA, CASE WHEN PF.TipoDestino = 'U' THEN COALESCE(RTRIM(UB1.Descripcion),'0') ELSE RTRIM(CH.Nombre) END AS DESTINO,
                                    CASE WHEN PF.Bins IS NULL THEN '-' ELSE PF.Bins END AS BINS, CASE WHEN RTRIM(ES.Nombre) IS NULL THEN '-' ELSE RTRIM(ES.Nombre) END AS ESPECIE, CASE WHEN RTRIM(VR.Nombre) IS NULL THEN '-' ELSE RTRIM(VR.Nombre) END AS VARIEDAD,
                                    CASE WHEN RTRIM(PF.Vacios) IS NULL THEN '' WHEN RTRIM(PF.Vacios) = 'N' THEN 'NO' WHEN RTRIM(PF.Vacios) = 'S' THEN 'SI' END AS VACIOS, ISNULL(PF.CantVacios,0) AS CANT_VACIOS, CASE WHEN RTRIM(PF.Cuellos) IS NULL THEN '' WHEN RTRIM(PF.Cuellos) = 'N' THEN 'NO' WHEN RTRIM(PF.Cuellos) = 'S' THEN 'SI' END AS CUELLOS,
                                    CASE WHEN RTRIM(PF.Obs) IS NULL THEN '' ELSE UPPER(LTRIM(RTRIM(PF.Obs))) END AS OBSERVACIONES, ISNULL(PF.IdTransportista,0) AS ID_TRANSPORTISTA, COALESCE(CONVERT(VARCHAR(15), RTRIM(TR.RazonSocial)),'') AS NOMBRE_TRANSPORTISTA, ISNULL(PF.IdCamion,0) AS ID_CAMION,
                                    COALESCE(RTRIM(CM.Nombre),'') AS NOMBRE_CAMION, ISNULL(PF.IdAcoplado,0) AS ID_ACOPLADO, COALESCE(RTRIM(AC.Nombre),'') AS NOMBRE_ACOPLADO, ISNULL(PF.IdChofer,0) AS ID_CHOFER, COALESCE(RTRIM(Chofer),'') AS NOMBRE_CHOFER,
                                    COALESCE(RTRIM(UB1.Descripcion),'0') AS DESTINO, COALESCE(RTRIM(UB2.Descripcion),'0') AS ORIGEN
                            FROM PedidoFlete AS PF LEFT JOIN
                                    Zona AS ZN ON ZN.IdZona = PF.IdZona LEFT JOIN
                                    Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                    Especie AS ES ON ES.IdEspecie = PF.IdEspecie LEFT JOIN 
                                    Variedad AS VR ON VR.IdVariedad = PF.IdVariedad LEFT JOIN
                                    Transportista AS TR ON TR.IdTransportista = PF.IdTransportista LEFT JOIN
                                    Camion AS CM ON CM.IdCamion = PF.IdCamion LEFT JOIN
                                    Acoplado AS AC ON AC.IdAcoplado = PF.IdAcoplado LEFT JOIN
                                    Chofer AS CF ON CF.IdChofer = PF.IdChofer LEFT JOIN
                                    Ubicacion AS UB1 ON UB1.IdUbicacion = PF.IdPlantaDestino LEFT JOIN
                                    Ubicacion AS UB2 ON UB2.IdUbicacion = PF.IdPlanta
                            WHERE PF.IdPedidoFlete = @@IdPedidoFlete
                            """
                    cursor.execute(sql, values)
                    consulta = cursor.fetchone()
                    if consulta:
                        data = []
                        ID_PF = str(consulta[0])
                        TIPO = str(consulta[1])
                        SOLICITA = str(consulta[2])
                        TIPO_DESTINO = str(consulta[3])
                        FECHA_PEDIDO = str(consulta[4])
                        HORA_PEDIDO = str(consulta[5])
                        FECHA_REQUERIDO = str(consulta[6])
                        HORA_REQUERIDO = str(consulta[7])
                        ZONA = str(consulta[8])
                        DESTINO = str(consulta[9])
                        BINS = str(consulta[10])
                        ESPECIE = str(consulta[11])
                        VARIEDAD = str(consulta[12])
                        VACIOS = str(consulta[13]) + ' - ' + str(consulta[14])
                        CUELLOS = str(consulta[15])
                        OBSERVACIONES = str(consulta[16])
                        ID_TRANSPORTISTA = str(consulta[17])
                        NOMBRE_TRANSPORTISTA = str(consulta[18])
                        ID_CAMION = str(consulta[19])
                        NOMBRE_CAMION = str(consulta[20])
                        ID_ACOPLADO = str(consulta[21])
                        NOMBRE_ACOPLADO = str(consulta[22])
                        ID_CHOFER = str(consulta[23])
                        NOMBRE_CHOFER = str(consulta[24])
                        DESTINO2 = str(consulta[25])
                        ORIGEN = str(consulta[26])
                        datos = {'ID':ID_PF,'Tipo':TIPO, 'Solicita':SOLICITA, 'TipoDestino':TIPO_DESTINO, 'FechaPedido':FECHA_PEDIDO, 'HoraPedido':HORA_PEDIDO, 'FechaRequerido':FECHA_REQUERIDO, 'HoraRequerido':HORA_REQUERIDO, 
                                    'Zona':ZONA , 'Destino': DESTINO, 'Bins': BINS, 'Especie':ESPECIE, 'Variedad':VARIEDAD, 'Vacios':VACIOS, 'Cuellos':CUELLOS, 'Obs':OBSERVACIONES, 'IdTransportista':ID_TRANSPORTISTA,
                                    'Transportista':NOMBRE_TRANSPORTISTA, 'IdCamion':ID_CAMION, 'Camion':NOMBRE_CAMION, 'IdAcoplado':ID_ACOPLADO, 'Acoplado':NOMBRE_ACOPLADO, 'IdChofer':ID_CHOFER, 'Chofer':NOMBRE_CHOFER,
                                    'Origen':ORIGEN, 'Destino2':DESTINO2}
                        data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No existen Pedidos de Flete pendientes de Asignación."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS FLETES",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['S3A'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def mostrar_detalles_viaje_rechazado(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            ID_CVN = request.POST.get('ID_CVN')
            values = [ID_CVN]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql =   """ 
                            SELECT CVN.ID_CVN AS ID_CVN, CA.NombreChofer AS CHOFER, CONVERT(VARCHAR(5), CVN.FechaAlta,103) + ' ' + CONVERT(VARCHAR(5), CVN.FechaAlta,108) + ' Hs.' AS FECHA,
                                CDCV.IdPedidoFlete AS ID_PEDIDO_FLETE, RTRIM(UB.Descripcion) AS ORIGEN, CASE WHEN RTRIM(PF.TipoDestino) = 'P' THEN RTRIM(CH.Nombre) ELSE RTRIM(UB2.Descripcion) END AS DESTINO,
	                            RTRIM(PF.Solicitante) AS SOLICITA
                            FROM Chofer_Viajes_Notificacion AS CVN LEFT JOIN
                                Chofer_Alta AS CA ON CA.ID_CA = CVN.ID_CA LEFT JOIN
                                Chofer_Detalle_Chacras_Viajes AS CDCV ON CDCV.ID_CVN = CVN.ID_CVN LEFT JOIN
                                S3A.dbo.PedidoFlete AS PF ON PF.IdPedidoFlete = CDCV.IdPedidoFlete LEFT JOIN
                                S3A.dbo.Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                S3A.dbo.Ubicacion AS UB ON UB.IdUbicacion = PF.IdPlanta LEFT JOIN
                                S3A.dbo.Ubicacion AS UB2 ON UB2.IdUbicacion = PF.IdPlantaDestino
                            WHERE CVN.ID_CVN = %s
                            """
                    cursor.execute(sql, values)
                    consulta = cursor.fetchall()
                    if consulta:
                        datos_unicos = {}
                        datos_repetidos = []
                        for row in consulta:
                            if not datos_unicos:
                                datos_unicos = {
                                    'ID_CVN': str(row[0]),
                                    'Chofer': str(row[1]),
                                    'Fecha': str(row[2])
                                }
                            
                            datos_repetidos.append({
                                'IdPedidoFlete': str(row[3]),
                                'Origen': str(row[4]),
                                'Destino': str(row[5]),
                                'Solicita': str(row[6])
                            })
                        return JsonResponse({'Message': 'Success', 'Datos': datos_unicos, 'Tabla':datos_repetidos})
                    else:
                        data = "No existen datos para el viaje rechazado."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS RECHZADOS",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def detalles_de_viajes_activos(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            ID_CA = request.POST.get('ID_CA')
            values = [ID_CA]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql =   """ 
                            DECLARE @@ID_CA INT;
                            SET @@ID_CA = %s;
                            SELECT CDCV.IdPedidoFlete AS ID_PEDIDO, RTRIM(UB.Descripcion) AS ORIGEN, CASE WHEN PF.TipoDestino = 'U' THEN RTRIM(UB2.Descripcion) ELSE RTRIM(CH.Nombre) END AS DESTINO,
                                CASE WHEN CDCV.LlegaChacra IS NULL THEN '--/-- --:--' ELSE (CONVERT(VARCHAR(5), CDCV.LlegaChacra, 103) + ' ' + CONVERT(VARCHAR(5), CDCV.LlegaChacra, 108) + ' Hs.') END AS HORA_LLEGADA,
                                COALESCE(CUV.Nombre,'-') AS LUGAR_VACIOS, CASE WHEN CVN.LlegaVacios IS NULL THEN '--/-- --:--' ELSE (CONVERT(VARCHAR(5), CVN.LlegaVacios, 103) + ' ' + CONVERT(VARCHAR(5), CVN.LlegaVacios, 108) + ' Hs.') END AS LLEGA_VACIOS
                            FROM Chofer_Alta AS CA LEFT JOIN
                                Chofer_Viajes_Notificacion AS CVN ON CVN.ID_CA = CA.ID_CA LEFT JOIN
                                Chofer_Detalle_Chacras_Viajes AS CDCV ON CDCV.ID_CVN = CVN.ID_CVN LEFT JOIN
                                Chofer_Ubicacion_Vacios AS CUV ON CUV.ID_CUV = CVN.ID_CUV LEFT JOIN
                                S3A.dbo.PedidoFlete AS PF ON PF.IdPedidoFlete = CDCV.IdPedidoFlete LEFT JOIN
                                S3A.dbo.Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                S3A.dbo.Ubicacion AS UB ON UB.IdUbicacion = PF.IdPlanta LEFT JOIN
                                S3A.dbo.Ubicacion AS UB2 ON UB2.IdUbicacion = PF.IdPlantaDestino
                            WHERE CA.ID_CA = @@ID_CA AND CVN.Estado = 'V' --AND CDCV.Estado = 'V'
                            """
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            ID_PF = str(row[0])
                            ORIGEN = str(row[1])
                            DESTINO = str(row[2])
                            HORA_LLEGADA = str(row[3])
                            VACIOS_LUGAR = str(row[4])
                            HORA_VACIOS = str(row[5])
                            datos = {'IdPF':ID_PF,'Origen':ORIGEN,'Destino':DESTINO,'HoraLlegada':HORA_LLEGADA}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'LugarVacios': VACIOS_LUGAR, 'HoraVacios': HORA_VACIOS, 'Datos': data})
                    else:
                        data = "No existen Destinos para ese Chofer."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS FLETES",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def verifica_pedidos_completa_combox(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            listado_id_pedidos = request.POST.getlist('IdPedidoFlete')
            listado_sql = ','.join(listado_id_pedidos)
            if verifica_mismo_tipo_destino(listado_id_pedidos):                
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql =   f""" 
                                SELECT PF.IdPedidoFlete AS ID_PEDIDO, CASE WHEN PF.TipoDestino = 'P' THEN RTRIM(CH.Nombre) ELSE RTRIM(UB.Descripcion) END AS DESTINO
                                FROM S3A.dbo.PedidoFlete AS PF LEFT JOIN 
                                        S3A.dbo.Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                        S3A.dbo.Ubicacion AS UB ON UB.IdUbicacion = PF.IdPlantaDestino
                                WHERE PF.IdPedidoFlete IN ({listado_sql})
                                """
                        cursor.execute(sql)
                        results = cursor.fetchall()
                        if results:
                            destinos = []
                            for row in results:
                                ID_PF = str(row[0])
                                DESTINO = str(row[1])
                                datos = {'IdPF':ID_PF, 'Destino':DESTINO}
                                destinos.append(datos)

                        sql2 =   """ 
                                SELECT CA.IdChofer AS ID_CHOFER, CA.NombreChofer AS CHOFER, CA.IdTransporte AS ID_TRANSPORTE,
                                        CA.IdCamion AS ID_CAMION, CA.IdAcoplado AS ID_ACOPLADO
                                FROM Chofer_Alta AS CA
                                WHERE CA.Estado = 'A'
                                ORDER BY CA.NombreChofer
                                """
                        cursor.execute(sql2)
                        results2 = cursor.fetchall()
                        if results2:
                            choferes = []
                            for row in results2:
                                ID_CHOFER = str(row[0])
                                CHOFER = str(row[1])
                                ID_TRANSPORTE = str(row[2])
                                ID_CAMION = str(row[3])
                                ID_ACOPLADO = str(row[4])
                                datos2 = {'IdChofer':ID_CHOFER, 'Chofer':CHOFER, 'IdTransporte':ID_TRANSPORTE,'IdCamion':ID_CAMION,'IdAcoplado':ID_ACOPLADO}
                                choferes.append(datos2)

                        sql3 =   """ 
                                SELECT CUV.ID_CUV AS ID_VACIOS, CUV.Nombre AS NOMBRE
                                FROM Chofer_Ubicacion_Vacios AS CUV
                                WHERE CUV.Estado = 'A'
                                ORDER BY CUV.Nombre
                                """
                        cursor.execute(sql3)
                        results3 = cursor.fetchall()
                        if results3:
                            vacios = []
                            for row in results3:
                                ID_VACIOS = str(row[0])
                                NOMBRE_VACIOS = str(row[1])
                                datos3 = {'IdVacios':ID_VACIOS, 'Nombre':NOMBRE_VACIOS}
                                vacios.append(datos3)

                        sql4 =   """ 
                                SELECT IdTransportista, RTRIM(CONVERT(VARCHAR(23), RazonSocial)) AS RAZON_SOCIAL
                                FROM S3A.dbo.Transportista 
                                WHERE Activo ='S' 
                                ORDER BY RazonSocial
                                """
                        cursor.execute(sql4)
                        results4 = cursor.fetchall()
                        if results4:
                            transportes = []
                            for row in results4:
                                ID_TRANSPORTE = str(row[0])
                                NOMBRE_TRANSPORTE = str(row[1])
                                datos4 = {'IdTransporte':ID_TRANSPORTE, 'Transporte':NOMBRE_TRANSPORTE}
                                transportes.append(datos4)

                        camiones, acoplados = crea_listado_datos()

                        if destinos and choferes:
                            return JsonResponse({'Message': 'Success', 'Destinos': destinos, 'Choferes': choferes, 'Vacios':vacios, 'Transportes':transportes, 'Camiones':camiones, 'Acoplados':acoplados})
                        else:
                            data = "No existen Destinos para ese Chofer."
                            return JsonResponse({'Message': 'Error', 'Nota': data})
                        
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS FLETES",request.user,error)
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    cursor.close()
                    connections['TRESASES_APLICATIVO'].close()
            
            return JsonResponse({'Message': 'Error', 'Nota': 'La agrupación de Pedidos debe ser siempre de un mismo Tipo. (PLANTA ó CAMBIO DOM.)'})
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def verifica_mismo_tipo_destino(listado_id):
    try:
        listado_sql = ','.join(listado_id)
        with connections['S3A'].cursor() as cursor:
            sql = f"""SELECT 
                    CASE 
                        WHEN MIN(PF.TipoDestino) = MAX(PF.TipoDestino) THEN 0
                        ELSE 1
                    END AS Resultado
                    FROM PedidoFlete AS PF
                    WHERE PF.IdPedidoFlete IN ({listado_sql})"""
            cursor.execute(sql)
            results = cursor.fetchone()
            if results:
                diferentes = str(results[0])
                if diferentes == '0':
                    return True
            return False
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","VERIFICA MISMO TIPO DESTINO","usuario",error)
        return False
    finally:
        cursor.close()
        connections['S3A'].close()

def crea_listado_datos():
    listado_camion = []
    listado_acoplado = []
    listado_chofer = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """ 
                    SELECT TR.IdTransportista AS ID_TRANSPORTE, CM.IdCamion AS ID_CAMION, UPPER(RTRIM(CM.Nombre) + ' - ' + RTRIM(CM.Patente)) AS NOMBRE_CAMION
                    FROM S3A.dbo.Transportista AS TR LEFT JOIN
                            S3A.dbo.Camion AS CM ON CM.IdTransportista = TR.IdTransportista
                    WHERE TR.IdTransportista IN (SELECT IdTransportista
                                                    FROM S3A.dbo.Transportista 
                                                    WHERE Activo ='S')
                    ORDER BY TR.IdTransportista
                 """
            cursor.execute(sql)
            results = cursor.fetchall()
            if results:
                for row in results:
                    IdTransporte = str(row[0])
                    Id = str(row[1])
                    Descripcion = str(row[2])
                    transporte = next((t for t in listado_camion if t['IdTransporte'] == IdTransporte), None)
                    if transporte is None:
                        transporte = {'IdTransporte': IdTransporte, 'Items': []}
                        listado_camion.append(transporte)
                    transporte['Items'].append({'IdCamion': Id, 'Descripcion': Descripcion})

            sql = """ 
                    SELECT TR.IdTransportista AS ID_TRANSPORTE, AC.IdAcoplado AS ID_CHOFER, UPPER(RTRIM(AC.Nombre) + ' - ' + RTRIM(AC.Patente)) AS NOMBRE_ACOPLADO
                    FROM S3A.dbo.Transportista AS TR LEFT JOIN
                            S3A.dbo.Acoplado AS AC ON AC.IdTransportista = TR.IdTransportista
                    WHERE TR.IdTransportista IN (SELECT IdTransportista
                                                    FROM S3A.dbo.Transportista 
                                                    WHERE Activo ='S')
                    ORDER BY TR.IdTransportista
                 """
            cursor.execute(sql)
            results = cursor.fetchall()
            if results:
                for row in results:
                    IdTransporte = str(row[0])
                    Id = str(row[1])
                    Descripcion = str(row[2])
                    transporte = next((t for t in listado_acoplado if t['IdTransporte'] == IdTransporte), None)
                    if transporte is None:
                        transporte = {'IdTransporte': IdTransporte, 'Items': []}
                        listado_acoplado.append(transporte)
                    transporte['Items'].append({'IdAcoplado': Id, 'Descripcion': Descripcion})
        return listado_camion, listado_acoplado
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","VERIFICA MISMO TIPO DESTINO","usuario",error)
        return listado_camion, listado_acoplado
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def multiple_asignacion(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            ID_CA = request.POST.get('ID_CA')
            values = [ID_CA]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql =   """ 
                            DECLARE @@ID_CA INT;
                            SET @@ID_CA = %s;
                            SELECT CDCV.IdPedidoFlete AS ID_PEDIDO, RTRIM(UB.Descripcion) AS ORIGEN, CASE WHEN PF.TipoDestino = 'U' THEN RTRIM(UB2.Descripcion) ELSE RTRIM(CH.Nombre) END AS DESTINO,
                                CASE WHEN CDCV.LlegaChacra IS NULL THEN '--/-- --:--' ELSE (CONVERT(VARCHAR(5), CDCV.LlegaChacra, 103) + ' ' + CONVERT(VARCHAR(5), CDCV.LlegaChacra, 108) + ' Hs.') END AS HORA_LLEGADA,
                                COALESCE(CUV.Nombre,'-') AS LUGAR_VACIOS, CASE WHEN CVN.LlegaVacios IS NULL THEN '--/-- --:--' ELSE (CONVERT(VARCHAR(5), CVN.LlegaVacios, 103) + ' ' + CONVERT(VARCHAR(5), CVN.LlegaVacios, 108) + ' Hs.') END AS LLEGA_VACIOS
                            FROM Chofer_Alta AS CA LEFT JOIN
                                Chofer_Viajes_Notificacion AS CVN ON CVN.ID_CA = CA.ID_CA LEFT JOIN
                                Chofer_Detalle_Chacras_Viajes AS CDCV ON CDCV.ID_CVN = CVN.ID_CVN LEFT JOIN
                                Chofer_Ubicacion_Vacios AS CUV ON CUV.ID_CUV = CVN.ID_CUV LEFT JOIN
                                S3A.dbo.PedidoFlete AS PF ON PF.IdPedidoFlete = CDCV.IdPedidoFlete LEFT JOIN
                                S3A.dbo.Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                S3A.dbo.Ubicacion AS UB ON UB.IdUbicacion = PF.IdPlanta LEFT JOIN
                                S3A.dbo.Ubicacion AS UB2 ON UB2.IdUbicacion = PF.IdPlantaDestino
                            WHERE CA.ID_CA = @@ID_CA AND CVN.Estado = 'V' AND CDCV.Estado = 'V'
                            """
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            ID_PF = str(row[0])
                            ORIGEN = str(row[1])
                            DESTINO = str(row[2])
                            HORA_LLEGADA = str(row[3])
                            VACIOS_LUGAR = str(row[4])
                            HORA_VACIOS = str(row[5])
                            datos = {'IdPF':ID_PF,'Origen':ORIGEN,'Destino':DESTINO,'HoraLlegada':HORA_LLEGADA}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'LugarVacios': VACIOS_LUGAR, 'HoraVacios': HORA_VACIOS, 'Datos': data})
                    else:
                        data = "No existen Destinos para ese Chofer."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR PEDIDOS FLETES",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def asiganciones_multiples(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_insertar')
        if user_has_permission:
            Usuario = str(request.user).upper()
            listado_id_pedidos = request.POST.getlist('IdPedidosFletes')
            IdChofer = request.POST.get('IdChofer')
            nombreChofer = request.POST.get('NombreChofer')
            IdTransporte = request.POST.get('IdTransporte')
            IdCamion = request.POST.get('IdCamion')
            IdAcoplado = request.POST.get('IdAcoplado') or None
            CantVacios = request.POST.get('CantVacios') or None
            IdUbiVacios = request.POST.get('IdUbiVacios') or None
            listado_sql = ','.join(listado_id_pedidos)
            valuesPedidoFlete = [IdTransporte,IdCamion,IdAcoplado,nombreChofer,IdChofer]
            valuesNotificacion = [IdChofer,IdUbiVacios,CantVacios,Usuario]
            try:
                with transaction.atomic():  
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql_update = f"""
                            UPDATE S3A.dbo.PedidoFlete
                            SET Prioridad = NULL, IdTransportista = %s, IdCamion = %s, IdAcoplado = %s, Chofer = %s, 
                            Estado = 'A', IdChofer = %s
                            WHERE IdPedidoFlete IN ({listado_sql})
                        """
                        cursor.execute(sql_update, valuesPedidoFlete)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows = cursor.fetchone()[0]

                        if affected_rows == 0:
                            raise Exception("No se actualizaron los datos de PedidosFletes.")

                        sql_insert = """
                            INSERT INTO Chofer_Viajes_Notificacion (ID_CA, EstadoNotificacion, ID_CUV, CantidadVac, Estado, FechaAlta, UserAlta)
                            OUTPUT INSERTED.ID_CVN
                            VALUES ((SELECT ID_CA FROM Chofer_Alta WHERE IdChofer = %s), 'P', %s, %s, 'A', GETDATE(), %s)
                        """
                        cursor.execute(sql_insert, valuesNotificacion)
                        ID_CVN = cursor.fetchone()[0]

                        if ID_CVN is None:
                            raise Exception("No se generó ID_CVN en Chofer_Viajes_Notificacion.")
                        
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]
                        if affected_rows2 == 0:
                            raise Exception("No se insertaron filas en Chofer_Viajes_Notificacion.")

                        for IdPedidoFlete in listado_id_pedidos:
                            sql_detalle = f"""
                                DECLARE @@ID_CVN INT;
                                DECLARE @@IdPedidoFlete INT;
                                SET @@ID_CVN = %s;
                                SET @@IdPedidoFlete = %s;

                                IF (SELECT DISTINCT PF.TipoDestino
                                    FROM S3A.dbo.PedidoFlete AS PF
                                    WHERE PF.IdPedidoFlete IN ({listado_sql})) = 'P'
                                BEGIN
                                    INSERT INTO Chofer_Detalle_Chacras_Viajes (ID_CVN, IdPedidoFlete, IdChacra, FechaAlta, Estado)
                                    VALUES (@@ID_CVN, @@IdPedidoFlete, (SELECT CASE WHEN IdChacra IS NULL THEN '0' ELSE IdChacra END 
                                    FROM S3A.dbo.PedidoFlete WHERE IdPedidoFlete = @@IdPedidoFlete), GETDATE(), 'A')
                                END
                                ELSE
                                BEGIN
                                    INSERT INTO Chofer_Detalle_Chacras_Viajes (ID_CVN, IdPedidoFlete, IdChacra, FechaAlta, Estado)
                                    VALUES (@@ID_CVN, @@IdPedidoFlete, (SELECT CASE WHEN IdPlantaDestino IS NULL THEN '0' ELSE IdPlantaDestino END 
                                    FROM S3A.dbo.PedidoFlete WHERE IdPedidoFlete = @@IdPedidoFlete), GETDATE(), 'A')
                                END
                            """
                            cursor.execute(sql_detalle, [ID_CVN, IdPedidoFlete])
                            cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                            affected_rows3 = cursor.fetchone()[0]
                            if affected_rows3 == 0:
                                raise Exception(f"No se insertó detalle para IdPedidoFlete {IdPedidoFlete}.")
                        if Existe_Viaje_Antes(IdChofer,ID_CVN):
                            Token = Obtener_Token(IdChofer)
                            if Token != '0':
                                notificaciones_Fruit_Truck(Token,"TE ASIGNARON UN NUEVO VIAJE! Entrá a la Aplicación para ver los destinos.","V","A",str(ID_CVN),"N")
                    return JsonResponse({'Message': 'Success', 'Nota': 'El viaje se creó correctamente.'})
            except Exception as e:
                return JsonResponse({'Message': 'Error', 'Nota': 'ERROR: ' + str(e)})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def asiganciones_individuales(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_insertar')
        if user_has_permission:
            Usuario = str(request.user).upper()
            IdPedidoFlete = request.POST.get('IdPedidoFlete')
            IdChofer = request.POST.get('IdChofer')
            nombreChofer = request.POST.get('NombreChofer')
            IdTransporte = request.POST.get('IdTransporte')
            IdCamion = request.POST.get('IdCamion')
            IdAcoplado = request.POST.get('IdAcoplado') or None
            CantVacios = request.POST.get('CantVacios') or None
            IdUbiVacios = request.POST.get('IdUbiVacios') or None
            valuesPedidoFlete = [IdTransporte,IdCamion,IdAcoplado,nombreChofer,IdChofer,IdPedidoFlete]
            valuesNotificacion = [IdChofer,IdUbiVacios,CantVacios,Usuario]
            try:
                with transaction.atomic():  
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql_update = """
                            UPDATE S3A.dbo.PedidoFlete
                            SET Prioridad = NULL, IdTransportista = %s, IdCamion = %s, IdAcoplado = %s, Chofer = %s, 
                            Estado = 'A', IdChofer = %s
                            WHERE IdPedidoFlete = %s
                        """
                        cursor.execute(sql_update, valuesPedidoFlete)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows = cursor.fetchone()[0]

                        if affected_rows == 0:
                            raise Exception("No se actualizaron los datos de PedidosFletes.")

                        sql_insert = """
                            INSERT INTO Chofer_Viajes_Notificacion (ID_CA, EstadoNotificacion, ID_CUV, CantidadVac, Estado, FechaAlta, UserAlta)
                            OUTPUT INSERTED.ID_CVN
                            VALUES ((SELECT ID_CA FROM Chofer_Alta WHERE IdChofer = %s), 'P', %s, %s, 'A', GETDATE(), %s)
                        """
                        cursor.execute(sql_insert, valuesNotificacion)
                        ID_CVN = cursor.fetchone()[0]
                        
                        if ID_CVN is None:
                            raise Exception("No se generó ID_CVN en Chofer_Viajes_Notificacion.")
                        
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]
                        if affected_rows2 == 0:
                            raise Exception("No se insertaron filas en Chofer_Viajes_Notificacion.")

                        
                        sql_detalle = """
                            DECLARE @@ID_CVN INT;
                            DECLARE @@IdPedidoFlete INT;
                            SET @@ID_CVN = %s;
                            SET @@IdPedidoFlete = %s;

                            IF (SELECT DISTINCT PF.TipoDestino
                                FROM S3A.dbo.PedidoFlete AS PF
                                WHERE PF.IdPedidoFlete = @@IdPedidoFlete) = 'P'
                            BEGIN
                                INSERT INTO Chofer_Detalle_Chacras_Viajes (ID_CVN, IdPedidoFlete, IdChacra, FechaAlta, Estado)
                                VALUES (@@ID_CVN, @@IdPedidoFlete, (SELECT CASE WHEN IdChacra IS NULL THEN '0' ELSE IdChacra END 
                                FROM S3A.dbo.PedidoFlete WHERE IdPedidoFlete = @@IdPedidoFlete), GETDATE(), 'A')
                            END
                            ELSE
                            BEGIN
                                INSERT INTO Chofer_Detalle_Chacras_Viajes (ID_CVN, IdPedidoFlete, IdChacra, FechaAlta, Estado)
                                VALUES (@@ID_CVN, @@IdPedidoFlete, (SELECT CASE WHEN IdPlantaDestino IS NULL THEN '0' ELSE IdPlantaDestino END 
                                FROM S3A.dbo.PedidoFlete WHERE IdPedidoFlete = @@IdPedidoFlete), GETDATE(), 'A')
                            END
                        """
                        cursor.execute(sql_detalle, [ID_CVN, IdPedidoFlete])
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows3 = cursor.fetchone()[0]
                        if affected_rows3 == 0:
                            raise Exception(f"No se insertó detalle para IdPedidoFlete {IdPedidoFlete}.")
                        if Existe_Viaje_Antes(IdChofer,ID_CVN):
                            Token = Obtener_Token(IdChofer)
                            if Token != '0':
                                notificaciones_Fruit_Truck(Token,"TE ASIGNARON UN NUEVO VIAJE! Entrá a la Aplicación para ver los destinos.","V","A",str(ID_CVN),"N")  #notificaciones_Fruit_Truck(Token, Body, Tipo, Habilitado, ID_CVN, Destinos):
                    return JsonResponse({'Message': 'Success', 'Nota': 'El viaje se creó correctamente.'})
            except Exception as e:
                return JsonResponse({'Message': 'Error', 'Nota': 'ERROR: ' + str(e)})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def Obtener_Token(ID_CA):
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """
                    SELECT COALESCE((SELECT IdFirebase FROM Chofer_Alta WHERE IdChofer = %s), '0') AS ID_FIREBASE
                    """
            cursor.execute(sql,[ID_CA])
            results = cursor.fetchone()
            if results:
                token = str(results[0])
                return token
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","OBTENER EL TOKEN","usuario",error)
        return '0'
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

def Existe_Viaje_Antes(IdChofer,ID_CVN):
    values = [ID_CVN,IdChofer]
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """

                    DECLARE @@ID_CVN INT;
                    DECLARE @@IdChofer INT;
                    SET @@ID_CVN = %s;
                    SET @@IdChofer = %s;
                    SELECT CASE 
                        WHEN EXISTS (
                            SELECT 1 
                            FROM Chofer_Viajes_Notificacion 
                            WHERE ID_CA = (SELECT ID_CA FROM Chofer_Alta WHERE IdChofer = @@IdChofer) 
                                AND (Estado = 'V' OR (ID_CVN < @@ID_CVN AND Estado = 'A'))
                        ) THEN 1 
                        ELSE 0 
                    END AS EXISTE_REGISTRO;
                    """
            cursor.execute(sql, values)
            results = cursor.fetchone()
            if results:
                existe = str(results[0])
                if existe == '0':
                    return True
                else:
                    return False
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","VERIFICA AL ENVÍAR NOTIFICACION","usuario",error)
        return True
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()
   
@login_required
@csrf_exempt
def mostrar_viajes_rechazados(request):
    if request.method == 'GET':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        usuario = str(request.user)
        if user_has_permission:
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """ 
                            SELECT CVN.ID_CVN AS ID_CVN, CA.NombreChofer AS CHOFER, CONVERT(VARCHAR(5), CVN.FechaAlta,103) + ' ' + CONVERT(VARCHAR(5), CVN.FechaAlta,108) + ' Hs.' AS FECHA,
                                    (SELECT COUNT(*) FROM Chofer_Detalle_Chacras_Viajes WHERE ID_CVN = CVN.ID_CVN AND Estado = 'R') AS CANTIDAD
                            FROM Chofer_Viajes_Notificacion AS CVN INNER JOIN
                                Chofer_Alta AS CA ON CA.ID_CA = CVN.ID_CA
                            WHERE CVN.Estado = 'R' AND CVN.FechaAlta >= DATEADD(day, -4, GETDATE());
                        """
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            ID_CVN = str(row[0])
                            CHOFER = str(row[1])
                            FECHA = str(row[2])
                            CANTIDAD = str(row[3])
                            datos = {'ID_CVN':ID_CVN, 'Chofer':CHOFER, 'Fecha':FECHA, 'Cantidad':CANTIDAD}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No existen Choferes en viaje."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR VIAJE RECHAZADOS",usuario,data)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})   
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def llevar_pendientes_rechazados(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_modificar')
        if user_has_permission:
            ID_CVN = request.POST.get('ID_CVN')
            values = [ID_CVN]
            try:
                 with transaction.atomic():
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """
                            DECLARE @ID_CVN INT;
                            SET @ID_CVN = %s; 

                            UPDATE S3A.dbo.PedidoFlete
                            SET Estado = 'P', IdTransportista = NULL, IdCamion = NULL, IdAcoplado = NULL, IdChofer = NULL, Chofer = NULL
                            WHERE IdPedidoFlete IN (
                                SELECT CDCV.IdPedidoFlete
                                FROM Chofer_Detalle_Chacras_Viajes AS CDCV
                                WHERE CDCV.ID_CVN = @ID_CVN AND CDCV.Estado = 'R'
                            );
                        """
                        cursor.execute(sql, values)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]
                        if affected_rows2 == 0:
                            raise Exception("No se actualizaron los Estados de PF.")

                        sql = """
                            DECLARE @ID_CVN INT;
                            SET @ID_CVN = %s; 

                            UPDATE Chofer_Detalle_Chacras_Viajes
                            SET Estado = 'E'
                            WHERE IdPedidoFlete IN (
                                SELECT CDCV.IdPedidoFlete
                                FROM Chofer_Detalle_Chacras_Viajes AS CDCV
                                WHERE CDCV.ID_CVN = @ID_CVN);
                        """
                        cursor.execute(sql, values)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows = cursor.fetchone()[0]
                        if affected_rows == 0:
                            raise Exception("No se actualizaron los Estados de CDCV.")
                        
                        sql = """
                            DECLARE @ID_CVN INT;
                            SET @ID_CVN = %s; 
                            
                            UPDATE Chofer_Viajes_Notificacion
                            SET Estado = 'E'
                            WHERE ID_CVN = @ID_CVN;
                        """
                        cursor.execute(sql, values)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]
                        if affected_rows2 == 0:
                            raise Exception("No se actualizaron los Estados de CVN.")

                    return JsonResponse({'Message': 'Success', 'Nota': 'Los pedidos se movieron a PENDIENTES correctamente.'})
            except Exception as e:
                error = str(e)
                print(error)
                insertar_registro_error_sql("LOGISTICA","MOVER PEDIDOS RECHZADOS",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
############ UBICACION DE LOS CHOFERES



@login_required
@csrf_exempt
def listado_choferes_activos(request):
    if request.method == 'GET':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        usuario = str(request.user)
        if user_has_permission:
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """ 
                            SELECT ID_CA AS ID_CA, NombreChofer AS CHOFER
                            FROM Chofer_Alta
                            WHERE Estado = 'A'
                            ORDER BY CHOFER
                        """
                    cursor.execute(sql)
                    results = cursor.fetchall()
                    if results:
                        data = [{'IdCA':'0', 'Chofer':'TODOS'}]
                        for row in results:
                            ID_CA = str(row[0])
                            CHOFER = str(row[1])
                            datos = {'IdCA':ID_CA, 'Chofer':CHOFER}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Choferes': data})
                    else:
                        data = "No existen Choferes habilitados."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                insertar_registro_error_sql("LOGISTICA","MOSTRAR UBICACION CHOFERES",usuario,data)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})   
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def detalle_ultima_ubicacion_choferes(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_ver')
        if user_has_permission:
            ID_CA = request.POST.get('ID_CA')
            values = [ID_CA]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql =   """ 
                            DECLARE @@ID_CA INT;
                            SET @@ID_CA = %s;
                            SELECT CDVC.ID_CDVC AS ID_CDVC, CDVC.Latitud AS LATITUD, CDVC.Longitud AS LONGITUD, CA.NombreChofer AS CHOFER,
                                    CONVERT(VARCHAR(5), CDVC.FechaAlta, 103) + ' ' + CONVERT(VARCHAR(5), CDVC.FechaAlta, 108) AS FECHA_HORA
                            FROM Chofer_Detalle_Viajes_Coordenadas AS CDVC INNER JOIN
                                    Chofer_Alta AS CA ON CA.ID_CA = CDVC.ID_CA
                            WHERE (CDVC.ID_CA =	@@ID_CA OR @@ID_CA ='0') AND
                                    CDVC.ID_CDVC = (SELECT MAX(ID_CDVC)
                                                    FROM Chofer_Detalle_Viajes_Coordenadas
                                                    WHERE ID_CA = CA.ID_CA)
                            """
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        data = []
                        for row in results:
                            ID_CDVC = str(row[0])
                            LATITUD = str(row[1])
                            LONGITUD = str(row[2])
                            CHOFER = str(row[3])
                            FECHA_HORA = str(row[4])
                            datos = {'ID':ID_CDVC,'Latitud':LATITUD,'Longitud':LONGITUD,'Chofer':CHOFER, 'Fecha':FECHA_HORA}
                            data.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': data})
                    else:
                        data = "No existen coordenadas para el Chofer seleccionado."
                        return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                error = str(e)
                insertar_registro_error_sql("LOGISTICA","DETALLE UBICACION CHOFERES",request.user,error)
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

def cantidad_destinos(ID_PEDIDO):
    values = [ID_PEDIDO]
    cantidad = 0
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """
                    SELECT COUNT(*) AS CANTIDAD
                    FROM Chofer_Detalle_Chacras_Viajes AS CDCV
                    WHERE CDCV.ID_CVN = (SELECT ID_CVN 
                                            FROM Chofer_Detalle_Chacras_Viajes 
                                            WHERE IdPedidoFlete = %s) AND Estado = 'A'
                    """
            cursor.execute(sql, values)
            results = cursor.fetchone()
            if results:
                cantidad = int(results[0])
                return cantidad
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","CANT DESTINOS","usuario",error)
        return cantidad
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

def idPedido_esta_viaje(ID_PEDIDO):
    values = [ID_PEDIDO]
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """
                    SELECT ID_CDCV
                    FROM Chofer_Detalle_Chacras_Viajes
                    WHERE IdPedidoFlete = %s AND Estado = 'V'
                    """
            cursor.execute(sql, values)
            results = cursor.fetchone()
            if results:
                return True
            return False
    except Exception as e:
        error = str(e)
        insertar_registro_error_sql("LOGISTICA","ESTA EN VIAJE","usuario",error)
        return False
    finally:
        cursor.close()
        connections['TRESASES_APLICATIVO'].close()

@login_required
@csrf_exempt
def mensaje_elimina_destino_viaje(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_borrar')
        if user_has_permission:
            ID_PEDIDO = request.POST.get('ID_PEDIDO')
            values = [ID_PEDIDO]
            if idPedido_esta_viaje(ID_PEDIDO):
                data = "El pedido ya se encuentra en viaje y no se puede mover."
                return JsonResponse({'Message': 'Error', 'Nota': data})
            else:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql =   """ 
                                DECLARE @@IdPedidoFlete INT;
                                SET @@IdPedidoFlete = %s;

                                SELECT CDCV.IdPedidoFlete AS ID_PEDIDO, CASE WHEN PF.TipoDestino = 'U' THEN RTRIM(UB.Descripcion) WHEN PF.TipoDestino = 'P' THEN RTRIM(CH.Nombre) END AS DESTINO, 
                                        (SELECT COUNT(*) AS CANTIDAD
                                                    FROM Chofer_Detalle_Chacras_Viajes AS CDCV
                                                    WHERE CDCV.ID_CVN = (SELECT ID_CVN 
                                                                            FROM Chofer_Detalle_Chacras_Viajes 
                                                                            WHERE IdPedidoFlete = @@IdPedidoFlete AND Estado = 'A') AND Estado = 'A') AS CANTIDAD, CDCV.ID_CVN AS ID_CVN
                                FROM Chofer_Detalle_Chacras_Viajes AS CDCV LEFT JOIN
                                        S3A.dbo.PedidoFlete AS PF ON PF.IdPedidoFlete = CDCV.IdPedidoFlete LEFT JOIN
                                        S3A.dbo.Chacra AS CH ON CH.IdChacra = PF.IdChacra LEFT JOIN
                                        S3A.dbo.Ubicacion AS UB ON UB.IdUbicacion = PF.IdPlantaDestino
                                WHERE CDCV.ID_CVN = (SELECT ID_CVN 
                                                        FROM Chofer_Detalle_Chacras_Viajes 
                                                        WHERE IdPedidoFlete = @@IdPedidoFlete AND Estado = 'A') AND CDCV.Estado = 'A'
                                """
                        cursor.execute(sql, values)
                        results = cursor.fetchall()
                        if results:
                            data = []
                            for row in results:
                                CANTIDAD = int(row[2])
                                ID_PEDIDO = str(row[0])
                                DESTINO = str(row[1])
                                ID_CVN = str(row[3])
                                datos = {'IdPedidoFlete':ID_PEDIDO,'Destino':DESTINO}
                                data.append(datos)
                            mensaje = ""
                            titulo = ""
                            if CANTIDAD > 1:
                                titulo = "ELIMINACIÓN DE DESTINO"
                                mensaje = "Se MOVERÁ el pedido N°: " + str(ID_PEDIDO) + " a PENDIENTES para ser ASIGNADO nuevamente. Desea continuar?"
                            else:
                                titulo = "ELIMINACION DE VIAJE"
                                mensaje = "Se MOVERÁ el pedido N°: " + str(ID_PEDIDO) + " a PENDIENTES para ser ASIGNADO nuevamente y se ELIMINARÁ el VIAJE que lo contiene ya que es el único destino. Desea continuar?"
                            return JsonResponse({'Message': 'Success', 'Tabla': data, 'Cantidad':CANTIDAD, 'Titulo':titulo, 'Mensaje':mensaje, 'IdCVN':ID_CVN})
                        else:
                            data = "No existen viajes para el pedido seleccionado."
                            return JsonResponse({'Message': 'Error', 'Nota': data})
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("LOGISTICA","DETALLE UBICACION CHOFERES",request.user,error)
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    cursor.close()
                    connections['TRESASES_APLICATIVO'].close()
                
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@login_required
@csrf_exempt
def mueve_destinos_a_pendientes(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Logistica.puede_borrar')
        if user_has_permission:
            Id_pedido = request.POST.get('ID_PEDIDO')
            Cantidad = int(request.POST.get('CANTIDAD'))
            Id_cvn = request.POST.get('ID_CVN')
            Usuario = str(request.user).upper()
            values1 = [Id_pedido,Id_cvn]
            values2 = [Id_pedido]
            values3 = [Usuario,Id_cvn]
            
            if Cantidad > 1:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql =   """ 
                                UPDATE Chofer_Detalle_Chacras_Viajes SET Estado = 'E', FechaBaja = GETDATE()
                                WHERE IdPedidoFlete = %s AND ID_CVN = %s
                                """
                        cursor.execute(sql, values1)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows1 = cursor.fetchone()[0]
                        sql2 =   """ 
                                UPDATE S3A.dbo.PedidoFlete
                                SET Estado = 'P', IdTransportista = NULL, IdCamion = NULL, IdAcoplado = NULL, IdChofer = NULL, Chofer = NULL
                                WHERE IdPedidoFlete = %s
                                """
                        cursor.execute(sql2, values2)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]

                        if (affected_rows1 > 0 and affected_rows2 > 0):
                            return JsonResponse({'Message': 'Success', 'Nota': 'El Pedido N°: ' + str(Id_pedido) + ' se movió a PENDIENTES correctamente.'})
                        else:
                            return JsonResponse({'Message': 'Error', 'Nota': 'El Pedido N°: ' + str(Id_pedido) + ' NO se PUDO a PENDIENTES correctamente.'})
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("LOGISTICA","ELIMINA DESTINO",request.user,error)
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    cursor.close()
                    connections['TRESASES_APLICATIVO'].close()
            else:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql =   """ 
                                UPDATE Chofer_Detalle_Chacras_Viajes SET Estado = 'E', FechaBaja = GETDATE()
                                WHERE IdPedidoFlete = %s AND ID_CVN = %s
                                """
                        cursor.execute(sql, values1)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows1 = cursor.fetchone()[0]
                        sql2 =   """ 
                                UPDATE S3A.dbo.PedidoFlete
                                SET Estado = 'P', IdTransportista = NULL, IdCamion = NULL, IdAcoplado = NULL, IdChofer = NULL, Chofer = NULL
                                WHERE IdPedidoFlete = %s
                                """
                        cursor.execute(sql2, values2)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows2 = cursor.fetchone()[0]
                        sql3 =   """ 
                                UPDATE Chofer_Viajes_Notificacion SET Estado = 'E', FechaElimina = GETDATE(), UserElimina = %s
                                WHERE ID_CVN = %s
                                """
                        cursor.execute(sql3, values3)
                        cursor.execute("SELECT @@ROWCOUNT AS AffectedRows")
                        affected_rows3 = cursor.fetchone()[0]

                        if (affected_rows1 > 0 and affected_rows2 > 0 and affected_rows3 > 0):
                            return JsonResponse({'Message': 'Success', 'Nota': 'El Pedido: ' + str(Id_pedido) + ' se movió a PENDIENTES correctamente.'})   
                        else:
                            return JsonResponse({'Message': 'Error', 'Nota': 'El Pedido N°: ' + str(Id_pedido) + ' NO se PUDO mover a PENDIENTES correctamente.'})
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("LOGISTICA","ELIMINA VIAJE / DESTINOS",request.user,error)
                    data = str(e)
                    return JsonResponse({'Message': 'Error', 'Nota': data})
                finally:
                    cursor.close()
                    connections['TRESASES_APLICATIVO'].close()             
        return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
















































from openpyxl.styles import PatternFill, Font, Border, Side
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from S3A.funcionesGenerales import Centros_Modulos_Usuarios, obtenerHorasArchivo
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseForbidden
from openpyxl.drawing.image import Image
from django.views.static import serve
from openpyxl.styles import Alignment
from django.shortcuts import render
from django.db import connections
from datetime import datetime
from io import BytesIO
import pandas as pd
import os


### RENDERIZADO

@login_required
def Md_Frio(request):
    user_has_permission = request.user.has_perm('Md_Frio.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Frio/index.html')
    return render (request, 'Md_Frio/404.html')

@login_required
def Autoriza_Horas_Extras(request):
    user_has_permission = request.user.has_perm('Md_Frio.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Frio/HorasExtras/horasExtras.html')
    return render (request, 'Md_Frio/404.html')

@login_required
def Listado_Horas_Extras(request):
    user_has_permission = request.user.has_perm('Md_Frio.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Frio/HorasExtras/listadoHoras.html')
    return render (request, 'Md_Frio/404.html')

### PETICIONES REQUEST      EXEC GENERAL_PERSONAL_HORAS_EXTRAS '2025-01-01', '2025-03-31', '', '', 'CENTROS-FRIO'

def Centros_por_Usuario(request):
    if request.method == 'GET':
        usuario = str(request.user).upper()
        try:
            listado_data = Centros_Modulos_Usuarios("CENTROS-FRIO",usuario)
            listado_estados = [{'IdEstado':'', 'Descripcion':'TODO'},{'IdEstado':'0', 'Descripcion':'AUTORIZADO'},
                               {'IdEstado':'3', 'Descripcion':'PENDIENTE'},{'IdEstado':'8', 'Descripcion':'ELIMINADO'}]
            if listado_data:
                return JsonResponse({'Message': 'Success', 'Datos': listado_data, 'Estados':listado_estados})
            else:
                data = "No se encontraron Centros de Costos Asignados."
                return JsonResponse({'Message': 'Error', 'Nota': data})                
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})

@csrf_exempt
def combox_lista_personal_centro(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            centro = str(request.POST.get('Centro'))
            usuario = str(request.user).upper()
            codigo = 'CENTROS-FRIO'
            values = [centro,usuario,codigo]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ 
                        EXEC GENERAL_PERSONAL_POR_CENTRO %s, %s, %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [{'Legajo': '', 'Nombre':'TODOS'}]
                    for row in consulta:
                        lista_data.append({
                            'Legajo': str(row[0]), 
                            'Nombre': str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt
def data_listado_horas_extras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Md_Frio.puede_ver')
        if user_has_permission:
            try:
                inicio = str(request.POST.get('Inicio'))
                final = str(request.POST.get('Final'))
                estado = str(request.POST.get('Estado'))
                legajo = str(request.POST.get('Legajo'))
                centro = str(request.POST.get('Centro'))
                usuario = str(request.user).upper()
                codigo = 'CENTROS-FRIO'
                values = [inicio,final,estado,legajo,centro,usuario,codigo]
                lista_data = jsonListadoHorasEstras(values)
                if lista_data:
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})  
                else:
                    data = 'No se pudo cargar los datos.'
                    return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Error', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
def jsonListadoHorasEstras(values):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute('EXEC MDG_LISTADO_HORAS_EXTRAS %s,%s,%s,%s,%s,%s,%s', values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    lista_data.append({
                        'ID_HEP':str(row[0]), 
                        'Legajo':str(row[1]), 
                        'Nombre':str(row[2]),  
                        'Desde':str(row[3]), 
                        'Hasta':str(row[4]), 
                        'Motivo':str(row[5]), 
                        'Descripcion':str(row[6]), 
                        'Cantidad':str(row[7]), 
                        'Autoriza':str(row[8]), 
                        'Centro':str(row[9]), 
                        'Solicita':str(row[10]), 
                        'Estado':str(row[11]), 
                        'Tipo':str(row[12]),
                        'E':str(row[13])
                    })
                return lista_data 
            else:
                return lista_data
    except Exception as e:
        return lista_data

@csrf_exempt    
def inserta_elimina_horas_extras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            with connections['S3A'].cursor() as cursor:
                metodo = request.POST.get('Metodo', '')
                permissions = {
                    'G': 'Md_Frio.puede_insertar',
                    'E': 'Md_Frio.puede_borrar'
                }
                if metodo in permissions:
                    if not request.user.has_perm(permissions[metodo]):
                        return JsonResponse({'Message': 'Error', 'Nota': 'No tiene permisos para realizar la petición'})
                    
                cantidad = int(request.POST.get('IDHEP', 0))
                metodo = str(request.POST.get('Metodo', '0'))
                usuario = str(request.user).upper()
                index = 0
                for i in range(cantidad):
                    id_key = f'IDhep{i}'
                    tipo_key = f'TipoHoraExtra{i}'
                    cantidad_key = f'CantidadHoras{i}'
                    id_value = request.POST.get(id_key)
                    tipo_value = request.POST.get(tipo_key)
                    cantidad_value = request.POST.get(cantidad_key)
                    values = [metodo,cantidad_value,id_value,usuario,tipo_value]
                    cursor.execute('EXEC MDF_HORAS_EXTRAS_INSERTA_ELIMINA %s, %s, %s, %s, %s', values)
                    index = index + 1

            if index == cantidad:
                return JsonResponse({'Message': 'Success', 'Nota': 'La petición se realizó correctamente.'}) 
            else: 
                return JsonResponse({'Message': 'Error', 'Nota': 'Uno o más items no se pudieron guardad o eliminar.'})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def data_listado_horas_extras_web(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Md_Frio.puede_ver')
        if user_has_permission:
            try:
                usuario = str(request.user).upper()
                codigo = 'CENTROS-FRIO'
                inicio = str(request.POST.get('Inicio'))
                final = str(request.POST.get('Final'))
                centro = str(request.POST.get('Centro'))
                legajo = str(request.POST.get('Legajo'))
                estado = str(request.POST.get('Estado'))
                archivo = str(request.POST.get('Archivo'))
                values = [usuario,codigo,inicio,final,legajo,centro,estado]
                lista_data = jsonListadoHorasExtrasWeb(values)
                if lista_data:
                    if archivo == "N":
                        return JsonResponse({'Message': 'Success', 'Datos': lista_data})  
                    else:
                        excel_response = general_excel_horas_extras(lista_data)            
                        return excel_response 
                else:
                    data = 'No se encontraron datos.'
                    return JsonResponse({'Message': 'Error', 'Nota': data})
            except Exception as e:
                data = str(e)
                return JsonResponse({'Message': 'Error', 'Nota': data})
        else:
            return JsonResponse ({'Message': 'Error', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def descarga_archivo_excel(request, filename):
    nombre = filename
    filename = 'Applications/Md_Frio/Archivos/Excel/' + filename
    if os.path.exists(filename):
        response = serve(request, os.path.basename(filename), os.path.dirname(filename))
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        return response
    else:
        raise Http404

def jsonListadoHorasExtrasWeb(values):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute('EXEC LISTADO_WEB_HORAS_EXTRAS %s,%s,%s,%s,%s,%s,%s', values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    lista_data.append({
                        'IdHora':str(row[0]), 
                        'Legajo':str(row[1]), 
                        'Nombre':str(row[2]),  
                        'Centro':str(row[3]), 
                        'Desde':str(row[4]), 
                        'Hasta':str(row[5]), 
                        'Motivo':str(row[6]), 
                        'Descripcion':str(row[7]), 
                        'Tipo':str(row[8]), 
                        'Cantidad':str(row[10]), 
                        'Solicita':str(row[9]), 
                        'Dispositivo':str(row[11]), 
                        'IdEstado':str(row[12]), 
                        'Estado':str(row[13])
                    })
                return lista_data 
            else:
                return lista_data
    except Exception as e:
        return lista_data
    
def general_excel_horas_extras(lista_data):
    try:
        df = pd.DataFrame(lista_data)
        df = df.drop('IdEstado', axis=1)
        df.fillna('', inplace=True)
        output = BytesIO()
        columns1 = ['ID HORA', 'LEGAJO', 'APELLIDO Y NOMBRE', 'CENTRO DE COSTO', 'DESDE', 'HASTA', 'MOTIVO', 'DESCRIPCIÓN', 'TIPO', 'CANTIDAD', 'SOLICITA', 'DISPOS.', 'ESTADO']
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=5, sheet_name='LISTADO HORAS EXTRAS',header=columns1)
            worksheet = writer.sheets['LISTADO HORAS EXTRAS']
            logo = Image('static/3A/images/TA.png')
            logo.width = 100
            logo.height = 60
            worksheet.add_image(logo, 'G2')
            worksheet['C4'] = 'LISTADO HORAS EXTRAS'
            worksheet['C4'].font = Font(size=14, bold=True)
            worksheet['C4'].alignment = Alignment(horizontal='center', vertical='center')
            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            worksheet['E2'] = fecha_actual
            worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
            header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
            header_font = Font(color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center')

            for cell in worksheet[6]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in worksheet.iter_rows(min_row=7, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border_style

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
                
        output.seek(0)
        nombre_excel = f'Listado_horas_extras_{obtenerHorasArchivo()}.xlsx'
        with open('Applications/Md_Frio/Archivos/Excel/'+ nombre_excel, 'wb') as f:
            f.write(output.getvalue())

        return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel}) 
    except Exception as e:
        data = str(e)
        return JsonResponse({'Message': 'Error', 'Nota': data})

















































def jsonListadoHorasEstrasS3A(values,tipo):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute('LISTADO_S3A_HORAS_EXTRAS %s,%s,%s,%s,%s', values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    if tipo == 'H':
                        lista_data.append({
                            'Cantidad':str(row[0]), 
                            'Tipo':str(row[1]), 
                            'Valor':str(row[2]),  
                            'IdCentro':str(row[3]), 
                            'Centro':str(row[4]), 
                            'Nombre':str(row[5]), 
                            'Legajo':str(row[6]), 
                            'Fecha':str(row[7])
                        })
                    if tipo == 'C':
                        lista_data.append({
                            'Legajo':str(row[0]), 
                            'Nombre':str(row[1]), 
                            'Desde':str(row[2]),  
                            'Hasta':str(row[3]), 
                            'Cantidad':str(row[4]), 
                            'Motivo':str(row[5]), 
                            'Autoriza':str(row[6]), 
                            'Descripcion':str(row[7]), 
                            'Tipo':str(row[8]), 
                            'Valor':str(row[9]), 
                            'IdCentro':str(row[10]), 
                            'Centro':str(row[11]), 
                            'Pagada':str(row[12])
                        })
                    if tipo == 'L':
                        lista_data.append({
                            'Legajo':str(row[0]), 
                            'Nombre':str(row[1]), 
                            'Desde':str(row[2]),  
                            'Hasta':str(row[3]), 
                            'Cantidad':str(row[4]), 
                            'Motivo':str(row[5]), 
                            'Autoriza':str(row[6]), 
                            'Descripcion':str(row[7]), 
                            'Tipo':str(row[8]), 
                            'Valor':str(row[9]), 
                            'IdCentro':str(row[10]), 
                            'Centro':str(row[11]), 
                            'Pagada':str(row[12])
                        })
                return lista_data 
            else:
                return lista_data
    except Exception as e:
        return lista_data















































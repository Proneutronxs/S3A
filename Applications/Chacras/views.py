from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.decorators import login_required
from S3A.funcionesGenerales import *
from django.views.static import serve
from django.db import connections
from django.http import JsonResponse
from django.http import HttpResponse, Http404
import openpyxl
from openpyxl.styles import Font, Alignment
import os


# Create your views here.


@login_required
@permission_required('Chacras.puede_ingresar', raise_exception=True)
def Chacras(request):
    return render (request, 'Chacras/chacras.html')


@login_required
def planillas(request):
    return render (request, 'Chacras/Planillas/planillas.html')

@login_required
def general(request):
    return render (request, 'Chacras/Planillas/general.html')


@login_required
@csrf_exempt
def listarAdicionales(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacra.puede_ver')
        if user_has_permission:
            desde = request.POST.get('Inicio')
            hasta = request.POST.get('Final')
            centros = request.POST.get('Centro')
            legajos = request.POST.get('Legajo')
            descripcion = request.POST.get('Descripcion')
            pagos = request.POST.get('Pago')
            values = [desde,hasta,centros,legajos,descripcion,pagos]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """  
                        DECLARE @@Inicio DATE;
                        DECLARE @@Final DATE;
                        DECLARE @@Centro INT;
                        DECLARE @@Pago VARCHAR(5);
                        DECLARE @@Legajo INT;
                        DECLARE @@Descripcion VARCHAR(5);

                        SET @@Inicio = %s
                        SET @@Final = %s
                        SET @@Centro = %s
                        SET @@Legajo = %s
                        SET @@Descripcion = %s 
                        SET @@Pago = %s

                        SELECT Planilla_Chacras.IdPlanilla AS ID, Planilla_Chacras.Legajo AS LEGAJO, CONVERT(VARCHAR(38),(TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple)) AS NOMBRE,
                                TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto AS CENTRO, Planilla_Chacras.Categoria AS CATEGORIA,
                                CASE Planilla_Chacras.Descripcion WHEN 'PT' THEN 'POR TANTO' WHEN 'PD' THEN 'POR DÍA' WHEN 'FE' THEN 'FERIADO' WHEN 'SD' THEN 'SAB/DOM' WHEN 'AR' THEN 'ARREGLO' END AS DESCRIPCION,
                                CASE Planilla_Chacras.Tarea WHEN 'P' THEN 'PODA' WHEN 'R' THEN 'RALEO' WHEN 'V' THEN 'VARIOS' WHEN 'T' THEN 'TRACTOR' END AS TAREA, Planilla_Chacras.Jornales AS CANT_DIAS,
                                CASE Planilla_Chacras.Tipo WHEN 'F' THEN ('FILA N°: ' + CONVERT(VARCHAR(10),Planilla_Chacras.NroFila)) WHEN 'P' THEN 'PLANTAS' WHEN 'M' THEN 'METROS' WHEN 'O' THEN 'OTROS' END AS TIPO,
                                CASE Planilla_Chacras.Pago WHEN 'R' THEN 'RECIBO' WHEN 'A' THEN 'ADICIONAL' WHEN 'D' THEN 'DIFERENCIA' END AS PAGO,
                                CONVERT(VARCHAR(20) ,RTRIM(S3A.dbo.Chacra.Nombre)) AS CHACRA,
                                CASE WHEN SUB_CONSULTA.CONTEO IS NULL THEN '0' ELSE SUB_CONSULTA.CONTEO END AS CONTEO_SUB,
                                CASE WHEN SUB_CONSULTA.CANTIDAD IS NULL THEN '1' ELSE SUB_CONSULTA.CANTIDAD END AS CANTIDAD_SUB,
                                CASE WHEN SUB_CONSULTA.IMPORTE IS NULL THEN CONVERT(VARCHAR,Planilla_Chacras.PrecioUnitario) ELSE CONVERT(VARCHAR,SUB_CONSULTA.IMPORTE) END AS IMPORTE_SUB,
                                CASE WHEN Planilla_Chacras.E IS NULL THEN '#fbbc05' WHEN Planilla_Chacras.E = 'P' THEN '#fbbc05' WHEN Planilla_Chacras.E = 'S' THEN '#34a853' END AS ESTADO,
                                CASE WHEN Planilla_Chacras.Descripcion = 'PD' AND Planilla_Chacras.Pago = 'A' THEN 'S' ELSE 'N' END AS EDIT_IMPORTE,
                                CASE WHEN Planilla_Chacras.Descripcion = 'PT' AND (Planilla_Chacras.Tarea = 'P' OR Planilla_Chacras.Tarea = 'R') THEN 'S' ELSE 'N' END AS EDIT_PREMIO,
		                        CASE WHEN Planilla_Chacras.ImportePremio IS NULL THEN CONVERT(VARCHAR,0) ELSE CONVERT(VARCHAR,Planilla_Chacras.ImportePremio) END AS PREMIO
                        FROM   Planilla_Chacras INNER JOIN
                                    TresAses_ISISPayroll.dbo.Empleados ON Planilla_Chacras.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado INNER JOIN
                                    TresAses_ISISPayroll.dbo.CentrosCostos ON Planilla_Chacras.Centro = TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo INNER JOIN
                                    S3A.dbo.Chacra ON Planilla_Chacras.Lote = S3A.dbo.Chacra.IdChacra LEFT JOIN
                                    (SELECT CASE QR WHEN '' THEN NULL ELSE QR END AS QR, Tarea, COUNT(*) AS CONTEO, CASE WHEN COUNT(*) = 1 THEN '1' WHEN COUNT(*) = 2 THEN '0.5' WHEN COUNT(*) = 3 THEN '0.33' WHEN COUNT(*) = 4 THEN '0.25' END AS CANTIDAD,
                                            ROUND(CONVERT(DECIMAL(10, 2), PrecioUnitario / COUNT(*)), 2) AS IMPORTE
                                        FROM Planilla_Chacras
                                        GROUP BY QR, Tarea,PrecioUnitario) AS SUB_CONSULTA ON Planilla_Chacras.QR = SUB_CONSULTA.QR 
                        WHERE (Planilla_Chacras.IdPlanilla > 80) AND CONVERT(DATE, Planilla_Chacras.Fecha) >= @@Inicio AND CONVERT(DATE, Planilla_Chacras.Fecha) <= @@Final
                            AND (@@Centro = '0' OR Planilla_Chacras.Centro = @@Centro)
                            AND (@@Pago = '0' OR Planilla_Chacras.Pago = @@Pago)
                            AND (@@Legajo = '0' OR Planilla_Chacras.Legajo = @@Legajo)
                            AND (@@Descripcion  = '0' OR Planilla_Chacras.Descripcion = @@Descripcion)
                            AND (Planilla_Chacras.E <> 'E' OR Planilla_Chacras.E IS NULL)
                        ORDER BY NOMBRE
                    """
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        listado_tabla = []
                        for row in results:
                            #ID	LEGAJO	NOMBRE	CENTRO	CATEGORIA	DESCRIPCION	TAREA	CANT_DIAS	TIPO	PAGO	CHACRA	CONTEO_SUB	CANTIDAD_SUB	IMPORTE_SUB	ESTADO
                            idAdiconal = str(row[0])
                            legajo = str(row[1])
                            nombre = str(row[2])
                            centro = str(row[3])
                            cat = str(row[4])
                            desc = str(row[5])
                            tarea = str(row[6])
                            dias = str(row[7])
                            tipo = str(row[8])
                            pago = str(row[9])
                            chacra = str(row[10])
                            cantidad = str(row[12])
                            importe = formatear_moneda(str(row[13]))
                            estado = str(row[14])
                            edit_importe = str(row[15])
                            edit_premio = str(row[16])
                            importe_premio = formatear_moneda(str(row[17]))
                            datos = {'Id': idAdiconal, 'Legajo': legajo, 'Nombre': nombre, 'Centro':centro, 'Categoria':cat, 'Descripcion':desc, 'Tarea':tarea, 'Dias':dias,
                                     'Tipo':tipo, 'Pago':pago, 'Chacra':chacra, 'Cantidad':cantidad, 'Importe':importe, 'Estado':estado, 'Eimporte':edit_importe,'Epremio':edit_premio, 
                                     'ImportePremio':importe_premio}
                            listado_tabla.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': listado_tabla})
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})

            except Exception as e:
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Hubo un error al intentar resolver la petición. ' + str(e) })
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    

@login_required
@csrf_exempt
def eliminaAdicionalTildado(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacras.puede_borrar')
        if user_has_permission:
            usuario = str(request.user)
            adicionales = request.POST.getlist('idCheck')
            index = 0
            for ad in adicionales:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """ 
                                UPDATE Planilla_Chacras SET E = 'E', FechaBaja = GETDATE(), UserBaja = %s WHERE IdPlanilla = %s
                                
                            """
                        cursor.execute(sql, [usuario.upper(), ad])
                        cursor.commit()
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("EMPAQUE","ELIMINA ADICIONAL TILDADO",usuario.upper(),error)
                finally:
                    connections['TRESASES_APLICATIVO'].close()
                index = index + 1
            if index == len(adicionales):
                return JsonResponse({'Message': 'Success', 'Nota': 'Se eliminaron los Adicionales.'})
            else:
                return JsonResponse({'Message': 'Success', 'Nota': 'Hubo un error al eliminar algunos Adicionales.'})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})


@login_required
@csrf_exempt
def insertaImporteAdicional(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacras.puede_insertar')
        if user_has_permission:
            usuario = str(request.user)
            importe = request.POST.get('Premio')
            adicionales = request.POST.getlist('idCheck')
            index = 0
            for ad in adicionales:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """ 
                                UPDATE Planilla_Chacras SET PrecioUnitario = %s, FechaBaja = GETDATE(), UserMod = %s WHERE IdPlanilla = %s
                            """
                        cursor.execute(sql, [importe, usuario.upper(), ad])
                        cursor.commit()
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("EMPAQUE","INSERTA IMPORTE TILDADO",usuario.upper(),error)
                finally:
                    connections['TRESASES_APLICATIVO'].close()
                index = index + 1
            if index == len(adicionales):
                return JsonResponse({'Message': 'Success', 'Nota': 'Se actualizó el importe de los Adicionales.'})
            else:
                return JsonResponse({'Message': 'Success', 'Nota': 'Hubo un error al eliminar algunos Adicionales.'})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@login_required
@csrf_exempt
def insertaPremioAdicional(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacras.puede_insertar')
        if user_has_permission:
            usuario = str(request.user)
            importe = request.POST.get('Premio')
            adicionales = request.POST.getlist('idCheck')
            index = 0
            for ad in adicionales:
                try:
                    with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                        sql = """ 
                                UPDATE Planilla_Chacras SET ImportePremio = %s, FechaMod = GETDATE(), UserMod = %s WHERE IdPlanilla = %s
                            """
                        cursor.execute(sql, [importe, usuario.upper(), ad])
                        cursor.commit()
                except Exception as e:
                    error = str(e)
                    insertar_registro_error_sql("EMPAQUE","INSERTA PREMIO TILDADO",usuario.upper(),error)
                finally:
                    connections['TRESASES_APLICATIVO'].close()
                index = index + 1
            if index == len(adicionales):
                return JsonResponse({'Message': 'Success', 'Nota': 'Se actualizó el premio de los Adicionales.'})
            else:
                return JsonResponse({'Message': 'Success', 'Nota': 'Hubo un error al intentar introducir el premio en algunos Adicionales.'})
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    

@login_required
@csrf_exempt
def detalleAdicional(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacra.puede_ver')
        if user_has_permission:
            idAdicional = request.POST.get('ID')
            values = [idAdicional]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """  
                            SELECT Planilla_Chacras.IdPlanilla AS ID, Planilla_Chacras.Legajo AS LEGAJO, CONVERT(VARCHAR(38),(TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple)) AS NOMBRE,
                            TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto AS CENTRO, Planilla_Chacras.Categoria AS CATEGORIA,
                            CASE Planilla_Chacras.Descripcion WHEN 'PT' THEN 'POR TANTO' WHEN 'PD' THEN 'POR DÍA' WHEN 'FE' THEN 'FERIADO' WHEN 'SD' THEN 'SAB/DOM' WHEN 'AR' THEN 'ARREGLO' END AS DESCRIPCION,
                            CASE Planilla_Chacras.Tarea WHEN 'P' THEN 'PODA' WHEN 'R' THEN 'RALEO' WHEN 'V' THEN 'VARIOS' WHEN 'T' THEN 'TRACTOR' END AS TAREA, Planilla_Chacras.Jornales AS CANT_DIAS,
                            CASE Planilla_Chacras.Tipo WHEN 'F' THEN ('FILA N°: ' + CONVERT(VARCHAR(10),Planilla_Chacras.NroFila)) WHEN 'P' THEN 'PLANTAS' WHEN 'M' THEN 'METROS' WHEN 'O' THEN 'OTROS' END AS TIPO,
                            CASE Planilla_Chacras.Pago WHEN 'R' THEN 'RECIBO' WHEN 'A' THEN 'ADICIONAL' WHEN 'D' THEN 'DIFERENCIA' END AS PAGO,
                            CONVERT(VARCHAR(20) ,RTRIM(S3A.dbo.Chacra.Nombre)) AS CHACRA,
                            CASE WHEN SUB_CONSULTA.CONTEO IS NULL THEN '0' ELSE SUB_CONSULTA.CONTEO END AS CONTEO_SUB,
                            CASE WHEN SUB_CONSULTA.CANTIDAD IS NULL THEN '1' ELSE SUB_CONSULTA.CANTIDAD END AS CANTIDAD_SUB,
                            CASE WHEN SUB_CONSULTA.IMPORTE IS NULL THEN CONVERT(VARCHAR,Planilla_Chacras.PrecioUnitario) ELSE CONVERT(VARCHAR,SUB_CONSULTA.IMPORTE) END AS IMPORTE_SUB,
                            CASE Planilla_Chacras.Cuadro WHEN '0' THEN '-' ELSE Planilla_Chacras.Cuadro END AS CUADRO,
                            Planilla_Chacras.Usuario AS ENCARGADO, Planilla_Chacras.Observaciones AS OBS
                            FROM   Planilla_Chacras INNER JOIN
                                TresAses_ISISPayroll.dbo.Empleados ON Planilla_Chacras.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado INNER JOIN
                                TresAses_ISISPayroll.dbo.CentrosCostos ON Planilla_Chacras.Centro = TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo INNER JOIN
                                S3A.dbo.Chacra ON Planilla_Chacras.Lote = S3A.dbo.Chacra.IdChacra LEFT JOIN
                                (SELECT CASE QR WHEN '' THEN NULL ELSE QR END AS QR, Tarea, COUNT(*) AS CONTEO, CASE WHEN COUNT(*) = 1 THEN '1' WHEN COUNT(*) = 2 THEN '0.5' WHEN COUNT(*) = 3 THEN '0.33' WHEN COUNT(*) = 4 THEN '0.25' END AS CANTIDAD,
                                        ROUND(CONVERT(DECIMAL(10, 2), PrecioUnitario / COUNT(*)), 2) AS IMPORTE
                                    FROM Planilla_Chacras
                                    GROUP BY QR, Tarea,PrecioUnitario) AS SUB_CONSULTA ON Planilla_Chacras.QR = SUB_CONSULTA.QR 
                            WHERE (Planilla_Chacras.IdPlanilla > 80) AND Planilla_Chacras.IdPlanilla = %s
                    """
                    cursor.execute(sql, values)
                    results = cursor.fetchone()
                    if results:
                        detalles = []
                        idAdiconal = str(results[0])
                        legajo = str(results[1])
                        nombre = str(results[2])
                        centro = str(results[3])
                        cat = str(results[4])
                        desc = str(results[5])
                        tarea = str(results[6])
                        dias = str(results[7])
                        tipo = str(results[8])
                        pago = str(results[9])
                        chacra = str(results[10])
                        cantidad = str(results[12])
                        importe = formatear_moneda(str(results[13]))
                        cuadro = str(results[14])
                        encargado = str(results[15])
                        obs = str(results[16])
                        datos = {'Id': idAdiconal, 'Legajo': legajo, 'Nombre': nombre, 'Centro':centro, 'Categoria':cat, 'Descripcion':desc, 'Tarea':tarea, 'Dias':dias,
                                    'Tipo':tipo, 'Pago':pago, 'Chacra':chacra, 'Cantidad':cantidad, 'Importe':importe, 'Cuadro':cuadro, 'Encargado':encargado,'Obs':obs}
                        detalles.append(datos)
                        return JsonResponse({'Message': 'Success', 'Datos': detalles})
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})

            except Exception as e:
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Hubo un error al intentar resolver la petición. ' + str(e) })
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})
    
@login_required
@csrf_exempt
def crearArchivos(request):
    if request.method == 'POST':
        user_has_permission = request.user.has_perm('Chacra.puede_ver')
        if user_has_permission:
            desde = request.POST.get('Inicio')
            hasta = request.POST.get('Final')
            centros = request.POST.get('Centro')
            legajos = request.POST.get('Legajo')
            descripcion = request.POST.get('Descripcion')
            pagos = request.POST.get('Pago')
            values = [desde,hasta,centros,legajos,descripcion,pagos]
            try:
                with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                    sql = """  
                        DECLARE @@Inicio DATE;
                        DECLARE @@Final DATE;
                        DECLARE @@Centro INT;
                        DECLARE @@Pago VARCHAR(5);
                        DECLARE @@Legajo INT;
                        DECLARE @@Descripcion VARCHAR(5);

                        SET @@Inicio = %s
                        SET @@Final = %s
                        SET @@Centro = %s
                        SET @@Legajo = %s
                        SET @@Descripcion = %s 
                        SET @@Pago = %s

                        SELECT Planilla_Chacras.IdPlanilla AS ID, Planilla_Chacras.Legajo AS LEGAJO, CONVERT(VARCHAR(38),(TresAses_ISISPayroll.dbo.Empleados.ApellidoEmple + ' ' + TresAses_ISISPayroll.dbo.Empleados.NombresEmple)) AS NOMBRE,
                            TresAses_ISISPayroll.dbo.CentrosCostos.AbrevCtroCosto AS CENTRO, Planilla_Chacras.Categoria AS CATEGORIA,
                            CASE Planilla_Chacras.Descripcion WHEN 'PT' THEN 'POR TANTO' WHEN 'PD' THEN 'POR DÍA' WHEN 'FE' THEN 'FERIADO' WHEN 'SD' THEN 'SAB/DOM' WHEN 'AR' THEN 'ARREGLO' END AS DESCRIPCION,
                            CASE Planilla_Chacras.Tarea WHEN 'P' THEN 'PODA' WHEN 'R' THEN 'RALEO' WHEN 'V' THEN 'VARIOS' WHEN 'T' THEN 'TRACTOR' END AS TAREA, Planilla_Chacras.Jornales AS CANT_DIAS,
                            CASE Planilla_Chacras.Tipo WHEN 'F' THEN ('FILA N°: ' + CONVERT(VARCHAR(10),Planilla_Chacras.NroFila)) WHEN 'P' THEN 'PLANTAS' WHEN 'M' THEN 'METROS' WHEN 'O' THEN 'OTROS' END AS TIPO,
                            CASE Planilla_Chacras.Pago WHEN 'R' THEN 'RECIBO' WHEN 'A' THEN 'ADICIONAL' WHEN 'D' THEN 'DIFERENCIA' END AS PAGO,
                            CASE WHEN SUB_CONSULTA.CANTIDAD IS NULL THEN '1' ELSE SUB_CONSULTA.CANTIDAD END AS CANTIDAD_UNI,
                            CASE WHEN SUB_CONSULTA.IMPORTE IS NULL THEN CONVERT(VARCHAR,Planilla_Chacras.PrecioUnitario) ELSE CONVERT(VARCHAR,SUB_CONSULTA.IMPORTE) END AS IMPORTE_UNI,
	                        CASE WHEN Planilla_Chacras.ImportePremio IS NULL THEN CONVERT(VARCHAR,0) ELSE CONVERT(VARCHAR,Planilla_Chacras.ImportePremio) END AS PREMIO,
                            CONVERT(VARCHAR(20) ,RTRIM(S3A.dbo.Chacra.Nombre)) AS CHACRA,CASE Planilla_Chacras.Cuadro WHEN '0' THEN '-' ELSE Planilla_Chacras.Cuadro END AS CUADRO, Planilla_Chacras.Observaciones AS OBS,
                            Planilla_Chacras.Usuario AS ENCARGADO
                        FROM   Planilla_Chacras INNER JOIN
                            TresAses_ISISPayroll.dbo.Empleados ON Planilla_Chacras.Legajo = TresAses_ISISPayroll.dbo.Empleados.CodEmpleado INNER JOIN
                            TresAses_ISISPayroll.dbo.CentrosCostos ON Planilla_Chacras.Centro = TresAses_ISISPayroll.dbo.CentrosCostos.Regis_CCo INNER JOIN
                            S3A.dbo.Chacra ON Planilla_Chacras.Lote = S3A.dbo.Chacra.IdChacra LEFT JOIN
                            (SELECT CASE QR WHEN '' THEN NULL ELSE QR END AS QR, Tarea, COUNT(*) AS CONTEO, CASE WHEN COUNT(*) = 1 THEN '1' WHEN COUNT(*) = 2 THEN '0.5' WHEN COUNT(*) = 3 THEN '0.33' WHEN COUNT(*) = 4 THEN '0.25' END AS CANTIDAD,
                                    ROUND(CONVERT(DECIMAL(10, 2), PrecioUnitario / COUNT(*)), 2) AS IMPORTE
                                FROM Planilla_Chacras
                                GROUP BY QR, Tarea,PrecioUnitario) AS SUB_CONSULTA ON Planilla_Chacras.QR = SUB_CONSULTA.QR 
                        WHERE (Planilla_Chacras.IdPlanilla > 80) AND CONVERT(DATE, Planilla_Chacras.Fecha) >= @@Inicio AND CONVERT(DATE, Planilla_Chacras.Fecha) <= @@Final
                            AND (@@Centro = '0' OR Planilla_Chacras.Centro = @@Centro)
                            AND (@@Pago = '0' OR Planilla_Chacras.Pago = @@Pago)
                            AND (@@Legajo = '0' OR Planilla_Chacras.Legajo = @@Legajo)
                            AND (@@Descripcion  = '0' OR Planilla_Chacras.Descripcion = @@Descripcion)
                            AND (Planilla_Chacras.E <> 'E' OR Planilla_Chacras.E IS NULL)
                        ORDER BY NOMBRE
                    """
                    
                    cursor.execute(sql, values)
                    results = cursor.fetchall()
                    if results:
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Lista_Adicionales"
                        bold_font = Font(bold=True)
                        center_alignment = Alignment(horizontal="center")
                        encabezados = [
                            "ID", "LEGAJO", "NOMBRE", "CENTRO", "CATEGORIA", "DESCRIPCION", "TAREA",
                            "CANT_DIAS", "TIPO", "PAGO", "CANTIDAD_UNI", "IMPORTE_UNI", "CHACRA", 
                            "CUADRO", "OBS", "ENCARGADO"
                        ]
                        for col_num, encabezado in enumerate(encabezados, start=1):
                            celda = ws.cell(row=1, column=col_num, value=encabezado)
                            celda.font = bold_font
                            celda.alignment = center_alignment
                        for idx, row in enumerate(results, start=2):  # Comienza en la fila 2
                            ws.cell(row=idx, column=1, value=str(row[0]))   # ID
                            ws.cell(row=idx, column=2, value=str(row[1]))   # LEGAJO
                            ws.cell(row=idx, column=3, value=str(row[2]))   # NOMBRE
                            ws.cell(row=idx, column=4, value=str(row[3]))   # CENTRO
                            ws.cell(row=idx, column=5, value=str(row[4]))   # CATEGORIA
                            ws.cell(row=idx, column=6, value=str(row[5]))   # DESCRIPCION
                            ws.cell(row=idx, column=7, value=str(row[6]))   # TAREA
                            ws.cell(row=idx, column=8, value=str(row[7]))   # CANT_DIAS
                            ws.cell(row=idx, column=9, value=str(row[8]))   # TIPO
                            ws.cell(row=idx, column=10, value=str(row[9]))  # PAGO
                            ws.cell(row=idx, column=11, value=str(row[10])) # CANTIDAD_UNI
                            ws.cell(row=idx, column=12, value=formatear_moneda(str(row[11]))) # IMPORTE_UNI
                            ws.cell(row=idx, column=13, value=formatear_moneda(str(row[12])))  # PREMIO
                            ws.cell(row=idx, column=13, value=str(row[13])) # CHACRA
                            ws.cell(row=idx, column=14, value=str(row[14])) # CUADRO
                            ws.cell(row=idx, column=15, value=str(row[15])) # OBS
                            ws.cell(row=idx, column=16, value=str(row[16])) # ENCARGADO

                        for column in ws.columns:
                            max_length = 0
                            column = list(column)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            ws.column_dimensions[column[0].column_letter].width = adjusted_width

                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename="Listado_Adicionales.xlsx"'
                        wb.save(response)
                        return response
                    else:
                        return JsonResponse({'Message': 'Not Found', 'Nota': 'No se encontraron datos.'})

            except Exception as e:
                print(e)
                return JsonResponse ({'Message': 'Not Found', 'Nota': 'Hubo un error al intentar resolver la petición. ' + str(e) })
            finally:
                cursor.close()
                connections['TRESASES_APLICATIVO'].close()
        else:
            return JsonResponse ({'Message': 'Not Found', 'Nota': 'No tiene permisos para resolver la petición.'})
    else:
        data = "No se pudo resolver la Petición"
        return JsonResponse({'Message': 'Error', 'Nota': data})



# for row in results:
#     ID = str(row[0])
#     LEGAJO = str(row[1])
#     NOMBRE = str(row[2])
#     CENTRO = str(row[3])
#     CATEGORIA = str(row[4])
#     DESCRIPCION = str(row[5])
#     TAREA = str(row[6])
#     CANT_DIAS = str(row[7])
#     TIPO = str(row[8])
#     PAGO = str(row[9])
#     CANTIDAD_UNI = str(row[10])
#     IMPORTE_UNI = formatear_moneda(str(row[11]))
#     PREMIO = formatear_moneda(str(row[12]))
#     CHACRA = str(row[13])
#     CUADRO = str(row[14])
#     OBS = str(row[15])
#     ENCARGADO = str(row[16])

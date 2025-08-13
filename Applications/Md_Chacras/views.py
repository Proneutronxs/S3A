
from S3A.funcionesGenerales import obtenerHorasArchivo, registroRealizado, formatear_moneda
from openpyxl.styles import PatternFill, Font, Border, Side
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseForbidden
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Protection
from django.views.static import serve
from openpyxl.styles import Alignment
from django.shortcuts import render
from django.db import connections
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from io import BytesIO
import pandas as pd
import xlrd
import json
import os

# Create your views here.


@login_required
def Md_Chacras(request):
    user_has_permission = request.user.has_perm('Md_Chacras.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Chacras/index.html')
    return render (request, 'Md_Chacras/404.html')

@login_required
def Presupuesto(request):
    user_has_permission = request.user.has_perm('Md_Chacras.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Chacras/Sipreta/presupuesto.html')
    return render (request, 'Md_Chacras/404.html')

@login_required
def listado_labores(request):
    user_has_permission = request.user.has_perm('Md_Chacras.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Chacras/Sipreta/listadoLabores.html')
    return render (request, 'Md_Chacras/404.html')

@login_required
def listado_labores_chacras(request):
    user_has_permission = request.user.has_perm('Md_Chacras.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Chacras/Sipreta/listadoLaboresChacra.html')
    return render (request, 'Md_Chacras/404.html')

@login_required
def Horas_Extras(request):
    user_has_permission = request.user.has_perm('Md_Chacras.puede_ingresar')
    if user_has_permission:
        return render (request, 'Md_Chacras/Mobile/horas_extras.html')
    return render (request, 'Md_Chacras/404.html')

def descarga_archivo_excel(request, filename):
    nombre = filename
    filename = 'Applications/Md_Chacras/Archivos/Excel/' + filename
    if os.path.exists(filename):
        response = serve(request, os.path.basename(filename), os.path.dirname(filename))
        response['Content-Disposition'] = f'attachment; filename="{nombre}"'
        return response
    else:
        raise Http404

@csrf_exempt
def data_listado_horas_extras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            inicio = str(request.POST.get('Incio'))
            final = str(request.POST.get('Final'))
            estado = str(request.POST.get('Estado'))
            tipo = str(request.POST.get('Tipo'))
            archivo = str(request.POST.get('Archivo'))
            values = [inicio,final,estado,tipo]
            lista_data = jsonHorasExtras(values)
            if lista_data:
                if archivo == "N":
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    excel_response = general_excel_horas_extras(lista_data)
                    return excel_response
            else:
                data = 'No se encontraron datos.'
                return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def jsonHorasExtras(values):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute('EXEC MDC_HORAS_EXTRAS %s,%s,%s,%s', values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    lista_data.append({
                        'IDhep':str(row[0]),
                        'Tipo':str(row[1]),
                        'Legajo':str(row[2]),
                        'Nombre':str(row[3]),
                        'Centro':str(row[4]),
                        'Desde':str(row[5]),
                        'Hasta':str(row[6]),
                        'Motivo':str(row[7]),
                        'Descripcion':str(row[8]),
                        'Cantidad':str(row[9]),
                        'Solicita':str(row[10]),
                        'Importe':str(row[11]),
                        'Estado':str(row[12])
                    })
                return lista_data
            else:
                return lista_data
    except Exception as e:
        return lista_data

def general_excel_horas_extras(lista_data):
    try:
        df = pd.DataFrame(lista_data)
        df.fillna('', inplace=True)
        output = BytesIO()
        columns1 = ['ID HORA', 'TIPO HORA', 'LEGAJO', 'APELLIDO Y NOMBRE', 'CENTRO DE COSTO', 'DESDE', 'HASTA', 'MOTIVO', 'DESCRIPCIÓN', 'CANTIDAD', 'SOLICITA', 'IMPORTE A', 'ESTADO']
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=5, sheet_name='LISTADO HORAS EXTRAS',header=columns1)
            worksheet = writer.sheets['LISTADO HORAS EXTRAS']
            logo = Image('static/3A/images/TA.png')
            logo.width = 100
            logo.height = 60
            worksheet.add_image(logo, 'G2')
            worksheet['C4'] = 'LISTADO HORAS EXTRAS'
            worksheet['C4'].font = Font(size=14, bold=True)
            worksheet['C4'].alignment = Alignment(horizontal='center', vertical='center')
            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            worksheet['E2'] = fecha_actual
            worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
            header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
            header_font = Font(color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center')

            for cell in worksheet[6]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in worksheet.iter_rows(min_row=7, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border_style

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

        output.seek(0)
        nombre_excel = f'Listado_horas_extras_{obtenerHorasArchivo()}.xlsx'
        with open('Applications/Md_Chacras/Archivos/Excel/'+ nombre_excel, 'wb') as f:
            f.write(output.getvalue())

        return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel})
    except Exception as e:
        data = str(e)
        return JsonResponse({'Message': 'Error', 'Nota': data})

def convertir_a_numerico_stock_ventas(df):
    try:
        df['Valor1'] = pd.to_numeric(df['Valor1'], errors='coerce')
        df['Valor2'] = pd.to_numeric(df['Valor2'], errors='coerce')
        df['Valor3'] = pd.to_numeric(df['Valor3'], errors='coerce')
        df['Valor4'] = pd.to_numeric(df['Valor4'], errors='coerce')
        df['Valor5'] = pd.to_numeric(df['Valor5'], errors='coerce')
        df['Valor6'] = pd.to_numeric(df['Valor6'], errors='coerce')
        df['Valor7'] = pd.to_numeric(df['Valor7'], errors='coerce')
        df['Valor8'] = pd.to_numeric(df['Valor8'], errors='coerce')
        df['Valor9'] = pd.to_numeric(df['Valor9'], errors='coerce')
        df['Valor10'] = pd.to_numeric(df['Valor10'], errors='coerce')
        df['Valor11'] = pd.to_numeric(df['Valor11'], errors='coerce')
        df['Valor12'] = pd.to_numeric(df['Valor12'], errors='coerce')
        df['Valor13'] = pd.to_numeric(df['Valor13'], errors='coerce')
        df['Total'] = pd.to_numeric(df['Total'], errors='coerce')
        return df
    except Exception as e:
        return None

@csrf_exempt
def inserta_elimina_horas_extras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            with connections['S3A'].cursor() as cursor:
                metodo = request.POST.get('Metodo', '')
                importe = request.POST.get('Importe', '')
                cantidad = int(request.POST.get('IDHEP', 0))
                usuario = str(request.user).upper()
                index = 0
                for i in range(cantidad):
                    id_key = f'IDhep{i}'
                    id_value = request.POST.get(id_key)
                    values = [metodo, importe, id_value, usuario]
                    cursor.execute('EXEC MDC_HORAS_EXTRAS_INSERTA_ELIMINA %s, %s, %s, %s', values)
                    index = index + 1
            if index == cantidad:
                return JsonResponse({'Message': 'Success', 'Nota': 'La petición se realizó correctamente.'})
            else:
                return JsonResponse({'Message': 'Error', 'Nota': 'Uno o más items no se pudieron guardad o eliminar.'})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

#SIPRETA → Sistema de Información para Poda y Raleo en Establecimientos de Tres Ases

def listado_combox_productor(request):
    if request.method == 'GET':
        try:
            lista_data = []
            lista_especies = []
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_PRODUCTORES_HABILITADOS
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        lista_data.append({
                            "IdProductor":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                sql_especies = """
                        EXEC LISTADO_ESPECIES_HABILITADAS
                """
                cursor.execute(sql_especies)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        lista_especies.append({
                            "IdEspecie":str(row[0]),
                            "Descripcion":str(row[1])
                        })


            if lista_data:
                return JsonResponse({'Message': 'Success', 'Datos': lista_data, 'Especies':lista_especies})
            else:
                data = "No se encontraron Datos."
                return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def listado_combox_chacras_x_productor(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idProductor = str(request.POST.get('IdProductor'))
            values = [idProductor]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_CHACRAS_X_PRODUCTOR %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] #{'IdChacra': '', 'Descripcion':'TODOS'}
                    for row in consulta:
                        lista_data.append({
                            "IdChacra":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data, 'Chacras':lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt
def listado_combox_variedades_x_especies(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idProductor = str(request.POST.get('IdEspecie'))
            values = [idProductor]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_VARIEDADES_X_ESPECIE  %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] #{'IdChacra': '', 'Descripcion':'TODOS'}
                    for row in consulta:
                        lista_data.append({
                            "IdVariedad":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})
    
@csrf_exempt
def listado_combox_cuadros_x_chacra(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idChacra = str(request.POST.get('IdChacra'))
            values = [idChacra]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """
                        EXEC LISTADO_CUADROS_X_CHACRA %s
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    lista_data = [] 
                    for row in consulta:
                        lista_data.append({
                            "IdCuadro":str(row[0]),
                            "Descripcion":str(row[1])
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': lista_data})
                else:
                    data = "No se encontraron Datos."
                    return JsonResponse({'Message': 'Error', 'Nota': data})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def listado_chacras_x_filas(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            idProductor = str(request.POST.get('IdProductor'))
            idChacra = str(request.POST.get('IdChacra'))
            tipo = str(request.POST.get('Tipo'))
            values = [idProductor,idChacra]
            listado_data = jsonListadoChacrasFilas(values)
            if tipo == 'TT':
                return JsonResponse({'Message': 'Success', 'Datos': listado_data})
            else:
                nombre_excel = crear_excel_presupuesto(listado_data,idChacra,idProductor)
                return JsonResponse({'Message': 'Success', 'Excel': nombre_excel})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def jsonListadoChacrasFilas(values):
    lista_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            cursor.execute('EXEC LISTADO_FILAS_POR_CHACRA %s,%s', values)
            consulta = cursor.fetchall()
            if consulta:
                for row in consulta:
                    lista_data.append({
                        "IdProductor":str(row[0]),
                        "Productor":str(row[1]),
                        "IdChacra":str(row[2]),
                        "Chacra":str(row[3]),
                        "Cuadro":str(row[4]),
                        "Fila":str(row[5]),
                        "IdEspecie":str(row[6]),
                        "Especie":str(row[7]),
                        "IdVariedad":str(row[8]),
                        "Variedad":str(row[9]),
                        "AñoPlantacion":str(row[10]),
                        "NroPlantas":str(row[11]),
                        "DFilas":str(row[12]),
                        "DPlantas":str(row[13]),
                        "SupPlanta":str(row[14]),
                        "Presupuesto":str(row[15]),
                        "QRFila":str(row[16]),
                        "IdCuadro":str(row[17])
                    })
                return lista_data
            else:
                return lista_data
    except Exception as e:
        return lista_data

def crear_excel_presupuesto(lista_data, chacra, productor):
    try:
        wb = Workbook()
        ws = wb.active

        columnas = [
            "PRODUCTOR", "ID CHACRA", "CHACRA", "ID CUADRO", "CUADRO", "FILA",
            "CANT. PLANTAS", "ID VARIEDAD", "VARIEDAD", "AÑO", "P. PODA", "P. RALEO"
        ]
        ws.append(columnas)

        for row in lista_data:
            fila = [
                row["Productor"],
                row["IdChacra"],
                row["Chacra"],
                row["IdCuadro"],
                row["Cuadro"],
                row["Fila"],
                row["NroPlantas"],
                row["IdVariedad"],
                row["Variedad"],
                "",  # AÑO
                "",  # P. PODA
                ""   # P. RALEO
            ]
            ws.append(fila)

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[col_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=12):
            for cell in row:
                if cell.col_idx in [10, 11, 12]:
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)

        ws.protection.sheet = True
        nombre = f'Listado_Chacra_{chacra}_{productor}.xlsx'
        wb.save('Applications/Md_Chacras/Archivos/Excel/' + nombre)
        return nombre
    except Exception as e:
        return str(e)

@csrf_exempt
def recibir_archivo_excel(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivoExcel']
            nombre_archivo = archivo_excel.name
            extension = os.path.splitext(nombre_archivo)[1].lower()
            lista_data = []

            if extension == '.xlsx':
                wb = load_workbook(archivo_excel)
                ws = wb.worksheets[0]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    validar_fila_chacras(row)
                    lista_data.append(row)

            elif extension == '.xls':
                libro = xlrd.open_workbook(file_contents=archivo_excel.read())
                hoja = libro.sheet_by_index(0)
                for i in range(1, hoja.nrows):
                    row = [None if cell == '' else cell for cell in hoja.row_values(i)]
                    row = tuple(row)
                    validar_fila_chacras(row)
                    lista_data.append(row)
            else:
                raise ValueError("Formato de archivo no soportado. Solo se aceptan .xls o .xlsx")

            if lista_data:
                json_resultado = json_chacras(lista_data)
                return JsonResponse({'Message': 'Success', 'Datos': json_resultado})
            else:
                raise ValueError("No se pudieron cargar los datos.")
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})

    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def recibir_archivo_excel_presupuesto(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            archivo_excel = request.FILES['archivoExcel']
            nombre_archivo = archivo_excel.name
            extension = os.path.splitext(nombre_archivo)[1].lower()
            lista_data = []
            if extension == '.xlsx':
                wb = load_workbook(archivo_excel)
                ws = wb.worksheets[0]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    validar_fila_presupuesto(row)
                    lista_data.append({
                        "Productor":row[0],
                        "IdChacra":row[1],
                        "Chacra":row[2],
                        "IdCuadro":row[3],
                        "Cuadro":row[4],
                        "Fila":row[5],
                        "CantPlantas":row[6],
                        "IdVariedad":row[7],
                        "Variedad":row[8],
                        "Año":row[9],
                        "Poda":row[10],
                        "Raleo":row[11]
                    })
            else:
                raise ValueError("Formato de archivo no soportado. Solo se aceptan .xlsx")
            if lista_data:
                return JsonResponse({'Message': 'Success', 'Datos': lista_data})
            else:
                raise ValueError("No se pudieron cargar los datos.")
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

class ValidacionError(Exception):
    pass

def validar_fila_chacras(row):
    try:
        legajo = int(row[1])
    except ValueError:
        raise ValidacionError("La columna 'ID CHACRA' debe contener sólo números enteros.")

    if not str(row[7]).strip():
        raise ValidacionError("La columna 'ID CUADRO' no puede estar vacía.")

    if row[8] is None:
        raise ValidacionError("La columna 'FILA' debe contener un valor.")

    try:
        variedad = int(row[11])
    except ValueError:
        raise ValidacionError("La columna 'ID VARIEDAD' debe contener sólo números enteros.")

    try:
        anio = int(row[13])
    except ValueError:
        raise ValidacionError("La columna 'AÑO' debe contener sólo números enteros.")

    try:
        anio = int(row[14])
    except ValueError:
        raise ValidacionError("La columna 'N° PLANTAS' debe contener sólo números enteros.")

    for i in [15, 16, 17, 18]:
        if not isinstance(row[i], (int, float, type(None))):
            raise ValidacionError(f"La columna {i+1} debe contener sólo números o estar vacía.")

def validar_fila_presupuesto(row):
    try:
        legajo = int(row[9])
    except ValueError:
        raise ValidacionError("La columna AÑO' debe contener sólo números enteros.")

    for i in [10, 11]:
        if not isinstance(row[i], (int, float, type(None))):
            raise ValidacionError(f"La columna {i+1} debe contener sólo números o estar vacía.")

def json_chacras(filas):
    resultado = {
        "Chacras": []
    }

    chacras_dict = {}

    for row in filas:
        IdFila = row[0]
        IdChacra = int(row[1])
        Chacra = (row[2] or "")
        RazonSocial = (row[5] or "")
        Productor = (row[6] or "")
        Cuadro = str(row[7]).replace('.0','')
        FilaNum = str(row[8]).replace('.0','')
        IdEspecie = str(row[9])
        Especie = (row[10] or "")
        IdVariedad = int(row[11])
        Variedad = (row[12] or "")
        Año = str(row[13]).replace('.0','')
        NroPlantas = str(row[14]).replace('.0','')
        DistFilas = str(row[15]).replace('.0','')
        DistPlantas = str(row[16]).replace('.0','')
        Superficie = f"{row[17]:.3f}" if row[17] else None
        Presupuesto = str(row[18]) if row[18] else None
        Tarea = str(row[19]) if row[19] else None

        # Agrupar por IdChacra
        if IdChacra not in chacras_dict:
            chacras_dict[IdChacra] = {
                "IdCracra": str(IdChacra),
                "Chacra": Chacra,
                "Productor": Productor,
                "Cuadros": []
            }

        chacra = chacras_dict[IdChacra]

        # Buscar cuadro existente
        cuadro_existente = next((c for c in chacra["Cuadros"] if Cuadro in c), None)

        if not cuadro_existente:
            cuadro_existente = {Cuadro: []}
            chacra["Cuadros"].append(cuadro_existente)

        # Crear fila
        nueva_fila = {
            "IdFila": IdFila,
            "Fila": FilaNum,
            "IdEspecie": IdEspecie,
            "Especie": Especie,
            "IdVariedad": IdVariedad,
            "Variedad": Variedad,
            "Año": Año,
            "NroPlantas": NroPlantas,
            "DistFilas": DistFilas,
            "DistPlantas": DistPlantas,
            "Superficie": Superficie,
            "Presupuesto": Presupuesto,
            "Tarea": Tarea
        }

        cuadro_existente[Cuadro].append(nueva_fila)

    resultado["Chacras"] = list(chacras_dict.values())
    return resultado

@csrf_exempt
def insertar_chacras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            Usuario = str(request.user).upper()
            chacras_json = request.POST.get('Chacras')
            datos_chacras = json.loads(chacras_json)
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                for chacra in datos_chacras['Chacras']:
                    id_chacra = chacra['IdCracra']
                    for cuadro in chacra['Cuadros']:
                        for cuadro_id, filas in cuadro.items():
                            values = [id_chacra, cuadro_id]
                            cursor.execute("EXEC SP_INSERTA_CHACRA_CUADRO %s,%s", values)
                            cursor.execute("SELECT ID_CUADRO FROM SPA_CUADRO WHERE ID_CHACRA = %s AND NOMBRE_CUADRO = %s ", values)
                            result = cursor.fetchone()
                            id_cuadro = result[0]
                            for fila in filas:
                                values_filas = [id_cuadro, fila['Fila'], fila['IdVariedad'], fila['Año'], fila['NroPlantas'], fila['DistFilas'], fila['DistPlantas'], fila['Superficie'], 'A']
                                cursor.execute("EXEC SPA_INSERTA_CHACRAS_CUADROS_FILAS %s,%s,%s,%s,%s,%s,%s,%s,%s", values_filas)
            return JsonResponse({'Message': 'Success', 'Nota': 'El archivo se subió correctamente.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def insertar_presupuesto(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            presupuesto_json = request.POST.get('Presupuesto')
            datos_presupuesto = json.loads(presupuesto_json)
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                for row in datos_presupuesto:
                    IdCuadro = row['IdCuadro']
                    IdFila = row['Fila']
                    IdVariedad = row['IdVariedad']
                    Año = row['Año']
                    Poda = row['Poda']
                    Raleo = row['Raleo']
                    values_poda = [IdFila, IdCuadro, IdVariedad, 'P', Año, Poda]
                    values_raleo = [IdFila, IdCuadro, IdVariedad, 'R', Año, Poda]

                    if Poda != None:
                        cursor.execute("EXEC SP_INSERTA_PRESUPUESTO %s,%s,%s,%s,%s,%s", values_poda)
                    if Raleo != None:
                        cursor.execute("EXEC SP_INSERTA_PRESUPUESTO %s,%s,%s,%s,%s,%s", values_raleo)

            return JsonResponse({'Message': 'Success', 'Nota': 'Los datos se guardaron correctamente.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

############# LISTADO LABORES

def carga_inicial_listado_labores(request):
    if request.method == 'GET':
        try:
            listado_productores = []
            listado_personal = []
            listado_labores = [{'Codigo': 'P', 'Descripcion':'PODA'},{'Codigo': 'R', 'Descripcion':'RALEO'}]
            listado_encargados = []
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                #### listado productores
                sql = """ EXEC LISTADO_PRODUCTORES_HABILITADOS """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_productores.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })

                #### listado personal
                sql = """
                        SELECT EM.CodEmpleado AS LEGAJO, ((CONVERT(VARCHAR, EM.CodEmpleado)) + ' - ' + EM.ApellidoEmple + ' ' + EM.NombresEmple) AS NOMBRES
                        FROM TresAses_ISISPayroll.dbo.Empleados AS EM INNER JOIN
                            TresAses_ISISPayroll.dbo.CentrosCostos AS CC ON CC.Regis_CCo = EM.Regis_CCo
                        WHERE CC.AbrevCtroCosto LIKE ('C%%')
                                AND EM.BajaDefinitivaEmple = '2'
                        ORDER BY EM.ApellidoEmple + ' ' + EM.NombresEmple
                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_personal.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })

                #### listado encargados
                sql = """
                        SELECT RTRIM(US.Usuario) AS ENCARGADO, CASE WHEN ( EM.ApellidoEmple + ' ' + EM.NombresEmple) IS NULL THEN US.Usuario
                            ELSE ( EM.ApellidoEmple + ' ' + EM.NombresEmple) END AS NOMBRES
                        FROM USUARIOS AS US LEFT JOIN
                            TresAses_ISISPayroll.dbo.Empleados AS EM ON EM.CodEmpleado = US.CodEmpleado
                        WHERE US.Estado = 'A'
                            AND US.CodEmpleado NOT IN('99999')
                            AND US.Tipo IN ('EC','G')
                        ORDER BY EM.ApellidoEmple + ' ' + EM.NombresEmple

                    """
                cursor.execute(sql)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_encargados.append({
                            'Codigo':str(row[0]),
                            'Descripcion':str(row[1])
                        })
            return JsonResponse({'Message': 'Success', 'Productores':listado_productores, 'Personal':listado_personal, 'Labores':listado_labores, 'Encargados':listado_encargados})
        except Exception as e:
            data = str(e)
            return JsonResponse({'Message': 'Error', 'Nota': data})
    else:
        return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def cuadro_personal_x_chacra(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_cuadros = []
            listado_personal = []
            idChacra = request.POST.get('IdChacra')
            values = [idChacra]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql_cuadros = """
                    SELECT ID_CUADRO, NOMBRE_CUADRO
                    FROM SPA_CUADRO
                    WHERE ID_CHACRA = %s
                    ORDER BY NOMBRE_CUADRO
                    """
                cursor.execute(sql_cuadros, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_cuadros.append({
                            "IdCuadro": str(row[0]),
                            "Cuadro":str(row[1])
                        })

                sql_personal = """
                    SELECT DISTINCT(ST.ID_LEGAJO)  AS LEGAJO,
                        CASE
                            WHEN SC.ID_CHACRA = '1000001' THEN (SELECT CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple)) FROM Rommik_isispayroll.dbo.Empleados WHERE CodEmpleado = ST.ID_LEGAJO)
                            ELSE (SELECT CONVERT(VARCHAR(30), (ApellidoEmple + ' ' + NombresEmple)) FROM TresAses_ISISPayroll.dbo.Empleados WHERE CodEmpleado = ST.ID_LEGAJO)
                        END AS NOMBRES
                    FROM SPA_TAREA AS ST INNER JOIN
                        SPA_QR AS QR ON QR.ID_QR = ST.ID_QR_FILA INNER JOIN
                        SPA_CUADRO AS SC ON SC.ID_CUADRO = QR.ID_CUADRO
                    WHERE (CONVERT(DATE,ST.FECHA) >= '2025-06-08')
                        AND	SC.ID_CHACRA = %s
                    ORDER BY NOMBRES

                    """
                cursor.execute(sql_personal,values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_personal.append({
                            "Legajo":str(row[0]),
                            "Nombre":str(row[1])
                        })
            return JsonResponse({'Message': 'Success', 'Cuadros': listado_cuadros, 'Personal': listado_personal})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def listado_detalle_labores(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idLegajo = str(request.POST.get('IdLegajo'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            idEncargado = str(request.POST.get('IdEncargado'))
            idLabor = str(request.POST.get('IdLabor'))
            values = [inicio,final,idLegajo,idChacra,idEncargado,idLabor,idCuadro]
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ EXEC SP_SELECT_DETALLE_LABORES %s, %s, %s, %s, %s, %s, %s  """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "LEGAJO":row[0],
                            "NOMBRES":row[1],
                            "FECHA":row[2],
                            "QR":row[3],
                            "ID_CUADRO":row[4],
                            "ID_CHACRA":row[5],
                            "ID_PRODUCTOR":row[6],
                            "PRODUCTOR":row[7],
                            "CHACRA":row[8],
                            "CUADRO":row[9],
                            "FILA":row[10],
                            "VARIEDADES":row[11],
                            "CANT_PLANTAS":row[12],
                            "LABOR":row[13],
                            "IMPORTE":row[14],
                            "ID_QR_FILA":row[15],
                            "PRESUPUESTO":row[16],
                            "SUPERFICIE":row[17],
                            "TIPO_DIA":row[20]
                        })
                    return JsonResponse({'Message': 'Success', 'Datos': listado_data})
                return JsonResponse({'Message': 'Error', 'Nota': 'No se encontraron datos.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})


# LEGAJO	NOMBRES	FECHA	QR	ID_CUADRO	ID_CHACRA	ID_PRODUCTOR	PRODUCTOR	CHACRA	CUADRO	FILA	VARIEDADES	CANT_PLANTAS	LABOR	IMPORTE_FILA	ID_QR_FILA
# 54009	URDANETA ALVAREZ SENEN ALBERTO	14/06/2025	5107	137	1001025	5405	TRES ASES S.A.	Z	1	4	RED DEL CHAÑAR	86	PODA	25000.00	5107
# 58015	CHAMBI JOSUE RUBEN	14/06/2025	5107	137	1001025	5405	TRES ASES S.A.	Z	1	4	RED DEL CHAÑAR	86	PODA	25000.00	5107

def convert_json_labores(json_data,tipo):
    try:
        if tipo == 'L':
            result = {}
            #print(json_data)
            for item in json_data:
                nombre = item['NOMBRES']
                legajo = item['LEGAJO']
                importe = float(item['IMPORTE'])
                detalle = {
                    'LEGAJO': str(legajo),  
                    'FECHA': item['FECHA'],
                    'TIPO_DIA': item['TIPO_DIA'],
                    'PRODUCTOR': item['PRODUCTOR'],
                    'CHACRA': item['CHACRA'],
                    'CUADRO': item['CUADRO'],
                    'FILA': item['FILA'],
                    'QR': item['QR'],
                    'LABOR': item['LABOR'],
                    'IMPORTE': formatear_moneda(importe),
                    'PRESUPUESTO': formatear_moneda(item['PRESUPUESTO']),
                    'VARIEDADES': item['VARIEDADES'],
                    'PLANTAS': item['CANT_PLANTAS'],
                    'SUPERFICIE': item['SUPERFICIE']
                }
                if nombre in result:
                    result[nombre]['IMPORTE_TOTAL'] += importe
                    result[nombre]['DETALLES'].append(detalle)
                else:
                    result[nombre] = {
                        'NOMBRES': nombre,
                        'IMPORTE_TOTAL': importe,
                        'DETALLES': [detalle]
                    }
            result_json = sorted(list(result.values()), key=lambda x: x['NOMBRES'])
            return result_json
        elif tipo == 'C':
            result = {}
            for item in json_data:
                chacra = item['CHACRA']
                importe = float(item['IMPORTE'])
                detalle = {
                    'LEGAJO': str(item['LEGAJO']),
                    'NOMBRES': item['NOMBRES'],
                    'FECHA': item['FECHA'],
                    'PRODUCTOR': item['PRODUCTOR'],
                    'CUADRO': item['CUADRO'],
                    'FILA': item['FILA'],
                    'QR': item['QR'],
                    'LABOR': item['LABOR'],
                    'IMPORTE': importe,
                    'PRESUPUESTO': item['PRESUPUESTO'],
                    'VARIEDADES': item['VARIEDADES'],
                    'PLANTAS': item['CANT_PLANTAS'],
                    'SUPERFICIE': item['SUPERFICIE']
                }
                if chacra in result:
                    result[chacra]['IMPORTE_TOTAL'] += importe
                    result[chacra]['DETALLES'].append(detalle)
                else:
                    result[chacra] = {
                        'CHACRA': chacra,
                        'IMPORTE_TOTAL': importe,
                        'DETALLES': [detalle]
                    }
            for chacra in result:
                result[chacra]['DETALLES'] = sorted(result[chacra]['DETALLES'], key=lambda x: x['NOMBRES'])

            result_json = list(result.values())
            return result_json
    except Exception as e:
        return 'e'

@csrf_exempt
def archivo_detalle_labores(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idLegajo = str(request.POST.get('IdLegajo'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            idEncargado = str(request.POST.get('IdEncargado'))
            idLabor = str(request.POST.get('IdLabor'))
            archivo = str(request.POST.get('Archivo'))
            tipo = str(request.POST.get('Tipo'))
            values = [inicio,final,idLegajo,idChacra,idEncargado,idLabor,idCuadro]
            filtros = request.POST.get('Filtros')
            filtros_dict = json.loads(filtros)
            if tipo == 'RP':
                values = [inicio,final,idLegajo,idChacra,idEncargado]
                nombre_excel = consulta_resumido_persona(values,filtros_dict)
                return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel})
            with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ EXEC SP_SELECT_DETALLE_LABORES %s, %s, %s, %s, %s, %s, %s  """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "LEGAJO":row[0],
                            "NOMBRES":row[1],
                            "FECHA":row[2],
                            "QR":row[3],
                            "ID_CUADRO":row[4],
                            "ID_CHACRA":row[5],
                            "ID_PRODUCTOR":row[6],
                            "PRODUCTOR":row[7],
                            "CHACRA":row[8],
                            "CUADRO":row[9],
                            "FILA":row[10],
                            "VARIEDADES":row[11],
                            "CANT_PLANTAS":row[12],
                            "LABOR":row[13],
                            "IMPORTE":row[14],
                            "ID_QR_FILA":row[15],
                            "PRESUPUESTO":row[16],
                            "SUPERFICIE":row[17],
                            "SUM_PLANTAS":row[18],
                            "SUM_SUPERFICIE":row[19],
                            "TIPO_DIA":row[20]
                        })
                        suma_plantas = row[18]
                        suma_superficie = row[19]
                    totales = {
                        "IMPORTE": sum(item["IMPORTE"] for item in listado_data),
                        "PRESUPUESTO": sum(item["PRESUPUESTO"] for item in listado_data),
                        "CANT_PLANTAS": sum(item["CANT_PLANTAS"] for item in listado_data),
                        "SUPERFICIE": sum(item["SUPERFICIE"] for item in listado_data)
                    }
                    totales_censo = {
                        "SUMA_PLANTAS":suma_plantas,
                        "SUMA_SUPERFICIE":suma_superficie,
                        "PORCEN_PLANTAS":round(float((totales["CANT_PLANTAS"]/suma_plantas)*100),2),
                        "PORCEN_SUPERFICIE":round(float((totales["SUPERFICIE"]/suma_superficie)*100),2) 
                    }
                    if archivo == 'excel':
                        nombre_excel = crear_excel_labores(listado_data,tipo,filtros_dict,totales,totales_censo)
                        return JsonResponse({'Message': 'Success', 'Archivo': nombre_excel})
                    if archivo == 'pdf':
                        return JsonResponse({'Message': 'Error', 'Nota': 'Este tipo de Archivo aún no esta disponible.'})
                return JsonResponse({'Message': 'Error', 'Nota': 'No se encontraron datos para generar el archivo.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def crear_excel_labores(jsonData,tipo,filtros,totales,total_censo):
    if tipo == 'DC':
        lista_data = convert_json_labores(jsonData,"C")
        try:
            df = pd.json_normalize(lista_data, 'DETALLES', ['CHACRA'])
            output = BytesIO()
            columns1 = ['CHACRA', 'NOMBRES', 'LEGAJO', 'FECHA', 'PRODUCTOR', 'CUADRO', 'FILA', 'QR', 'LABOR', 'IMPORTE', 'VARIEDADES', 'PLANTAS']
            df = df[columns1]
            #df.fillna('', inplace=True)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=5, sheet_name='RESUMIDO POR LEGAJO', header=columns1)
                worksheet = writer.sheets['RESUMIDO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'G2')
                worksheet['C4'] = 'DETALLE LABORES RESUMIDO POR LEGAJO'
                worksheet['C4'].font = Font(size=14, bold=True)
                worksheet['C4'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                for cell in worksheet[6]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=7, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[6:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width

                # for cell in worksheet['J']:
                #    if cell.row > 6 and isinstance(cell.value, (int, float)):
                #       cell.number_format = '#.##0,0'
                    
            output.seek(0)
            nombre_excel = f'Listado_Labores_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Chacras/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'
    if tipo == 'DP':
        lista_data = convert_json_labores(jsonData,"L")
        try:
            df = pd.json_normalize(lista_data, 'DETALLES', ['NOMBRES'])
            output = BytesIO()
            columns1 = ['NOMBRES', 'LEGAJO', 'CHACRA', 'FECHA', 'TIPO_DIA', 'PRODUCTOR', 'CUADRO', 'FILA', 'QR', 'LABOR', 'IMPORTE', 'PRESUPUESTO', 'VARIEDADES', 'PLANTAS', 'SUPERFICIE']
            df = df[columns1]
            #df.fillna('', inplace=True)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='DETALLADO POR LEGAJO', header=columns1)
                worksheet = writer.sheets['DETALLADO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'J2')
                worksheet['E5'] = 'LABORES POR LEGAJO'
                worksheet['E5'].font = Font(size=14, bold=True)
                worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros['DESDE']
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros['HASTA']
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A4'] = 'Productor:'
                worksheet['B4'] = filtros['PRODUCTOR']
                worksheet['A4'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros['CHACRA']
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A6'] = 'Cuadro:'
                worksheet['B6'] = filtros['CUADRO']
                worksheet['A6'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Personal:'
                worksheet['B7'] = filtros['PERSONAL']
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Labor:'
                worksheet['B8'] = filtros['LABOR']
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width
                    
                worksheet.cell(row=worksheet.max_row + 1, column=9).value = "TOTALES:"
                worksheet.cell(row=worksheet.max_row, column=10).value = formatear_moneda(totales["IMPORTE"])
                worksheet.cell(row=worksheet.max_row, column=11).value = formatear_moneda(totales["PRESUPUESTO"])
                worksheet.cell(row=worksheet.max_row, column=13).value = totales["CANT_PLANTAS"]
                worksheet.cell(row=worksheet.max_row, column=14).value = totales["SUPERFICIE"]
                
                for i in range(9, worksheet.max_column + 1):
                    cell = worksheet.cell(row=worksheet.max_row, column=i)
                    cell.font = Font(bold=True)
                    cell.border = border_style
                
                # Obtén la fila actual
                fila_actual = worksheet.max_row + 2

                # Agrega la primera fila de información
                worksheet.cell(row=fila_actual, column=12).value = "TOTAL CENSO:"
                worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                worksheet.cell(row=fila_actual, column=13).value = total_censo["SUMA_PLANTAS"]
                worksheet.cell(row=fila_actual, column=14).value = total_censo["SUMA_SUPERFICIE"]

                # Agrega la segunda fila de información
                fila_actual += 1
                worksheet.cell(row=fila_actual, column=12).value = "% REALIZADO:"
                worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                worksheet.cell(row=fila_actual, column=13).value = str(total_censo["PORCEN_PLANTAS"]) + ' %'
                worksheet.cell(row=fila_actual, column=14).value = str(total_censo["PORCEN_SUPERFICIE"]) + ' %'

                # Aplica borde a las celdas
                for i in range(12, 15):
                    for j in range(fila_actual - 1, fila_actual + 1):
                        cell = worksheet.cell(row=j, column=i)
                        cell.border = border_style
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    for cell in row:
                        if cell.column_letter in ['J', 'K', 'M', 'N']:
                            cell.alignment = Alignment(horizontal='right')
                    
            output.seek(0)
            nombre_excel = f'Listado_Labores_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Chacras/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'
    
def crear_excel_resumido(tipo,lista_data,filtros):
    if tipo == 'RP':
        try:
            df = pd.DataFrame(lista_data)
            output = BytesIO()
            columns1 = ['NOMBRES', 'LEGAJO', 'CANT_PLANTAS', 'LABOR', 'IMPORTE', 'VALOR_REFERENCIA', 'DIAS_TRABAJADOS', 'IMPORTE_DIA']
            df = df[columns1]
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='RESUMIDO POR LEGAJO')
                worksheet = writer.sheets['RESUMIDO POR LEGAJO']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'J2')
                worksheet['E5'] = 'RESUMIDO POR LEGAJO'
                worksheet['E5'].font = Font(size=14, bold=True)
                worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros['DESDE']
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros['HASTA']
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A4'] = 'Productor:'
                worksheet['B4'] = filtros['PRODUCTOR']
                worksheet['A4'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros['CHACRA']
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Personal:'
                worksheet['B7'] = filtros['PERSONAL']
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Labor:'
                worksheet['B8'] = filtros['LABOR']
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width
                    
                # worksheet.cell(row=worksheet.max_row + 1, column=9).value = "TOTALES:"
                # worksheet.cell(row=worksheet.max_row, column=10).value = formatear_moneda(totales["IMPORTE"])
                # worksheet.cell(row=worksheet.max_row, column=11).value = formatear_moneda(totales["PRESUPUESTO"])
                # worksheet.cell(row=worksheet.max_row, column=13).value = totales["CANT_PLANTAS"]
                # worksheet.cell(row=worksheet.max_row, column=14).value = totales["SUPERFICIE"]
                
                # for i in range(9, worksheet.max_column + 1):
                #     cell = worksheet.cell(row=worksheet.max_row, column=i)
                #     cell.font = Font(bold=True)
                #     cell.border = border_style
                
                # Obtén la fila actual
                # fila_actual = worksheet.max_row + 2

                # # Agrega la primera fila de información
                # worksheet.cell(row=fila_actual, column=12).value = "TOTAL CENSO:"
                # worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                # worksheet.cell(row=fila_actual, column=13).value = total_censo["SUMA_PLANTAS"]
                # worksheet.cell(row=fila_actual, column=14).value = total_censo["SUMA_SUPERFICIE"]

                # # Agrega la segunda fila de información
                # fila_actual += 1
                # worksheet.cell(row=fila_actual, column=12).value = "% REALIZADO:"
                # worksheet.cell(row=fila_actual, column=12).font = Font(bold=True)
                # worksheet.cell(row=fila_actual, column=13).value = str(total_censo["PORCEN_PLANTAS"]) + ' %'
                # worksheet.cell(row=fila_actual, column=14).value = str(total_censo["PORCEN_SUPERFICIE"]) + ' %'

                # # Aplica borde a las celdas
                # for i in range(12, 15):
                #     for j in range(fila_actual - 1, fila_actual + 1):
                #         cell = worksheet.cell(row=j, column=i)
                #         cell.border = border_style
                # for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                #     for cell in row:
                #         if cell.column_letter in ['J', 'K', 'M', 'N']:
                #             cell.alignment = Alignment(horizontal='right')
                    
            output.seek(0)
            nombre_excel = f'Resumido_Persona_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Chacras/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'

#/home/sides/MAIN S3A/S3A/Applications/Md_Chacras/Archivos/Excel

def consulta_resumido_persona(values,filtros):
    listado_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
            sql = """ 
                EXEC SP_SELECT_RESUMEN_LABORES %s, %s, %s, %s, %s

              """
            cursor.execute(sql, values)
            results = cursor.fetchall()
            if results:
                for row in results:
                    listado_data.append({
                        "NOMBRES":row[0],
                        "LEGAJO":row[1],
                        "CANT_PLANTAS":row[2],
                        "LABOR":row[3],
                        "IMPORTE":row[4],
                        "VALOR_REFERENCIA":row[5],
                        "DIAS_TRABAJADOS":row[6],
                        "IMPORTE_DIA":row[7],
                    })
            nombre_excel = crear_excel_resumido('RP',listado_data,filtros)
        return nombre_excel
    except Exception as e:
        return 'e'


@csrf_exempt
def archivo_detalle_labores_chacras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            # idLabor = str(request.POST.get('IdLabor'))
            idEspecie = str(request.POST.get('IdEspecie'))
            idVariedad = str(request.POST.get('IdVariedad'))
            values = [inicio,final,idChacra,idCuadro,idEspecie,idVariedad]
            jsonDetalleChacras = consulta_detallado_chacra(values)
            if jsonDetalleChacras:
                return JsonResponse({'Message': 'Success', 'Datos': jsonDetalleChacras})
            return JsonResponse({'Message': 'Error', 'Nota': 'No se encontraron datos.'})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

@csrf_exempt
def crea_archivo_detalle_labores_chacras(request):
    if not request.user.is_authenticated:
        return JsonResponse({'Message': 'Not Authenticated', 'Redirect': '/'})
    if request.method == 'POST':
        try:
            listado_data = []
            inicio = str(request.POST.get('Inicio'))
            final = str(request.POST.get('Final'))
            idChacra = str(request.POST.get('IdChacra'))
            idCuadro = str(request.POST.get('IdCuadro'))
            # idLabor = str(request.POST.get('IdLabor'))
            idEspecie = str(request.POST.get('IdEspecie'))
            idVariedad = str(request.POST.get('IdVariedad'))
            tipo = str(request.POST.get('Tipo'))
            values = [inicio,final,idChacra,idCuadro,idEspecie,idVariedad]
            filtros = request.POST.get('Filtros')
            filtros_dict = json.loads(filtros)
            if tipo == 'DP':
                jsonData = consulta_detallado_chacra(values)
                nombre_archivo = crear_excel_labores_chacra(jsonData,tipo,filtros_dict,'','')
                return JsonResponse({'Message': 'Success', 'Archivo': nombre_archivo})
            else:
                jsonData, jsonGeneral = consulta_resumido_chacra(values)
                nombre_archivo = crear_excel_labores_chacra(jsonData,tipo,filtros_dict,jsonGeneral,'')
                return JsonResponse({'Message': 'Success', 'Archivo': nombre_archivo})
        except Exception as e:
            return JsonResponse({'Message': 'Error', 'Nota': str(e)})
    return JsonResponse({'Message': 'No se pudo resolver la petición.'})

def consulta_detallado_chacra(values):
    listado_data = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ 
                    DECLARE @Inicio DATE;
                    DECLARE @Final DATE;
                    DECLARE @IdChacra VARCHAR(10);
                    DECLARE @IdCuadro VARCHAR(10);
                    DECLARE @IdEspecie VARCHAR(10);
                    DECLARE @IdVariedad VARCHAR(10);

                    SET @Inicio = %s;
                    SET @Final = %s;
                    SET @IdChacra = %s;
                    SET @IdCuadro = %s;
                    SET @IdEspecie = %s;
                    SET @IdVariedad = %s;

                    SELECT      CD.ID_CHACRA, RTRIM(TA_CH.Nombre) AS NOM_CHACRA, CONVERT(VARCHAR(10), ST.FECHA, 103) AS FECHA, QR.ID_CUADRO, CD.NOMBRE_CUADRO AS CUADRO, LB.NOMBRE_LABOR,
                                (SELECT STRING_AGG(RTRIM(V.Nombre), ', ') AS Resultado
                                FROM    SPA_FILAS AS S INNER JOIN
                                            S3A.dbo.Variedad AS V ON V.IdVariedad = S.ID_VARIEDAD
                                WHERE (S.ID_CUADRO = CD.ID_CUADRO) AND (S.ID_FILA = FL.ID_FILA)) AS VARIEDADES,
                                (SELECT SUM(NRO_PLANTAS) AS Expr1
                                FROM    SPA_FILAS AS S
                                WHERE (ID_CUADRO = CD.ID_CUADRO) AND (ID_FILA = FL.ID_FILA)) AS CANT_PLANTAS, ST.VALOR AS IMPORTE_FILA, CASE WHEN PR.VALOR_REFERENCIA IS NULL THEN 0 ELSE PR.VALOR_REFERENCIA END AS PRESUPUESTO,
                                CASE 
                                    WHEN PR.VALOR_REFERENCIA IS NULL OR PR.VALOR_REFERENCIA = 0 THEN 0 
                                    ELSE FORMAT(((ST.VALOR - PR.VALOR_REFERENCIA) / PR.VALOR_REFERENCIA) * 100, 'N2') 
                                END AS PORCENTAJE_SUPERADO,
                                (SELECT (SUM(SUPERFICIE) * SUM(NRO_PLANTAS)) FROM SPA_FILAS AS S WHERE ID_CUADRO = CD.ID_CUADRO AND ID_FILA = FL.ID_FILA) AS SUP_TRABAJADA, FL.ID_FILA AS ID_FILA,
                                (SELECT E.IdEspecie FROM S3A.dbo.Variedad AS E WHERE E.IdVariedad = FL.ID_VARIEDAD) AS ID_ESPECIE
                    FROM        SPA_TAREA AS ST INNER JOIN
                                SPA_QR AS QR ON ST.ID_QR_FILA = QR.ID_QR INNER JOIN
                                SPA_CUADRO AS CD ON QR.ID_CUADRO = CD.ID_CUADRO INNER JOIN
                                SPA_FILAS AS FL ON QR.ID_CUADRO = FL.ID_CUADRO AND QR.ID_FILA = FL.ID_FILA AND QR.ID_VARIEDAD = FL.ID_VARIEDAD INNER JOIN
                                S3A.dbo.Chacra AS TA_CH ON TA_CH.IdChacra = CD.ID_CHACRA INNER JOIN
                                SPA_LABOR AS LB ON ST.ID_LABOR = LB.ID_LABOR LEFT JOIN
                                SPA_PRESUPUESTO AS PR ON QR.ID_FILA = PR.ID_FILA AND QR.ID_CUADRO = PR.ID_CUADRO AND QR.ID_VARIEDAD = PR.ID_VARIEDAD INNER JOIN
                                S3A.dbo.Variedad AS V ON V.IdVariedad = FL.ID_VARIEDAD
                    WHERE	    CONVERT(DATE, ST.FECHA) >= @Inicio
                                AND CONVERT(DATE, ST.FECHA) <= @Final
                                AND (@IdChacra = CD.ID_CHACRA OR @IdChacra = '')
                                AND (@IdCuadro = CD.ID_CUADRO OR @IdCuadro = '')
                                AND (@IdEspecie = CONVERT(VARCHAR,V.IdEspecie) OR @IdEspecie = '')
                                AND (@IdVariedad = FL.ID_VARIEDAD OR @IdVariedad = '')
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "ID_CHACRA":row[0],
                            "CHACRA":row[1],
                            "FECHA":row[2],
                            "ID_CUADRO":row[3],
                            "CUADRO":row[4],
                            "LABOR":row[5],
                            "VARIEDADES":row[6],
                            "CANT_PLANTAS":row[7],
                            "IMPORTE_FILA":formatear_moneda(str(row[8])),
                            "PRESUPUESTO":formatear_moneda(str(row[9])),
                            "PORCENTAJE":row[10],
                            "SUPERFICIE":row[11],
                            "FILA":row[12],
                        })
        return listado_data
    except Exception as e:
        return listado_data
    
def consulta_resumido_chacra(values):
    listado_data = []
    listado_general = []
    try:
        with connections['TRESASES_APLICATIVO'].cursor() as cursor:
                sql = """ 
                    DECLARE @Inicio DATE;
                    DECLARE @Final DATE;
                    DECLARE @IdChacra VARCHAR(10);
                    DECLARE @IdCuadro VARCHAR(10);
                    DECLARE @IdEspecie VARCHAR(10);
                    DECLARE @IdVariedad VARCHAR(10);

                    SET @Inicio = %s;
                    SET @Final = %s;
                    SET @IdChacra = %s;    ---- CHACRA 1000026 CUADRO 106 CANT FILAS 9 VALOR 360.000 SUPERFICIE 0.405 CANT PLANTAS 504
                    SET @IdCuadro = %s;
                    SET @IdEspecie = %s;
                    SET @IdVariedad = %s;

                    SELECT RTRIM(CH.Nombre) AS NOM_CHACRA, COUNT(FL.ID_FILA) AS FILA, SUM(FL.NRO_PLANTAS) AS CANT_PLANTAS, LB.NOMBRE_LABOR AS LABOR, SUM(ST.VALOR) AS IMPORTE, 
                            CASE WHEN PRE.VALOR_REFERENCIA IS NULL THEN 0 ELSE SUM(PRE.VALOR_REFERENCIA) END AS PRESUPUESTO,
                            CASE 
                                WHEN PRE.VALOR_REFERENCIA IS NULL OR SUM(PRE.VALOR_REFERENCIA) = 0 
                                THEN 0 
                                ELSE FORMAT(((SUM(ST.VALOR) - SUM(PRE.VALOR_REFERENCIA)) / SUM(PRE.VALOR_REFERENCIA)) * 100, 'N2')
                            END AS PORCENTAJE_SUPERADO, SUM(FL.SUPERFICIE) AS SUPERFICIE,
                            (SELECT SUM(SUPERFICIE) FROM SPA_FILAS WHERE ID_CUADRO IN (SELECT ID_CUADRO FROM SPA_CUADRO WHERE ID_CHACRA = CD.ID_CHACRA)) AS SUP_TOTAL,
                            CASE 
                                WHEN (SELECT SUM(SUPERFICIE) FROM SPA_FILAS WHERE ID_CUADRO IN (SELECT ID_CUADRO FROM SPA_CUADRO WHERE ID_CHACRA = CD.ID_CHACRA)) = 0 
                                THEN 0 
                                ELSE ((SUM(FL.SUPERFICIE) * 1.0 / (SELECT SUM(SUPERFICIE) FROM SPA_FILAS WHERE ID_CUADRO IN (SELECT ID_CUADRO FROM SPA_CUADRO WHERE ID_CHACRA = CD.ID_CHACRA))) * 100)
                            END AS PORCENTAJE_TRABAJADO
                    FROM   SPA_TAREA AS ST INNER JOIN
                                SPA_QR AS QR ON ST.ID_QR_FILA = QR.ID_QR INNER JOIN
                                SPA_FILAS AS FL ON QR.ID_FILA = FL.ID_FILA AND QR.ID_CUADRO = FL.ID_CUADRO AND QR.ID_VARIEDAD = FL.ID_VARIEDAD INNER JOIN
                                SPA_CUADRO AS CD ON CD.ID_CUADRO = FL.ID_CUADRO INNER JOIN
                                S3A.dbo.Chacra AS CH ON CH.IdChacra = CD.ID_CHACRA INNER JOIN
                                SPA_LABOR AS LB ON LB.ID_LABOR = ST.ID_LABOR LEFT JOIN
                                SPA_PRESUPUESTO AS PRE ON FL.ID_FILA = PRE.ID_FILA AND FL.ID_CUADRO = PRE.ID_CUADRO AND FL.ID_VARIEDAD = PRE.ID_VARIEDAD INNER JOIN
                                S3A.dbo.Variedad AS V ON V.IdVariedad = FL.ID_VARIEDAD
                    WHERE	CONVERT(DATE, ST.FECHA) >= @Inicio
                            AND CONVERT(DATE, ST.FECHA) <= @Final
                            AND (@IdChacra = CD.ID_CHACRA OR @IdChacra = '')
                            AND (@IdCuadro = CD.ID_CUADRO OR @IdCuadro = '')
                            AND (@IdEspecie = CONVERT(VARCHAR,V.IdEspecie) OR @IdEspecie = '')
                            AND (@IdVariedad = FL.ID_VARIEDAD OR @IdVariedad = '')
                    GROUP BY RTRIM(CH.Nombre),LB.NOMBRE_LABOR,PRE.VALOR_REFERENCIA,CD.ID_CHACRA
                    ORDER BY RTRIM(CH.Nombre)
                    """
                cursor.execute(sql, values)
                consulta = cursor.fetchall()
                if consulta:
                    for row in consulta:
                        listado_data.append({
                            "CHACRA":row[0],
                            "FILAS":row[1],
                            "PLANTAS":row[2],
                            "LABOR":row[3],
                            "IMPORTE":formatear_moneda(str(row[4])),
                            "PRESUPUESTO":formatear_moneda(str(row[5])),
                            "PORCENTAJE_SUPERADO":row[6],
                            "SUPERFICIE":row[7],
                        })
                        listado_general.append({
                            "CHACRA":row[0],
                            "TOTAL_SUPERFICIE":round(row[8],2),
                            "PORCENTAJE_TRABAJADO":round(row[9],2),
                        })
        return listado_data,listado_general
    except Exception as e:
        return listado_data,listado_general

def crear_excel_labores_chacra(jsonData,tipo,filtros,listado_general,total_censo):
    if tipo == 'RP':
        try:
            df_detalles = pd.DataFrame(jsonData)
            df_general = pd.DataFrame(listado_general)
            df_detalles = df_detalles.rename(columns={
                'PORCENTAJE_SUPERADO': '% SUPERADO',
            })
            df_general = df_general.rename(columns={
                'TOTAL_SUPERFICIE': 'TOTAL SUP.',
                'PORCENTAJE_TRABAJADO': '% TRABAJADO',
            })

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_detalles.to_excel(writer, index=False, startrow=9, sheet_name='RESUMIDO POR CHACRA')
                worksheet = writer.sheets['RESUMIDO POR CHACRA']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'G2')
                worksheet['D4'] = 'RESUMIDO POR CHACRA'
                worksheet['D4'].font = Font(size=14, bold=True)
                worksheet['D4'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros.get('DESDE', '')
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros.get('HASTA', '')
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros.get('CHACRA', '')
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A6'] = 'Cuadro:'
                worksheet['B6'] = filtros.get('CUADRO', '')
                worksheet['A6'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Especie:'
                worksheet['B7'] = filtros.get('ESPECIE', '')
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Variedad:'
                worksheet['B8'] = filtros.get('VARIEDAD', '')
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in worksheet.iter_rows(min_row=10, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[6:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width

                # Alinear columnas 2, 3, 5, 6, 7 y 8 de df_detalles a la derecha
                for row in worksheet.iter_rows(min_row=11, max_row=worksheet.max_row, min_col=2, max_col=3):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right')
                for row in worksheet.iter_rows(min_row=11, max_row=worksheet.max_row, min_col=5, max_col=8):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right')

                # Agregar tabla general
                start_row = worksheet.max_row + 3
                df_general.to_excel(writer, index=False, startrow=start_row - 1, startcol=5, sheet_name='RESUMIDO POR CHACRA', header=True)
                for col_idx, col in enumerate(df_general.columns, start=6):  # 6 porque startcol=5 y se suma 1 para la columna correcta
                    cell = worksheet.cell(row=start_row, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                for row in worksheet.iter_rows(min_row=start_row + 1, min_col=6, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                # Alinear columnas 7 y 8 de df_general a la derecha
                start_row_general = worksheet.max_row - len(df_general) + 1
                for row in worksheet.iter_rows(min_row=start_row_general, max_row=worksheet.max_row, min_col=7, max_col=8):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right')

            output.seek(0)
            nombre_excel = f'Listado_Resumido_Chacra_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Chacras/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'
    if tipo == 'DP':
        try:
            df = pd.DataFrame(jsonData)
            columns1 = ['CHACRA', 'FECHA', 'CUADRO', 'FILA', 'CANT_PLANTAS', 'LABOR', 'VARIEDADES', 'IMPORTE_FILA', 'PRESUPUESTO', 'PORCENTAJE', 'SUPERFICIE']
            df = df[columns1]
            df = df.rename(columns={
                'CANT_PLANTAS': 'N° PTAS',
                'IMPORTE_FILA': 'IMPORTE',
                'PORCENTAJE': '%',
            })
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, startrow=9, sheet_name='DETALLADO POR CHACRA')
                worksheet = writer.sheets['DETALLADO POR CHACRA']
                logo = Image('static/3A/images/TA.png')  
                logo.width = 80
                logo.height = 50
                worksheet.add_image(logo, 'J2')
                worksheet['E5'] = 'DETALLADO POR CHACRA'
                worksheet['E5'].font = Font(size=14, bold=True)
                worksheet['E5'].alignment = Alignment(horizontal='center', vertical='center')
                fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                worksheet['E2'] = fecha_actual
                worksheet['E2'].alignment = Alignment(horizontal='right', vertical='center')
                header_fill = PatternFill(start_color="44546a", end_color="44546a", fill_type="solid")
                header_font = Font(color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')

                worksheet['A2'] = 'Desde:'
                worksheet['B2'] = filtros.get('DESDE', '')
                worksheet['A2'].font = Font(size=12, bold=True)

                worksheet['A3'] = 'Hasta:'
                worksheet['B3'] = filtros.get('HASTA', '')
                worksheet['A3'].font = Font(size=12, bold=True)

                worksheet['A5'] = 'Chacra:'
                worksheet['B5'] = filtros.get('CHACRA', '')
                worksheet['A5'].font = Font(size=12, bold=True)

                worksheet['A6'] = 'Cuadro:'
                worksheet['B6'] = filtros.get('CUADRO', '')
                worksheet['A6'].font = Font(size=12, bold=True)

                worksheet['A7'] = 'Especie:'
                worksheet['B7'] = filtros.get('ESPECIE', '')
                worksheet['A7'].font = Font(size=12, bold=True)

                worksheet['A8'] = 'Variedad:'
                worksheet['B8'] = filtros.get('VARIEDAD', '')
                worksheet['A8'].font = Font(size=12, bold=True)

                for cell in worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=10, max_row=worksheet.max_row, min_col=2, max_col=5):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')

                for row in worksheet.iter_rows(min_row=10, max_row=worksheet.max_row, min_col=8, max_col=11):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right')

                for row in worksheet.iter_rows(min_row=11, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = border_style

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col[10:]:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                                pass
                    adjusted_width = (max_length + 8)
                    worksheet.column_dimensions[column].width = adjusted_width

            output.seek(0)
            nombre_excel = f'Listado_Detallado_Chacra_{str(datetime.now().strftime("%d_%m_%Y_%H_%M_%S"))}.xlsx'
            with open('Applications/Md_Chacras/Archivos/Excel/'+nombre_excel, 'wb') as f:
                f.write(output.getvalue())

            return nombre_excel
        except Exception as e:
            return 'e'































